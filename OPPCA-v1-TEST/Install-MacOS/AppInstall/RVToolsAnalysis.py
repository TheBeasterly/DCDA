import os
import sys
import openpyxl as xl
import pandas as pd
import tkinter as tk
from tkinter import filedialog

def main(label2):

################# DEFS #################

    ## Remove Formatting
    def removeFormatting(destfile):
      """Removes all formatting from an excel workbook.

      Args:
        destfile: An xl workbook object.
      """

      for worksheet in destfile.worksheets:
        for cell in worksheet.iter_rows():
          for c in cell:
            c.style = 'Normal'

    ## Match File Servers
    def match_fs(destfile):
      """Matches file servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      fs_str = ["file", "fs", "nas", "share", "ftp"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in fs_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 1).value = "Yes"

    ## Match SQL DBs
    def match_sql(destfile):
      """Matches SQL DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      sql_str = ["sql"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in sql_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 2).value = "Yes"

    ## Match Oracle DBs
    def match_orcl(destfile):
      """Matches Oracle DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      orcl_str = ["orcl", "oracle"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in orcl_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 3).value = "Yes"

    ## Match PostGres DBs
    def match_pgres(destfile):
      """Matches PostGres DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      pgres_str = ["pgres", "postgres"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in pgres_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 4).value = "Yes"

    ## Match Possible DBs
    def match_gendb(destfile):
      """Matches Possible DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      gendb_str = ["db", "database"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in gendb_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 2).value = "Check"
            cell.offset(0, 3).value = "Check"
            cell.offset(0, 4).value = "Check"

    ## Match Exchange Servers
    def match_exch(destfile):
      """Matches Exchange Servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      exch_str = ["exch", "exchange"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in exch_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 5).value = "Yes"

    ## Match TestDev
    def match_tstdev(destfile):
      """Matches TestDev in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      tstdev_str = ["tst", "test", "dev"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in tstdev_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 6).value = "Yes"

    ## Set No Values
    # Set Used Range of Column A
    def get_used_range(worksheet):
      """Gets the used range of a worksheet.

      Args:
        worksheet: An xl worksheet object.

      Returns:
        A tuple of two cell objects, representing the start and end of the used range of the worksheet.
      """

      # Get the first and last row numbers in the worksheet.
      first_row_number = worksheet.min_row
      last_row_number = worksheet.max_row

      # Iterate over the cells in column A and find the first non-empty cell.
      start_cell = None
      for row_index in range(first_row_number, last_row_number + 1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          start_cell = cell
          break

      # If the first non-empty cell in column A is None, then the used range of the worksheet is empty.
      if start_cell is None:
        return None

      # Find the last non-empty cell in column A.
      end_cell = None
      for row_index in range(last_row_number, first_row_number - 1, -1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          end_cell = cell
          break

      # Return the start and end cells of the used range of the worksheet.
      return start_cell, end_cell

    #Set No Values in columns B through G
    def set_no_values(destfile, used_range_a):
      """Sets the value of all empty cells in columns B to G to "No", based on the used range of column A.

      Args:
        destfile: An xl workbook object.
        used_range_a: A tuple of two cell objects, representing the start and end of the used range of column A.
      """

      # Get the start and end row numbers of the used range of column A.
      start_row_number = used_range_a[0].row
      end_row_number = used_range_a[1].row

      # Iterate over the rows in the used range of column A.
      for row_index in range(start_row_number, end_row_number + 1):

        # Iterate over the cells in columns B to G.
        for column_index in range(2, 8):

          # Get the cell object.
          cell = destfile.active.cell(row=row_index, column=column_index)

          # If the cell value is empty, then set it to "No".
          if cell.value is None:
            cell.value = "No"


    ## Insert Columns for MiB to GB Math
    # Find "Provisioned MiB" on vInfo Sheet
    def vinfo_findprovmib(vinfo_ws):
      """Finds the cell in row A of the "vInfo" worksheet that contains the string "Provisioned MiB".

      Args:
        vinfo_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Provisioned MiB", or None if the string is not found.
      """

      for row in vinfo_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Provisioned MiB":
            return cell

      return None

    # Insert Column after "Provisioned MiB" on vInfo Sheet
    def vinfo_provmib_inscol(vinfo_ws, vinfoprovmib_cell):
      """Inserts a new column to the right of the specified cell in the "vInfo" worksheet.

      Args:
        vinfo_ws: An xl worksheet object.
        vinfoprovmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vinfoprovmib_cell.column
      vinfo_ws.insert_cols(column_index + 1)
      new_column_index = vinfoprovmib_cell.column + 1

      # Set the header of the newly added column
      vinfo_ws.cell(row=1, column=new_column_index).value = "Provisioned GB"

      # Calculate and insert "Provisioned GB" values
      for i in range(2, vinfo_ws.max_row + 1):
        prov_mib_value = vinfo_ws.cell(row=i, column=column_index).value

         # Handle cases where the 'Provisioned MiB' cell is empty or not numeric
        if prov_mib_value is None or not isinstance(prov_mib_value, (int, float)):
            prov_gb_value = None  # Set to None or a default value
        else:
            prov_gb_value = round(prov_mib_value / 953.7, 2)

        vinfo_ws.cell(row=i, column=new_column_index).value = prov_gb_value

    # Find "In Use MiB" on vInfo Sheet
    def vinfo_findinusemib(vinfo_ws):
      """Finds the cell in row A of the "vInfo" worksheet that contains the string "In Use MiB".

      Args:
        vinfo_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "In Use MiB", or None if the string is not found.
      """

      for row in vinfo_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "In Use MiB":
            return cell

      return None

    # Insert Column after "In Use MiB" on vInfo Sheet
    def vinfo_inusemib_inscol(vinfo_ws, vinfoinusemib_cell):
      """Inserts a new column to the right of the specified cell in the "vInfo" worksheet.

      Args:
        vinfo_ws: An xl worksheet object.
        vinfoinusemib_cell: The cell to insert the new column to the right of.
      """

      column_index = vinfoinusemib_cell.column
      vinfo_ws.insert_cols(column_index + 1)
      new_column_index = vinfoinusemib_cell.column + 1

      # Set the header of the newly added column
      vinfo_ws.cell(row=1, column=new_column_index).value = "In Use GB"

      # Calculate and insert "In Use GB" values
      for i in range(2, vinfo_ws.max_row + 1):
        inuse_mib_value = vinfo_ws.cell(row=i, column=column_index).value

         # Handle cases where the 'In Use MiB' cell is empty or not numeric
        if inuse_mib_value is None or not isinstance(inuse_mib_value, (int, float)):
            inuse_gb_value = None  # Set to None or a default value
        else:
            inuse_gb_value = round(inuse_mib_value / 953.7, 2)

        vinfo_ws.cell(row=i, column=new_column_index).value = inuse_gb_value

    # Find "Capacity MiB" on vPartition Sheet
    def vpart_findcapacitymib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Capacity MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Capacity MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Capacity MiB":
            return cell

      return None

    # Insert Column after "Capacity MiB" on vPartition Sheet
    def vpart_capacitymib_inscol(vpart_ws, vpartcapmib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartcapmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartcapmib_cell.column
      vpart_ws.insert_cols(column_index + 1)
      new_column_index = vpartcapmib_cell.column + 1

      # Set the header of the newly added column
      vpart_ws.cell(row=1, column=new_column_index).value = "Capacity GB"

      # Calculate and insert "Capacity GB" values
      for i in range(2, vpart_ws.max_row + 1):
        cap_mib_value = vpart_ws.cell(row=i, column=column_index).value

        # Handle cases where the 'Capacity MiB' cell is empty or not numeric
        if cap_mib_value is None or not isinstance(cap_mib_value, (int, float)):
            cap_gb_value = None  # Set to None or a default value
        else:
            cap_gb_value = round(cap_mib_value / 953.7, 2)

        vpart_ws.cell(row=i, column=new_column_index).value = cap_gb_value

    # Find "Consumed MiB" on vPartition Sheet
    def vpart_findconsumedmib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Consumed MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Consumed MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Consumed MiB":
            return cell

      return None

    # Insert Column after "Consumed MiB" on vPartition Sheet
    def vpart_consumedmib_inscol(vpart_ws, vpartconsmib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartconsmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartconsmib_cell.column
      vpart_ws.insert_cols(column_index + 1)
      new_column_index = vpartconsmib_cell.column + 1

      # Set the header of the newly added column
      vpart_ws.cell(row=1, column=new_column_index).value = "Consumed GB"

      # Calculate and insert "Consumed GB" values
      for i in range(2, vpart_ws.max_row + 1):
        cons_mib_value = vpart_ws.cell(row=i, column=column_index).value

        # Handle cases where the 'Consumed MiB' cell is empty or not numeric
        if cons_mib_value is None or not isinstance(cons_mib_value, (int, float)):
            cons_gb_value = None  # Set to None or a default value
        else:
            cons_gb_value = round(cons_mib_value / 953.7, 2)

        vpart_ws.cell(row=i, column=new_column_index).value = cons_gb_value

    # Find "Free MiB" on vPartition Sheet
    def vpart_findfreemib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Free MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Free MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Free MiB":
            return cell

      return None

    # Insert Column after "Free MiB" on vPartition Sheet
    def vpart_freemib_inscol(vpart_ws, vpartfreemib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartfreemib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartfreemib_cell.column
      vpart_ws.insert_cols(column_index + 1)
      new_column_index = vpartfreemib_cell.column + 1

      # Set the header of the newly added column
      vpart_ws.cell(row=1, column=new_column_index).value = "Free GB"

      # Calculate and insert "Free GB" values
      for i in range(2, vpart_ws.max_row + 1):
        free_mib_value = vpart_ws.cell(row=i, column=column_index).value

        # Handle cases where the 'Free MiB' cell is empty or not numeric
        if free_mib_value is None or not isinstance(free_mib_value, (int, float)):
            free_gb_value = None  # Set to None or a default value
        else:
            free_gb_value = round(free_mib_value / 953.7, 2)

        vpart_ws.cell(row=i, column=new_column_index).value = free_gb_value


    ## Compare VMs on vInfo Sheet to vPart Sheet
    def compare_vms(vinfo_ws, vpart_ws):
      """Compares the vInfo worksheet to the vPartition worksheet and sets the "HasTools" column in the vInfo worksheet to "Yes" if a match is found, or "No" if no match is found.

      Args:
        vinfo_ws: An xl Worksheet object for the vInfo worksheet.
        vpart_ws: An xl Worksheet object for the vPartition worksheet.
      """

      # Declare the vinfo_cell variable.
      vinfo_cell = None

      # Iterate over all of the rows in the vInfo worksheet.
      for row in vinfo_ws.rows:

        # Get the cell value in the first column of the row.
        vinfo_cell_value = row[0].value

        # If the cell value is not zero or blank, then check if there is a matching cell in the vPartition worksheet.
        if vinfo_cell_value != 0 and vinfo_cell_value != "":

          # Assign the current row to the vinfo_cell variable.
          vinfo_cell = row[0]

          # Iterate over all of the rows in the vPartition worksheet.
          for vpart_row in vpart_ws.rows:

            # Get the cell value in the first column of the row.
            vpart_cell_value = vpart_row[0].value

            # If a matching cell is found, then set the "HasTools" column in the vInfo worksheet to "Yes".
            if vpart_cell_value == vinfo_cell_value:
              vinfo_cell.offset(0, 7).value = "Yes"
              break

          # If no matching cell is found, then set the "HasTools" column in the vInfo worksheet to "No".
          else:
            vinfo_cell.offset(0, 7).value = "No"

      # Set the header for the "HasTools" column.
      vinfo_ws["H1"].value = "HasTools"


    ## Delete Unnecessary Columns
    def del_cols_vInfo(destfile_name):
      # Load the Excel file into a Pandas DataFrame
      df1 = pd.read_excel(destfile_name, sheet_name='vInfo')
      df2 = pd.read_excel(destfile_name, sheet_name='vPartition')

      # Get a list of all valid column indices
      valid_column_indices1 = list(df1.columns)
      valid_column_indices2 = list(df2.columns)

      # Get the column indices to keep
      keep_cols_vInfo = [col_idx for col_idx in valid_column_indices1 if col_idx in ['VM', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'HasTools', 'Disks', 'Total disk capacity', 'Provisioned MiB', 'Provisioned GB', 'In Use MiB', 'In Use GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']]
      keep_cols_vPart = [col_idx for col_idx in valid_column_indices2 if col_idx in ['VM', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'Disk', 'Capacity MiB', 'Capacity GB', 'Consumed MiB', 'Consumed GB', 'Free MiB', 'Free GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']]

      # Create a new DataFrame with only the columns that you want to keep
      new_df1 = df1.loc[:, keep_cols_vInfo]
      new_df2 = df2.loc[:, keep_cols_vPart]

      # Save the adjusted worksheets to the same workbook
      writer = pd.ExcelWriter(destfile_name, mode='w')
      new_df1.to_excel(writer, sheet_name='vInfo', index=False)
      new_df2.to_excel(writer, sheet_name='vPartition', index=False)
      writer.close()


    ## Filter First Row
    def filter_rows(destfile):
      """Filters the first row on all sheets.

      Args:
        destfile: An xl workbook object.
      """

      for ws in destfile.worksheets:
        # Set the auto-filter range to the first row of each sheet.
        ws.auto_filter.ref = "A1:{}".format(ws.dimensions.split(':')[1])


################# RUNTIME #################


    ## Get the Excel file name from the user
    root = tk.Tk()
    root.withdraw()

    try:
      srcfile_name = filedialog.askopenfilename()

      try:
        ## Open the Excel file
        srcfile = xl.load_workbook(srcfile_name)

        ## Create a new Excel file
        destfile_name = srcfile_name[:-5] + "-EDITED.xlsx"
        srcfile.save(destfile_name)

        ## Open new Excel file and remove extra sheets
        destfile = xl.load_workbook(destfile_name)
        
        keep_sheets = ['vInfo', 'vDisk', 'vPartition']
        for sheetName in destfile.sheetnames:
            if sheetName not in keep_sheets:
                del destfile[sheetName]
        destfile.save(destfile_name)

        ## Remove Formatting
        destfile = xl.load_workbook(destfile_name)
        removeFormatting(destfile)
        destfile.save(destfile_name)

        ## Add Columns
        destfile = xl.load_workbook(destfile_name)

        for worksheet in destfile.worksheets:
          worksheet.insert_cols(2, 6)

        destfile.save(destfile_name)

        ## Rename columns B:G in all sheets to "IsFile", "IsSQL", "IsOrcl", "IsPGres", "IsExch", and "IsTestDev" 
        destfile = xl.load_workbook(destfile_name)

        for worksheet in destfile.worksheets:
          worksheet['B1'] = 'IsFile'
          worksheet['C1'] = 'IsSQL'
          worksheet['D1'] = 'IsOrcl'
          worksheet['E1'] = 'IsPGres'
          worksheet['F1'] = 'IsExch'
          worksheet['G1'] = 'IsTestDev'

        # Insert "HasTools" Column in vInfo Sheet
        destfile["vInfo"].insert_cols(8, 1)
        destfile["vInfo"]['H1'] = 'HasTools'

        destfile.save(destfile_name)

        ## Match Workload Types
        destfile = xl.load_workbook(destfile_name)
        match_fs(destfile)
        match_sql(destfile)
        match_orcl(destfile)
        match_pgres(destfile)
        match_gendb(destfile)
        match_exch(destfile)
        match_tstdev(destfile)
        for sheetname in destfile.sheetnames:
          destfile.active = destfile[sheetname]
          used_range_a = get_used_range(destfile.active)
          set_no_values(destfile, used_range_a)
        destfile.save(destfile_name)

        ## Del Cols
        del_cols_vInfo(destfile_name)

        ## Insert Columns for MiB to GB math
        destfile = xl.load_workbook(destfile_name)
        vinfo_ws = destfile["vInfo"]
        vpart_ws = destfile["vPartition"]
        for sheet in destfile:
            destfile[sheet.title].views.sheetView[0].tabSelected = False
        removeFormatting(destfile)

        # Insert Columns in vInfo Sheet
        destfile.active = vinfo_ws
        vinfoprovmib_cell = vinfo_findprovmib(vinfo_ws)
        if vinfoprovmib_cell is not None:
          vinfo_provmib_inscol(vinfo_ws, vinfoprovmib_cell)
        vinfoinusemib_cell = vinfo_findinusemib(vinfo_ws)
        if vinfoinusemib_cell is not None:
          vinfo_inusemib_inscol(vinfo_ws, vinfoinusemib_cell)

        # Insert Columns in vPartition Sheet
        destfile.active = vpart_ws
        vpartcapmib_cell = vpart_findcapacitymib(vpart_ws)
        if vpartcapmib_cell is not None:
          vpart_capacitymib_inscol(vpart_ws, vpartcapmib_cell)
        vpartconsmib_cell = vpart_findconsumedmib(vpart_ws)
        if vpartconsmib_cell is not None:
          vpart_consumedmib_inscol(vpart_ws, vpartconsmib_cell)
        vpartfreemib_cell = vpart_findfreemib(vpart_ws)
        if vpartfreemib_cell is not None:
          vpart_freemib_inscol(vpart_ws, vpartfreemib_cell)

        destfile.save(destfile_name)

        ## Compare VMs
        destfile = xl.load_workbook(destfile_name)
        vinfo_ws = destfile["vInfo"]
        vpart_ws = destfile["vPartition"]
        compare_vms(vinfo_ws, vpart_ws)
        destfile.save(destfile_name)

        ## Filter First Row
        destfile = xl.load_workbook(destfile_name)
        filter_rows(destfile)
        destfile.save(destfile_name)
      except Exception as e:
        label2.config(text=f"Source File import aborted. Analysis canceled.")
      else:
        label2.config(text=f"Success!\n \nRVTools Analysis Completed\n \nAnalyzed file will be new file appended with -EDITED and saved in same directory as Source File\n \nYou may now click the Exit Button to close this Program, or choose another file to analyze")
    except Exception as e:
      label2.config(text=f"Error selecting file: {str(e)}")


if __name__ == "__main__":
    main()