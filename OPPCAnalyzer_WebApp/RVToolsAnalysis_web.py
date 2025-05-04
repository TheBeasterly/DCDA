# RVToolsAnalysis_web.py
import os
import sys
import openpyxl as xl
from openpyxl.styles import PatternFill # Keep this if removeFormatting needs it
import pandas as pd
import traceback
import time # Added for SSE delays
import json # Added for SSE final message

# =================================================================================================================================================================================================================================================
# ============================================================================================= ALL ORIGINAL HELPER FUNCTIONS FROM RVToolsAnalysis.py =============================================================================================
# =================================================================================================================================================================================================================================================

## Remove Formatting
def removeFormatting(destfile):
  """Removes all formatting from an excel workbook.

  Args:
    destfile: An xl workbook object.
  """
  try:
    for worksheet in destfile.worksheets:
      # Iterate over all cells that have data or formatting
      for row in worksheet.iter_rows():
        for cell in row:
            # Resetting style explicitly might be needed depending on openpyxl version
            # Using 'Normal' style assumes it exists. If not, other methods might be needed.
            # A safer approach might be to reset individual properties if 'Normal' fails.
             if cell.has_style:
                cell.font = xl.styles.Font()
                cell.border = xl.styles.Border()
                cell.fill = xl.styles.PatternFill()
                cell.number_format = 'General' # Reset number format
                cell.protection = xl.styles.Protection()
                cell.alignment = xl.styles.Alignment()
                # Explicitly set style attribute if available and necessary
                # cell.style = 'Normal' # This can sometimes cause issues if 'Normal' isn't defined
  except Exception as e:
    print(f"Warning: Error during removeFormatting: {e}")
    # Continue processing even if formatting removal fails partially

## Match File Servers
def match_fs(destfile):
  """Matches file servers in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  fs_str = ["file", "fs", "nas", "share", "ftp"]
  try:
    for worksheet in destfile.worksheets:
       # Check if column A exists and has values
       if worksheet.max_column < 1: continue # Skip empty sheets

       for cell in worksheet["A"]: # Iterate through cells in column A
          if cell.row == 1: continue # Skip header row
          if cell.value is None or not isinstance(cell.value, str): continue # Skip empty/non-string cells

          cell_value_lower = cell.value.lower()
          match_found = False
          for search_string in fs_str:
            if search_string in cell_value_lower: # Simple substring check
              match_found = True
              break

          if match_found:
            # Assuming 'IsFile' is now in Column C (index 3)
            if worksheet.max_column >= 3:
                worksheet.cell(row=cell.row, column=3).value = "Yes"
            else:
                print(f"Warning: Column C (IsFile) not found in sheet '{worksheet.title}' for row {cell.row}")

  except Exception as e:
      print(f"Error during match_fs: {e}")
      traceback.print_exc()


## Match SQL DBs
def match_sql(destfile):
  """Matches SQL DBs in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  sql_str = ["sql"]
  try:
    for worksheet in destfile.worksheets:
       if worksheet.max_column < 1: continue

       for cell in worksheet["A"]:
          if cell.row == 1: continue
          if cell.value is None or not isinstance(cell.value, str): continue

          cell_value_lower = cell.value.lower()
          match_found = False
          for search_string in sql_str:
            if search_string in cell_value_lower:
              match_found = True
              break

          if match_found:
            # Assuming 'IsSQL' is now in Column D (index 4)
            if worksheet.max_column >= 4:
                worksheet.cell(row=cell.row, column=4).value = "Yes"
            else:
                 print(f"Warning: Column D (IsSQL) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_sql: {e}")
      traceback.print_exc()


## Match Oracle DBs
def match_orcl(destfile):
  """Matches Oracle DBs in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  orcl_str = ["orcl", "oracle"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in orcl_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsOrcl' is now in Column E (index 5)
          if worksheet.max_column >= 5:
              worksheet.cell(row=cell.row, column=5).value = "Yes"
          else:
               print(f"Warning: Column E (IsOrcl) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_orcl: {e}")
      traceback.print_exc()


## Match PostGres DBs
def match_pgres(destfile):
  """Matches PostGres DBs in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  pgres_str = ["pgres", "postgres"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in pgres_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsPGres' is now in Column F (index 6)
          if worksheet.max_column >= 6:
              worksheet.cell(row=cell.row, column=6).value = "Yes"
          else:
              print(f"Warning: Column F (IsPGres) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_pgres: {e}")
      traceback.print_exc()


## Match Possible DBs (General)
def match_gendb(destfile):
  """Matches Possible (generic) DBs in all sheets based on VM name in Column A.
     Sets 'Check' in SQL, Oracle, PGres columns if a generic term is found
     and the specific match wasn't already 'Yes'.

  Args:
    destfile: An xl workbook object.
  """
  gendb_str = ["db", "database"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      # Define columns for DB checks (assuming they are D, E, F)
      db_check_cols = [4, 5, 6] # D=IsSQL, E=IsOrcl, F=IsPGres

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in gendb_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Set 'Check' only if the specific DB columns are not already 'Yes'
          for col_idx in db_check_cols:
             if worksheet.max_column >= col_idx:
                 target_cell = worksheet.cell(row=cell.row, column=col_idx)
                 if target_cell.value != "Yes":
                     target_cell.value = "Check"
             else:
                  print(f"Warning: Column {col_idx} not found in sheet '{worksheet.title}' for row {cell.row} during gendb check.")
  except Exception as e:
      print(f"Error during match_gendb: {e}")
      traceback.print_exc()


## Match Exchange Servers
def match_exch(destfile):
  """Matches Exchange Servers in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  exch_str = ["exch", "exchange"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in exch_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsExch' is now in Column G (index 7)
          if worksheet.max_column >= 7:
              worksheet.cell(row=cell.row, column=7).value = "Yes"
          else:
              print(f"Warning: Column G (IsExch) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_exch: {e}")
      traceback.print_exc()


## Match TestDev
def match_tstdev(destfile):
  """Matches Test/Dev systems in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  tstdev_str = ["tst", "test", "dev"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in tstdev_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsTestDev' is now in Column H (index 8)
          if worksheet.max_column >= 8:
              worksheet.cell(row=cell.row, column=8).value = "Yes"
          else:
              print(f"Warning: Column H (IsTestDev) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_tstdev: {e}")
      traceback.print_exc()


## Set No Values
# Helper to get used range of Column A (more robustly)
def get_last_row_in_col_a(worksheet):
    """Finds the last row index in Column A that contains data."""
    last_row = 0
    # Iterate backwards from max_row is safer if there are gaps
    for row_idx in range(worksheet.max_row, 0, -1):
        cell = worksheet.cell(row=row_idx, column=1) # Column A
        if cell.value is not None and str(cell.value).strip() != "":
            last_row = row_idx
            break
    # If sheet seems empty based on max_row but has data, check dimensions?
    # For simplicity, we rely on max_row being somewhat accurate.
    # Or iterate from 1 up to max_row if needed.
    if last_row == 0 and worksheet.max_row > 0: # Fallback check if max_row is misleading
        for row_idx in range(1, worksheet.max_row + 1):
             cell = worksheet.cell(row=row_idx, column=1)
             if cell.value is not None and str(cell.value).strip() != "":
                 last_row = max(last_row, row_idx)
    return last_row

# Set No Values in columns C through H
def set_no_values(destfile):
    """Sets the value of all empty cells in columns C to H to "No",
       up to the last row containing data in column A for each sheet.

    Args:
      destfile: An xl workbook object.
    """
    try:
        # Iterate through the relevant sheets
        for ws_name in ['vInfo', 'vDisk', 'vPartition']:
            if ws_name in destfile.sheetnames:
                worksheet = destfile[ws_name]
                if worksheet.max_row <= 1: continue # Skip header-only or empty sheets

                last_row_a = get_last_row_in_col_a(worksheet)
                if last_row_a == 0: # If column A seems empty, maybe use worksheet.max_row cautiously
                    last_row_a = worksheet.max_row if worksheet.max_row > 1 else 1
                    print(f"Warning: Could not determine last row in Col A for sheet '{ws_name}'. Using worksheet max_row {last_row_a}.")


                # Iterate over the rows determined by column A's content.
                # Start from row 2 to skip header
                for row_index in range(2, last_row_a + 1):
                    # Iterate over the columns C to H (indices 3 to 8).
                    for column_index in range(3, 9):
                        # Check if column exists before accessing
                        if worksheet.max_column >= column_index:
                            cell = worksheet.cell(row=row_index, column=column_index)
                            # If the cell value is empty (None), set it to "No".
                            if cell.value is None or str(cell.value).strip() == "":
                                cell.value = "No"
                        # else: No need to warn for every cell, assume column exists if header was set
            else:
                 print(f"Warning: Sheet '{ws_name}' not found in workbook for set_no_values.")

    except Exception as e:
        print(f"Error during set_no_values: {e}")
        traceback.print_exc()

## Insert Columns for MiB to GB Math

# --- vInfo Sheet ---
def vinfo_find_col(vinfo_ws, header_name):
    """Finds the column index (1-based) for a given header name in the first row."""
    for col_idx in range(1, vinfo_ws.max_column + 1):
        if vinfo_ws.cell(row=1, column=col_idx).value == header_name:
            return col_idx
    return None

def vinfo_insert_gb_col(vinfo_ws, mib_header, gb_header):
    """Inserts a GB column after the MiB column and calculates values."""
    mib_col_idx = vinfo_find_col(vinfo_ws, mib_header)
    if mib_col_idx is None:
        print(f"Warning: Column '{mib_header}' not found in vInfo sheet.")
        return False # Indicate failure

    gb_col_idx = mib_col_idx + 1
    vinfo_ws.insert_cols(gb_col_idx)
    vinfo_ws.cell(row=1, column=gb_col_idx).value = gb_header

    for i in range(2, vinfo_ws.max_row + 1):
        mib_cell = vinfo_ws.cell(row=i, column=mib_col_idx)
        gb_cell = vinfo_ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = mib_cell.value
            if mib_value is not None and isinstance(mib_value, (int, float)):
                # ********** CHANGED HERE **********
                gb_value = round(mib_value / 953.7, 2)
                # **********************************
                gb_cell.value = gb_value
                gb_cell.number_format = '0.00' # Apply number format
            # else: leave gb_cell blank if MiB is not a number
        except (ValueError, TypeError):
             # Handle potential errors if value is string that can't be converted
             gb_cell.value = None # Or set to 0 or an error string if preferred
    print(f"Processed '{gb_header}' column in vInfo.")
    return True # Indicate success

# --- vDisk Sheet ---
def vdisk_find_col(vdisk_ws, header_name):
    """Finds the column index (1-based) for a given header name in the first row."""
    for col_idx in range(1, vdisk_ws.max_column + 1):
        if vdisk_ws.cell(row=1, column=col_idx).value == header_name:
            return col_idx
    return None

def vdisk_insert_gb_col(vdisk_ws, mib_header, gb_header):
    """Inserts a GB column after the MiB column and calculates values."""
    mib_col_idx = vdisk_find_col(vdisk_ws, mib_header)
    if mib_col_idx is None:
        print(f"Warning: Column '{mib_header}' not found in vDisk sheet.")
        return False

    gb_col_idx = mib_col_idx + 1
    vdisk_ws.insert_cols(gb_col_idx)
    vdisk_ws.cell(row=1, column=gb_col_idx).value = gb_header

    for i in range(2, vdisk_ws.max_row + 1):
        mib_cell = vdisk_ws.cell(row=i, column=mib_col_idx)
        gb_cell = vdisk_ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = mib_cell.value
            if mib_value is not None and isinstance(mib_value, (int, float)):
                # ********** CHANGED HERE **********
                gb_value = round(mib_value / 953.7, 2)
                # **********************************
                gb_cell.value = gb_value
                gb_cell.number_format = '0.00'
            # else: leave gb_cell blank
        except (ValueError, TypeError):
             gb_cell.value = None
    print(f"Processed '{gb_header}' column in vDisk.")
    return True

# --- vPartition Sheet ---
def vpart_find_col(vpart_ws, header_name):
    """Finds the column index (1-based) for a given header name in the first row."""
    for col_idx in range(1, vpart_ws.max_column + 1):
        if vpart_ws.cell(row=1, column=col_idx).value == header_name:
            return col_idx
    return None

def vpart_insert_gb_col(vpart_ws, mib_header, gb_header):
    """Inserts a GB column after the MiB column and calculates values."""
    mib_col_idx = vpart_find_col(vpart_ws, mib_header)
    if mib_col_idx is None:
        print(f"Warning: Column '{mib_header}' not found in vPartition sheet.")
        return False

    gb_col_idx = mib_col_idx + 1
    vpart_ws.insert_cols(gb_col_idx)
    vpart_ws.cell(row=1, column=gb_col_idx).value = gb_header

    for i in range(2, vpart_ws.max_row + 1):
        mib_cell = vpart_ws.cell(row=i, column=mib_col_idx)
        gb_cell = vpart_ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = mib_cell.value
            if mib_value is not None and isinstance(mib_value, (int, float)):
                # ********** CHANGED HERE **********
                gb_value = round(mib_value / 953.7, 2)
                # **********************************
                gb_cell.value = gb_value
                gb_cell.number_format = '0.00'
            # else: leave gb_cell blank
        except (ValueError, TypeError):
             gb_cell.value = None
    print(f"Processed '{gb_header}' column in vPartition.")
    return True

# -- Disk Count Value --
def vdisk_diskcount_val(vdisk_ws):
    """Inserts a value of '1' in the 'DiskCount' Column on the vDisk worksheet.
       Assumes 'DiskCount' header is already present."""
    diskcount_col_idx = vdisk_find_col(vdisk_ws, "DiskCount")
    if diskcount_col_idx is None:
        print("Warning: 'DiskCount' column header not found in vDisk sheet.")
        return False

    # Set the value to 1 for all data rows
    for i in range(2, vdisk_ws.max_row + 1):
        vdisk_ws.cell(row=i, column=diskcount_col_idx).value = 1
    print("Processed 'DiskCount' values in vDisk.")
    return True


## Compare VMs on vInfo Sheet to vPart Sheet
def compare_vms(vinfo_ws, vpart_ws):
  """Compares VMs in vInfo (Col A) to vPartition (Col A) and sets 'HasTools'
     (assumed Col I / index 9 in vInfo) to 'Yes' or 'No'.

  Args:
    vinfo_ws: An xl Worksheet object for the vInfo worksheet.
    vpart_ws: An xl Worksheet object for the vPartition worksheet.
  """
  if vinfo_ws is None or vpart_ws is None:
      print("Error: vInfo or vPartition worksheet not provided for compare_vms.")
      return

  # --- Pre-load vPartition VM names into a set for faster lookup ---
  vpart_vms = set()
  if vpart_ws.max_row > 1: # Check if there's data beyond the header
      # Assuming VM names are in the first column (A)
      for cell in vpart_ws['A'][1:]: # Skip header row A1
          if cell.value is not None and str(cell.value).strip() != "":
              vpart_vms.add(str(cell.value)) # Add VM name as string
  print(f"Found {len(vpart_vms)} unique VM names in vPartition for comparison.")

  # --- Iterate through vInfo and check against the set ---
  hastools_col_idx = 9 # Assuming 'HasTools' is Column I

  # Ensure the HasTools column exists (at least header)
  if vinfo_ws.max_column < hastools_col_idx:
      print(f"Warning: 'HasTools' column (expected index {hastools_col_idx}) not found in vInfo sheet.")
      # Optionally add the column if missing? For now, we assume it was added earlier.
      # vinfo_ws.insert_cols(hastools_col_idx)
      # vinfo_ws.cell(row=1, column=hastools_col_idx).value = "HasTools"
      # return # Or continue if you add it

  # Check header just in case
  if vinfo_ws.cell(row=1, column=hastools_col_idx).value != "HasTools":
       print(f"Warning: Column {hastools_col_idx} in vInfo is not named 'HasTools'. Proceeding anyway.")


  if vinfo_ws.max_row > 1:
      # Iterate through rows in vInfo (Column A for VM name)
      for row_idx in range(2, vinfo_ws.max_row + 1):
          vinfo_vm_cell = vinfo_ws.cell(row=row_idx, column=1) # Col A
          hastools_cell = vinfo_ws.cell(row=row_idx, column=hastools_col_idx) # Col I

          if vinfo_vm_cell.value is not None and str(vinfo_vm_cell.value).strip() != "":
              vm_name = str(vinfo_vm_cell.value)
              # Check if the VM name exists in the set from vPartition
              if vm_name in vpart_vms:
                  hastools_cell.value = "Yes"
              else:
                  hastools_cell.value = "No"
          else:
              # Handle empty VM name cell in vInfo if necessary
              hastools_cell.value = "No" # Or None, or skip


## Delete Unnecessary Columns (Using Pandas)
def del_cols_vInfo(destfile_path):
    """Deletes unnecessary columns using Pandas for efficiency.
       Reads from and writes back to the specified Excel file path.
    """
    try:
        # Define columns to keep for each sheet
        keep_cols = {
            'vInfo': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'HasTools', 'Disks', 'Total disk capacity', 'Provisioned MiB', 'Provisioned GB', 'In Use MiB', 'In Use GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vDisk': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'DiskCount', 'Disk', 'Capacity MiB', 'Capacity GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vPartition': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'Disk', 'Capacity MiB', 'Capacity GB', 'Consumed MiB', 'Consumed GB', 'Free MiB', 'Free GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']
        }

        # Read all relevant sheets into a dictionary of DataFrames
        # Use openpyxl engine to preserve sheet structure better if possible
        try:
             all_sheets_df = pd.read_excel(destfile_path, sheet_name=['vInfo', 'vDisk', 'vPartition'], engine='openpyxl')
        except ValueError as e:
             # Handle case where one sheet might be missing
             print(f"Warning reading sheets with pandas: {e}. Trying to read existing sheets individually.")
             all_sheets_df = {}
             xls = pd.ExcelFile(destfile_path, engine='openpyxl')
             for sheet_name in ['vInfo', 'vDisk', 'vPartition']:
                 if sheet_name in xls.sheet_names:
                     all_sheets_df[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)


        # Create a Pandas Excel writer object using openpyxl engine to overwrite
        # Use mode='w' to overwrite the file with only the processed sheets
        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in all_sheets_df.items():
                if sheet_name in keep_cols:
                    # Find which columns to keep ACTUALLY exist in the DataFrame
                    cols_to_keep_in_df = [col for col in keep_cols[sheet_name] if col in df.columns]
                    if len(cols_to_keep_in_df) < len(keep_cols[sheet_name]):
                         missing_cols = set(keep_cols[sheet_name]) - set(cols_to_keep_in_df)
                         print(f"Warning: Columns missing in '{sheet_name}' for deletion step: {missing_cols}")

                    # Select only the columns to keep
                    new_df = df[cols_to_keep_in_df]
                    # Write the modified DataFrame back to the Excel file
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"Processed columns for sheet: {sheet_name}")
                # else: Sheets not in keep_cols are implicitly dropped by mode='w'

    except FileNotFoundError:
        print(f"Error: File not found for del_cols_vInfo: {destfile_path}")
        raise # Re-raise the exception to be caught by the main handler
    except Exception as e:
        print(f"Error during del_cols_vInfo (Pandas): {e}")
        traceback.print_exc()
        raise # Re-raise the exception


## Truncate vPartition Storage Capacity Cols into vInfo (Using Pandas)
def trunc_cols_vPart(destfile_path):
    """Aggregates storage data from vPartition and merges it into a new vSummary sheet.
       Reads from and writes back to the specified Excel file path.
    """
    try:
        # Read necessary sheets
        # Use openpyxl engine. Handle potential missing sheets.
        sheets_to_read = ['vInfo', 'vPartition']
        dfs = {}
        try:
            xls = pd.ExcelFile(destfile_path, engine='openpyxl')
            for sheet_name in sheets_to_read:
                 if sheet_name in xls.sheet_names:
                     dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                 else:
                     # If a required sheet is missing, we cannot proceed
                     raise ValueError(f"Required sheet '{sheet_name}' not found in '{destfile_path}' for truncation.")
            # Also read vDisk if it exists, to preserve it
            if 'vDisk' in xls.sheet_names:
                 dfs['vDisk'] = pd.read_excel(xls, sheet_name='vDisk')

        except FileNotFoundError:
             print(f"Error: File not found for trunc_cols_vPart: {destfile_path}")
             raise
        except ValueError as e: # Catch the missing sheet error from above
             print(f"Error: {e}")
             raise

        tcdf1 = dfs['vInfo']
        tcdf3 = dfs['vPartition']

        # Identify storage-related columns in vPartition
        storage_columns = ['Capacity GB', 'Consumed GB', 'Free GB']
        # Filter to only those that actually exist in the DataFrame
        existing_storage_columns = [col for col in storage_columns if col in tcdf3.columns]
        if not existing_storage_columns:
             print("Warning: No storage columns ('Capacity GB', 'Consumed GB', 'Free GB') found in vPartition. Skipping aggregation.")
              # Decide how to proceed: maybe create vSummary with just vInfo data?
              # For now, we'll just save existing sheets and skip vSummary creation.
             vSummary_df = tcdf1.copy() # Create vSummary based on vInfo only
             # Fall through to save section

        else:
            print(f"Aggregating vPartition columns: {existing_storage_columns}")
            # Ensure 'VM' column exists for grouping
            if 'VM' not in tcdf3.columns:
                 raise ValueError("'VM' column not found in vPartition sheet, cannot aggregate.")

            # Aggregate storage data from vPartition, converting non-numeric to 0 before sum
            for col in existing_storage_columns:
                 tcdf3[col] = pd.to_numeric(tcdf3[col], errors='coerce').fillna(0)

            aggregated_df = (
                tcdf3.groupby('VM', as_index=False) # Keep 'VM' as a column
                [existing_storage_columns].sum() # Sum only the numeric columns
            )
            # aggregated_df will have 'VM' and the summed storage columns

            # Ensure vInfo also has a 'VM' column
            if 'VM' not in tcdf1.columns:
                 raise ValueError("'VM' column not found in vInfo sheet, cannot merge.")

            # Merge aggregated data into vInfo using a left merge
            vSummary_df = pd.merge(tcdf1, aggregated_df, on='VM', how='left')

            # Fill NaN values that might result from the merge (VMs in vInfo but not vPart)
            # Important: Fill only the newly merged columns with 0
            for col in existing_storage_columns:
                 if col in vSummary_df.columns: # Check if column exists after merge
                      vSummary_df[col].fillna(0, inplace=True)


            # --- Reorder columns (Optional but nice) ---
            # Place aggregated columns after "In Use GB" if it exists
            if 'In Use GB' in vSummary_df.columns:
                try:
                    in_use_gb_index = vSummary_df.columns.get_loc('In Use GB')
                    # Columns before 'In Use GB' (inclusive)
                    cols_before = vSummary_df.columns[:in_use_gb_index + 1].tolist()
                    # Columns after 'In Use GB' (excluding the newly added ones)
                    cols_after = [col for col in vSummary_df.columns[in_use_gb_index + 1:] if col not in existing_storage_columns]

                    # Construct the new order: before + aggregated + after
                    new_column_order = cols_before + existing_storage_columns + cols_after
                    vSummary_df = vSummary_df[new_column_order]
                except KeyError:
                     print("Warning: 'In Use GB' column not found for reordering vSummary.")
                     # Keep default merge order
            else:
                 print("Warning: 'In Use GB' column not found in vSummary base data.")
            # --- End Reorder ---


        # --- Save the results ---
        # Use mode='w' to write only the sheets we want (vSummary, vInfo, vDisk, vPartition)
        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            # Write vSummary (which contains merged data or just vInfo data if aggregation failed)
            vSummary_df.to_excel(writer, sheet_name='vSummary', index=False)
            print("Written vSummary sheet.")

            # Write back the original (or processed by del_cols) vInfo and vPartition sheets
            # Ensure we use the dataframes read at the start of *this* function
            if 'vInfo' in dfs:
                dfs['vInfo'].to_excel(writer, sheet_name='vInfo', index=False)
                print("Written vInfo sheet.")
            if 'vDisk' in dfs: # Preserve vDisk if it existed
                dfs['vDisk'].to_excel(writer, sheet_name='vDisk', index=False)
                print("Written vDisk sheet.")
            if 'vPartition' in dfs:
                dfs['vPartition'].to_excel(writer, sheet_name='vPartition', index=False)
                print("Written vPartition sheet.")

    except FileNotFoundError:
         print(f"Error: File not found for trunc_cols_vPart: {destfile_path}")
         raise
    except ValueError as e: # Catch errors like missing VM columns or required sheets
         print(f"Error during trunc_cols_vPart (Pandas Preprocessing): {e}")
         raise
    except Exception as e:
        print(f"Error during trunc_cols_vPart (Pandas): {e}")
        traceback.print_exc()
        raise


## Remove Excess vSummary Cols (Using openpyxl AFTER Pandas)
def trimvSum1(destfile):
    """Removes specific columns by index from the vSummary sheet using openpyxl.
       Note: Column indices might change if the Pandas merge/reorder changes.
       This function assumes a specific column order post-trunc_cols_vPart.
       It's generally safer to delete by *name* if possible, but the original used indices.
    """
    if 'vSummary' not in destfile.sheetnames:
        print("Warning: vSummary sheet not found for trimming.")
        return

    vSum_ws = destfile['vSummary']
    try:
        # Original indices were: delete 3 starting at 22, delete 1 at 13, delete 1 at 11
        # Let's find columns by name if possible for robustness
        headers = [cell.value for cell in vSum_ws[1]] # Get header row values

        # Find indices by name (adjust names if they differ slightly)
        prov_mib_idx = headers.index('Provisioned MiB') + 1 if 'Provisioned MiB' in headers else None
        in_use_mib_idx = headers.index('In Use MiB') + 1 if 'In Use MiB' in headers else None
        # Which columns were originally at 22, 23, 24? Need to know what they were.
        # Assuming they might be Datacenter, Cluster, OS Config, OS Tools?
        # Example: If we want to delete 'Datacenter', 'Cluster'
        cols_to_delete_by_name = ['Provisioned MiB', 'In Use MiB'] # Add other names if known

        deleted_count = 0
        # Delete from right to left to avoid index shifting issues
        # Find indices first
        indices_to_delete = []
        if in_use_mib_idx and in_use_mib_idx <= vSum_ws.max_column: indices_to_delete.append(in_use_mib_idx)
        if prov_mib_idx and prov_mib_idx <= vSum_ws.max_column: indices_to_delete.append(prov_mib_idx)
        # Add indices of other columns if found by name

        # Sort indices in descending order for safe deletion
        indices_to_delete.sort(reverse=True)

        if not indices_to_delete:
             print("Warning: Columns specified for deletion in trimvSum1 ('Provisioned MiB', 'In Use MiB') not found.")
             # Attempt original index-based deletion as fallback? Risky.
             # print("Attempting original index-based deletion (may be incorrect):")
             # try:
             #      vSum_ws.delete_cols(22, 3) # Risky
             #      vSum_ws.delete_cols(13, 1) # Risky
             #      vSum_ws.delete_cols(11, 1) # Risky
             # except Exception as idx_e:
             #      print(f"Index-based deletion failed: {idx_e}")
             return # Exit if columns not found by name

        print(f"Attempting to delete columns at indices: {indices_to_delete}")
        for col_idx in indices_to_delete:
             vSum_ws.delete_cols(col_idx, 1)
             print(f"Deleted column at index {col_idx}")


    except ValueError as ve:
        print(f"Warning: Could not find one or more columns by name for deletion in trimvSum1: {ve}. Check headers.")
    except Exception as e:
        print(f"Error during trimvSum1 (openpyxl column deletion): {e}")
        traceback.print_exc()


## Consolidate Consumed GB Column (Using openpyxl AFTER Pandas)
def consol_vSum(destfile):
    """Copies 'In Use GB' to 'Consumed GB' if 'Consumed GB' is empty/blank on vSummary sheet."""

    if 'vSummary' not in destfile.sheetnames:
        print("Warning: vSummary sheet not found for consolidation.")
        return

    vSum_ws = destfile['vSummary']
    try:
        # Find the column indices for 'Consumed GB' and 'In Use GB'
        consumed_gb_col_idx = None
        in_use_gb_col_idx = None
        headers = [cell.value for cell in vSum_ws[1]] # Header row
        for idx, header in enumerate(headers, 1): # 1-based index
            if header == 'Consumed GB':
                consumed_gb_col_idx = idx
            elif header == 'In Use GB':
                in_use_gb_col_idx = idx

        if consumed_gb_col_idx is None or in_use_gb_col_idx is None:
            missing = []
            if consumed_gb_col_idx is None: missing.append("'Consumed GB'")
            if in_use_gb_col_idx is None: missing.append("'In Use GB'")
            print(f"Error: {', '.join(missing)} column(s) not found in vSummary for consolidation.")
            return # Cannot proceed

        # Iterate over rows, starting from the second row
        for row_idx in range(2, vSum_ws.max_row + 1):
            consumed_gb_cell = vSum_ws.cell(row=row_idx, column=consumed_gb_col_idx)
            in_use_gb_cell = vSum_ws.cell(row=row_idx, column=in_use_gb_col_idx)

            # Check if 'Consumed GB' is empty or blank (None or empty string)
            consumed_value = consumed_gb_cell.value
            if consumed_value is None or (isinstance(consumed_value, str) and consumed_value.strip() == ""):
                 # Copy value and also number format if possible
                 consumed_gb_cell.value = in_use_gb_cell.value
                 if in_use_gb_cell.has_style and in_use_gb_cell.number_format:
                      consumed_gb_cell.number_format = in_use_gb_cell.number_format


    except Exception as e:
        print(f"Error during consol_vSum (openpyxl): {e}")
        traceback.print_exc()


## Filter First Row
def filter_rows(destfile):
  """Applies auto-filter to the header row on all sheets.

  Args:
    destfile: An xl workbook object.
  """
  try:
    for ws in destfile.worksheets:
      # Check if sheet has any data
      if ws.max_row > 0 and ws.max_column > 0:
          # Set the auto-filter range to encompass all columns in the first row
          # Use get_column_letter utility
          last_col_letter = xl.utils.get_column_letter(ws.max_column)
          filter_range = f"A1:{last_col_letter}1"
          ws.auto_filter.ref = filter_range
          # print(f"Applied auto_filter '{filter_range}' to sheet '{ws.title}'") # Optional log
      # else: print(f"Skipping filter on empty sheet '{ws.title}'") # Optional log
  except Exception as e:
      print(f"Error applying filters: {e}")
      traceback.print_exc()




# ================================================================================================================================================================================================================================================
# ===================================================================================================== MAIN PROCESSING FUNCTION FOR WEB APP =====================================================================================================
# ================================================================================================================================================================================================================================================

def process_rvtools_file(input_filepath, output_folder, original_basename):
    """
    Processes the RVTools Excel file, yielding SSE-formatted progress updates.

    Args:
        input_filepath (str): The full path to the *saved* RVTools Excel file.
        output_folder (str): The folder where the processed file should be saved.
        original_basename (str): The original filename as uploaded by the user.

    Yields:
        str: Server-Sent Event formatted strings (e.g., "data: Processing step...\n\n").
             Includes a final 'result' event with JSON payload indicating outcome.
    """
    # Helper to format SSE messages for the client
    def sse_message(data):
        # Replace newlines in message data to avoid breaking SSE format
        data = str(data).replace('\n', ' ')
        return f"data: {data}\n\n"

    # Helper to yield a final result event to the client
    def yield_result(success, message):
        payload = json.dumps({"success": success, "message": message})
        yield f"event: result\ndata: {payload}\n\n"

    # --- Start Processing ---
    yield sse_message(f"Starting analysis for: {original_basename}")
    time.sleep(0.1) # Give a small delay for messages to flush

    destfile = None # Initialize destfile to None for error handling scope
    try:
        # --- Basic File Setup ---
        if not os.path.exists(input_filepath):
             print(f"ERROR: Input file not found: {input_filepath}") # Server log
             yield from yield_result(False, f"Server Error: Input file missing.")
             return # Stop generation
        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
                yield sse_message(f"Created output folder: {output_folder}")
                print(f"Created output folder: {output_folder}") # Server log
            except OSError as e:
                print(f"ERROR: Failed to create output folder '{output_folder}': {e}") # Server log
                yield from yield_result(False, f"Server Error: Cannot create folder.")
                return

        # Construct output filename (same logic as before)
        if original_basename:
            base_name_no_ext = os.path.splitext(original_basename)[0]
            output_filename = f"{base_name_no_ext}-EDITED.xlsx"
        else:
            temp_base = os.path.splitext(os.path.basename(input_filepath))[0]
            output_filename = f"{temp_base}-ORIGINAL_NAME_MISSING-EDITED.xlsx"
            yield sse_message("Warning: Original filename missing, using temporary name.")
            print("Warning: Original filename was not provided; using temporary name as base.") # Server log

        output_filepath = os.path.join(output_folder, output_filename)
        yield sse_message(f"Output file will be: {output_filename}")
        print(f"Output file will be: {output_filepath} (Derived from original: '{original_basename}')") # Server log
        time.sleep(0.1)

        # --- Core Processing Logic (yielding progress) ---

        # Step 1: Load source workbook & save copy
        yield sse_message("Step 1: Loading source workbook and creating copy...")
        time.sleep(0.1)
        srcfile = xl.load_workbook(input_filepath)
        srcfile.save(output_filepath)
        srcfile.close()
        destfile = xl.load_workbook(output_filepath) # Assign to destfile
        yield sse_message(" -> Workbook copied.")
        time.sleep(0.1)

        # Step 2: Remove extra sheets
        yield sse_message("Step 2: Removing unnecessary sheets...")
        time.sleep(0.1)
        keep_sheets = ['vInfo', 'vDisk', 'vPartition']
        sheets_to_delete = [s for s in destfile.sheetnames if s not in keep_sheets]
        if not any(s in destfile.sheetnames for s in keep_sheets):
             print("ERROR: Required sheets missing.") # Server log
             if destfile:
                 try: destfile.close()
                 except: pass
             yield from yield_result(False, "Error: Required sheets (vInfo, vDisk, vPartition) not found.")
             return
        for sheetName in sheets_to_delete:
            del destfile[sheetName]
        yield sse_message(f" -> Kept sheets: {destfile.sheetnames}")
        time.sleep(0.1)

        # Step 3: Remove Formatting
        yield sse_message("Step 3: Removing initial formatting...")
        time.sleep(0.1)
        removeFormatting(destfile) # Assuming helper funcs don't yield
        yield sse_message(" -> Formatting removed.")
        time.sleep(0.1)

        # Step 4: Add and Rename Columns
        yield sse_message("Step 4: Adding and renaming category/status columns...")
        time.sleep(0.1)
        target_sheets = [s for s in keep_sheets if s in destfile.sheetnames]
        for sheet_name in target_sheets:
            worksheet = destfile[sheet_name]
            worksheet.insert_cols(3, 6)
            worksheet['C1'] = 'IsFile'; worksheet['D1'] = 'IsSQL'; worksheet['E1'] = 'IsOrcl'
            worksheet['F1'] = 'IsPGres'; worksheet['G1'] = 'IsExch'; worksheet['H1'] = 'IsTestDev'
            if sheet_name == 'vInfo':
                worksheet.insert_cols(9, 1); worksheet['I1'] = 'HasTools'
            elif sheet_name == 'vDisk':
                worksheet.insert_cols(9, 1); worksheet['I1'] = 'DiskCount'
        yield sse_message(" -> Columns added and renamed.")
        time.sleep(0.1)

        # Step 5: Match Workload Types & Set Defaults
        yield sse_message("Step 5: Matching workload types...")
        time.sleep(0.1)
        match_fs(destfile); match_sql(destfile); match_orcl(destfile); match_pgres(destfile)
        match_exch(destfile); match_tstdev(destfile); match_gendb(destfile)
        yield sse_message(" -> Workload types matched.")
        yield sse_message("Step 5b: Setting default 'No' for unmatched...")
        time.sleep(0.1)
        set_no_values(destfile)
        yield sse_message(" -> Default 'No' values set.")
        time.sleep(0.1)

        # Step 6: Compare vInfo to vPartition for 'HasTools'
        yield sse_message("Step 6: Comparing vInfo and vPartition for 'HasTools'...")
        time.sleep(0.1)
        vinfo_ws = destfile['vInfo'] if 'vInfo' in destfile else None
        vpart_ws = destfile['vPartition'] if 'vPartition' in destfile else None
        if vinfo_ws and vpart_ws:
             compare_vms(vinfo_ws, vpart_ws)
             yield sse_message(" -> 'HasTools' status updated in vInfo.")
        else:
             yield sse_message(" -> Skipped 'HasTools' comparison (vInfo or vPartition missing).")
        time.sleep(0.1)

        # Step 7: Fill DiskCount in vDisk
        yield sse_message("Step 7: Filling 'DiskCount' values in vDisk...")
        time.sleep(0.1)
        if 'vDisk' in destfile:
            vdisk_ws = destfile['vDisk']
            if not vdisk_diskcount_val(vdisk_ws):
                yield sse_message(" -> Warning: Failed to fill DiskCount values.")
                print("Warning: Failed to fill DiskCount values.") # Server log
            else:
                yield sse_message(" -> 'DiskCount' values set.")
        else:
             yield sse_message(" -> Skipped 'DiskCount' (vDisk sheet missing).")
        time.sleep(0.1)

        # Step 8: Insert GB Columns and Calculate
        yield sse_message("Step 8: Calculating GB values from MiB...")
        time.sleep(0.1)
        if 'vInfo' in destfile:
            vinfo_ws = destfile['vInfo']
            vinfo_insert_gb_col(vinfo_ws, "Provisioned MiB", "Provisioned GB")
            vinfo_insert_gb_col(vinfo_ws, "In Use MiB", "In Use GB")
        if 'vDisk' in destfile:
            vdisk_ws = destfile['vDisk']
            vdisk_insert_gb_col(vdisk_ws, "Capacity MiB", "Capacity GB")
        if 'vPartition' in destfile:
            vpart_ws = destfile['vPartition']
            vpart_insert_gb_col(vpart_ws, "Capacity MiB", "Capacity GB")
            vpart_insert_gb_col(vpart_ws, "Consumed MiB", "Consumed GB")
            vpart_insert_gb_col(vpart_ws, "Free MiB", "Free GB")
        yield sse_message(" -> GB columns added and calculated.")
        time.sleep(0.1)

        # --- Save before Pandas ---
        yield sse_message("Saving intermediate results before Pandas processing...")
        time.sleep(0.1)
        destfile.save(output_filepath)
        destfile.close() # Close openpyxl handle before pandas writes
        destfile = None # Clear variable as it's closed
        yield sse_message(" -> Intermediate save complete.")
        time.sleep(0.1)

        # Step 9: Delete unnecessary columns (Pandas)
        yield sse_message("Step 9: Deleting columns using Pandas...")
        time.sleep(0.1)
        try:
            del_cols_vInfo(output_filepath) # Reads/writes file
            yield sse_message(" -> Columns deleted.")
        except Exception as pd_del_err:
            print(f"ERROR during Pandas column deletion: {pd_del_err}") # Server log
            yield from yield_result(False, f"Error deleting columns via Pandas: {pd_del_err}")
            return
        time.sleep(0.1)

        # Step 10: Aggregate vPartition (Pandas)
        yield sse_message("Step 10: Creating vSummary sheet using Pandas...")
        time.sleep(0.1)
        try:
            trunc_cols_vPart(output_filepath) # Reads/adds sheet/writes file
            yield sse_message(" -> vSummary sheet created.")
        except Exception as pd_trunc_err:
            print(f"ERROR during Pandas summary creation: {pd_trunc_err}") # Server log
            yield from yield_result(False, f"Error creating vSummary via Pandas: {pd_trunc_err}")
            return
        time.sleep(0.1)

        # --- Reload with openpyxl ---
        yield sse_message("Reloading workbook for final cleanup...")
        time.sleep(0.1)
        destfile = xl.load_workbook(output_filepath) # Re-assign destfile

        # Step 11: Final Formatting, Trim, Consolidate
        yield sse_message("Step 11: Final cleanup (Formatting, Trim, Consolidate)...")
        time.sleep(0.1)
        removeFormatting(destfile)
        if 'vSummary' in destfile:
            trimvSum1(destfile); consol_vSum(destfile)
            yield sse_message(" -> vSummary sheet trimmed and consolidated.")
        else:
            yield sse_message(" -> Skipped vSummary cleanup (sheet not found).")
            print("Warning: vSummary sheet not found after Pandas operations.") # Server log
        time.sleep(0.1)

        # Step 12: Apply Filters
        yield sse_message("Step 12: Applying auto-filters...")
        time.sleep(0.1)
        filter_rows(destfile)
        yield sse_message(" -> Filters applied.")
        time.sleep(0.1)

        # Final save
        yield sse_message("Saving final processed workbook...")
        time.sleep(0.1)
        destfile.save(output_filepath)
        yield sse_message(" -> Final save complete.")
        if destfile:
            try: destfile.close()
            except: pass
            destfile = None # Clear closed workbook
        time.sleep(0.5) # Pause before final message

        # --- Success ---
        yield sse_message("--- Analysis Completed Successfully ---")
        print(f"--- RVTools Analysis Completed Successfully for: {original_basename} ---") # Server log
        # Yield the final result event with success and the output filename
        yield from yield_result(True, output_filename) # output_filename is the correctly formatted name

    except Exception as e:
        # Attempt to close file handle if open
        if destfile:
            try: destfile.close()
            except Exception as close_err:
                 print(f"Warning: Error closing workbook during exception handling: {close_err}") # Server log

        # Log full error to server console
        print(f"\n--- ERROR during SSE processing for {original_basename} ---")
        traceback.print_exc()
        print("--- End Error Traceback ---")

        # Yield user-friendly error message and the final result event
        error_message = f"An error occurred: {type(e).__name__}"
        yield sse_message(f"ERROR: {error_message} - Check server logs for details.")
        time.sleep(0.1)
        yield from yield_result(False, error_message) # Send final error status to client

# ===============================================================================================================================================================================================================================================
# ================================================================================================================ END OF SCRIPT ================================================================================================================
# ===============================================================================================================================================================================================================================================