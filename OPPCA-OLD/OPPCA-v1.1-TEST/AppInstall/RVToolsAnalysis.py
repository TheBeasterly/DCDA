import os
import sys
import openpyxl as xl
import pandas as pd
import tkinter as tk
from tkinter import filedialog

def main(label2):

################# DEFS #################

    ## Remove Formatting
    def removeFormatting(destfile):
      """Removes all formatting from an excel workbook.

      Args:
        destfile: An xl workbook object.
      """

      for worksheet in destfile.worksheets:
        for cell in worksheet.iter_rows():
          for c in cell:
            c.style = 'Normal'

    ## Match File Servers
    def match_fs(destfile):
      """Matches file servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      fs_str = ["file", "fs", "nas", "share", "ftp"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in fs_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 2).value = "Yes"

    ## Match SQL DBs
    def match_sql(destfile):
      """Matches SQL DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      sql_str = ["sql"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in sql_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 3).value = "Yes"

    ## Match Oracle DBs
    def match_orcl(destfile):
      """Matches Oracle DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      orcl_str = ["orcl", "oracle"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in orcl_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 4).value = "Yes"

    ## Match PostGres DBs
    def match_pgres(destfile):
      """Matches PostGres DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      pgres_str = ["pgres", "postgres"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in pgres_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 5).value = "Yes"

    ## Match Possible DBs
    def match_gendb(destfile):
      """Matches Possible DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      gendb_str = ["db", "database"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in gendb_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 3).value = "Check"
            cell.offset(0, 4).value = "Check"
            cell.offset(0, 5).value = "Check"

    ## Match Exchange Servers
    def match_exch(destfile):
      """Matches Exchange Servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      exch_str = ["exch", "exchange"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in exch_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 6).value = "Yes"

    ## Match TestDev
    def match_tstdev(destfile):
      """Matches TestDev in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      tstdev_str = ["tst", "test", "dev"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in tstdev_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 7).value = "Yes"

    ## Set No Values
    # Set Used Range of Column A
    def get_used_range(worksheet):
      """Gets the used range of a worksheet.

      Args:
        worksheet: An xl worksheet object.

      Returns:
        A tuple of two cell objects, representing the start and end of the used range of the worksheet.
      """

      # Get the first and last row numbers in the worksheet.
      first_row_number = worksheet.min_row
      last_row_number = worksheet.max_row

      # Iterate over the cells in column A and find the first non-empty cell.
      start_cell = None
      for row_index in range(first_row_number, last_row_number + 1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          start_cell = cell
          break

      # If the first non-empty cell in column A is None, then the used range of the worksheet is empty.
      if start_cell is None:
        return None

      # Find the last non-empty cell in column A.
      end_cell = None
      for row_index in range(last_row_number, first_row_number - 1, -1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          end_cell = cell
          break

      # Return the start and end cells of the used range of the worksheet.
      return start_cell, end_cell

    #Set No Values in columns C through H
    def set_no_values(destfile, used_range_a):
      """Sets the value of all empty cells in columns C to H to "No", based on the used range of column A.

      Args:
        destfile: An xl workbook object.
        used_range_a: A tuple of two cell objects, representing the start and end of the used range of column A.
      """

      # Get the start and end row numbers of the used range of column A.
      start_row_number = used_range_a[0].row
      end_row_number = used_range_a[1].row

      # Iterate over the rows in the used range of column A.
      for row_index in range(start_row_number, end_row_number + 1):

        # Iterate over the cells in columns C to H.
        for column_index in range(3, 9):

          # Get the cell object.
          cell = destfile.active.cell(row=row_index, column=column_index)

          # If the cell value is empty, then set it to "No".
          if cell.value is None:
            cell.value = "No"


    ## Insert Columns for MiB to GB Math
    # Find "Provisioned MiB" on vInfo Sheet
    def vinfo_findprovmib(vinfo_ws):
      """Finds the cell in row A of the "vInfo" worksheet that contains the string "Provisioned MiB".

      Args:
        vinfo_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Provisioned MiB", or None if the string is not found.
      """

      for row in vinfo_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Provisioned MiB":
            return cell

      return None

    # Insert Column after "Provisioned MiB" on vInfo Sheet
    def vinfo_provmib_inscol(vinfo_ws, vinfoprovmib_cell):
      """Inserts a new column to the right of the specified cell in the "vInfo" worksheet.

      Args:
        vinfo_ws: An xl worksheet object.
        vinfoprovmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vinfoprovmib_cell.column
      vinfo_ws.insert_cols(column_index + 1)
      new_column_index = vinfoprovmib_cell.column + 1

      # Set the header of the newly added column
      vinfo_ws.cell(row=1, column=new_column_index).value = "Provisioned GB"

      # Calculate and insert "Provisioned GB" values
      for i in range(2, vinfo_ws.max_row + 1):
        prov_mib_value = vinfo_ws.cell(row=i, column=column_index).value

         # Handle cases where the 'Provisioned MiB' cell is empty or not numeric
        if prov_mib_value is None or not isinstance(prov_mib_value, (int, float)):
            prov_gb_value = None  # Set to None or a default value
        else:
            prov_gb_value = round(prov_mib_value / 953.7, 2)

        vinfo_ws.cell(row=i, column=new_column_index).value = prov_gb_value

    # Find "In Use MiB" on vInfo Sheet
    def vinfo_findinusemib(vinfo_ws):
      """Finds the cell in row A of the "vInfo" worksheet that contains the string "In Use MiB".

      Args:
        vinfo_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "In Use MiB", or None if the string is not found.
      """

      for row in vinfo_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "In Use MiB":
            return cell

      return None

    # Insert Column after "In Use MiB" on vInfo Sheet
    def vinfo_inusemib_inscol(vinfo_ws, vinfoinusemib_cell):
      """Inserts a new column to the right of the specified cell in the "vInfo" worksheet.

      Args:
        vinfo_ws: An xl worksheet object.
        vinfoinusemib_cell: The cell to insert the new column to the right of.
      """

      column_index = vinfoinusemib_cell.column
      vinfo_ws.insert_cols(column_index + 1)
      new_column_index = vinfoinusemib_cell.column + 1

      # Set the header of the newly added column
      vinfo_ws.cell(row=1, column=new_column_index).value = "In Use GB"

      # Calculate and insert "In Use GB" values
      for i in range(2, vinfo_ws.max_row + 1):
        inuse_mib_value = vinfo_ws.cell(row=i, column=column_index).value

         # Handle cases where the 'In Use MiB' cell is empty or not numeric
        if inuse_mib_value is None or not isinstance(inuse_mib_value, (int, float)):
            inuse_gb_value = None  # Set to None or a default value
        else:
            inuse_gb_value = round(inuse_mib_value / 953.7, 2)

        vinfo_ws.cell(row=i, column=new_column_index).value = inuse_gb_value

    # Find "Capacity MiB" on vDisk Sheet
    def vdisk_findcapacitymib(vdisk_ws):
      """Finds the cell in row A of the "vDisk" worksheet that contains the string "Capacity MiB".

      Args:
        vdisk_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Capacity MiB", or None if the string is not found.
      """

      for row in vdisk_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Capacity MiB":
            return cell

      return None

    # Insert Column after "Capacity MiB" on vDisk Sheet
    def vdisk_capacitymib_inscol(vdisk_ws, vdiskcapmib_cell):
      """Inserts a new column to the right of the specified cell in the "vDisk" worksheet.

      Args:
        vdisk_ws: An xl worksheet object.
        vdiskcapmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vdiskcapmib_cell.column
      vdisk_ws.insert_cols(column_index + 1)
      new_column_index = vdiskcapmib_cell.column + 1

      # Set the header of the newly added column
      vdisk_ws.cell(row=1, column=new_column_index).value = "Capacity GB"

      # Calculate and insert "Capacity GB" values
      for i in range(2, vdisk_ws.max_row + 1):
        cap_mib_value = vdisk_ws.cell(row=i, column=column_index).value

        # Handle cases where the 'Capacity MiB' cell is empty or not numeric
        if cap_mib_value is None or not isinstance(cap_mib_value, (int, float)):
            cap_gb_value = None  # Set to None or a default value
        else:
            cap_gb_value = round(cap_mib_value / 953.7, 2)

        vdisk_ws.cell(row=i, column=new_column_index).value = cap_gb_value

    # Find "DiskCount" on vDisk Sheet
    def vdisk_finddiskcount(vdisk_ws):
      """Finds the cell in row 1 of the "vDisk" worksheet that contains the string "DiskCount".

      Args:
        vdisk_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "DiskCount", or None if the string is not found.
      """

      for row in vdisk_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "DiskCount":
            return cell

      return None

    # Insert Value of "1" in "DiskCount" Column on vDisk Sheet
    def vdisk_diskcount_val(vdisk_ws, vdiskdiskcount_cell):
      """Inserts a value of "1" in the "DiskCount" Column on the "vDisk" worksheet.

      Args:
        vdisk_ws: An xl worksheet object.
        vdiskdiskcount_cell: The cell to insert the new column to the right of.
      """

      column_index = vdiskdiskcount_cell.column

      # Set the value of the cell in row 1 of the column to "DiskCount".
      vdisk_ws.cell(row=1, column=column_index).value = "DiskCount"

      for i in range(2, vdisk_ws.max_row + 1):
        vdisk_ws.cell(row=i, column=column_index).value = 1

    # Find "Capacity MiB" on vPartition Sheet
    def vpart_findcapacitymib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Capacity MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Capacity MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Capacity MiB":
            return cell

      return None

    # Insert Column after "Capacity MiB" on vPartition Sheet
    def vpart_capacitymib_inscol(vpart_ws, vpartcapmib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartcapmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartcapmib_cell.column
      vpart_ws.insert_cols(column_index + 1)
      new_column_index = vpartcapmib_cell.column + 1

      # Set the header of the newly added column
      vpart_ws.cell(row=1, column=new_column_index).value = "Capacity GB"

      # Calculate and insert "Capacity GB" values
      for i in range(2, vpart_ws.max_row + 1):
        vpartcap_mib_value = vpart_ws.cell(row=i, column=column_index).value

        # Handle cases where the 'Capacity MiB' cell is empty or not numeric
        if vpartcap_mib_value is None or not isinstance(vpartcap_mib_value, (int, float)):
            vpartcap_gb_value = None  # Set to None or a default value
        else:
            vpartcap_gb_value = round(vpartcap_mib_value / 953.7, 2)

        vpart_ws.cell(row=i, column=new_column_index).value = vpartcap_gb_value

    # Find "Consumed MiB" on vPartition Sheet
    def vpart_findconsumedmib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Consumed MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Consumed MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Consumed MiB":
            return cell

      return None

    # Insert Column after "Consumed MiB" on vPartition Sheet
    def vpart_consumedmib_inscol(vpart_ws, vpartconsmib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartconsmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartconsmib_cell.column
      vpart_ws.insert_cols(column_index + 1)
      new_column_index = vpartconsmib_cell.column + 1

      # Set the header of the newly added column
      vpart_ws.cell(row=1, column=new_column_index).value = "Consumed GB"

      # Calculate and insert "Consumed GB" values
      for i in range(2, vpart_ws.max_row + 1):
        cons_mib_value = vpart_ws.cell(row=i, column=column_index).value

        # Handle cases where the 'Consumed MiB' cell is empty or not numeric
        if cons_mib_value is None or not isinstance(cons_mib_value, (int, float)):
            cons_gb_value = None  # Set to None or a default value
        else:
            cons_gb_value = round(cons_mib_value / 953.7, 2)

        vpart_ws.cell(row=i, column=new_column_index).value = cons_gb_value

    # Find "Free MiB" on vPartition Sheet
    def vpart_findfreemib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Free MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Free MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Free MiB":
            return cell

      return None

    # Insert Column after "Free MiB" on vPartition Sheet
    def vpart_freemib_inscol(vpart_ws, vpartfreemib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartfreemib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartfreemib_cell.column
      vpart_ws.insert_cols(column_index + 1)
      new_column_index = vpartfreemib_cell.column + 1

      # Set the header of the newly added column
      vpart_ws.cell(row=1, column=new_column_index).value = "Free GB"

      # Calculate and insert "Free GB" values
      for i in range(2, vpart_ws.max_row + 1):
        free_mib_value = vpart_ws.cell(row=i, column=column_index).value

        # Handle cases where the 'Free MiB' cell is empty or not numeric
        if free_mib_value is None or not isinstance(free_mib_value, (int, float)):
            free_gb_value = None  # Set to None or a default value
        else:
            free_gb_value = round(free_mib_value / 953.7, 2)

        vpart_ws.cell(row=i, column=new_column_index).value = free_gb_value


    ## Compare VMs on vInfo Sheet to vPart Sheet
    def compare_vms(vinfo_ws, vpart_ws):
      """Compares the vInfo worksheet to the vPartition worksheet and sets the "HasTools" column in the vInfo worksheet to "Yes" if a match is found, or "No" if no match is found.

      Args:
        vinfo_ws: An xl Worksheet object for the vInfo worksheet.
        vpart_ws: An xl Worksheet object for the vPartition worksheet.
      """

      # Declare the vinfo_cell variable.
      vinfo_cell = None

      # Iterate over all of the rows in the vInfo worksheet.
      for row in vinfo_ws.rows:

        # Get the cell value in the first column of the row.
        vinfo_cell_value = row[0].value

        # If the cell value is not zero or blank, then check if there is a matching cell in the vPartition worksheet.
        if vinfo_cell_value != 0 and vinfo_cell_value != "":

          # Assign the current row to the vinfo_cell variable.
          vinfo_cell = row[0]

          # Iterate over all of the rows in the vPartition worksheet.
          for vpart_row in vpart_ws.rows:

            # Get the cell value in the first column of the row.
            vpart_cell_value = vpart_row[0].value

            # If a matching cell is found, then set the "HasTools" column in the vInfo worksheet to "Yes".
            if vpart_cell_value == vinfo_cell_value:
              vinfo_cell.offset(0, 8).value = "Yes"
              break

          # If no matching cell is found, then set the "HasTools" column in the vInfo worksheet to "No".
          else:
            vinfo_cell.offset(0, 8).value = "No"

      # Set the header for the "HasTools" column.
      vinfo_ws["I1"].value = "HasTools"


    ## Delete Unnecessary Columns
    def del_cols_vInfo(destfile_name):
      # Load the Excel file into a Pandas DataFrame
      df1 = pd.read_excel(destfile_name, sheet_name='vInfo')
      df2 = pd.read_excel(destfile_name, sheet_name='vDisk')
      df3 = pd.read_excel(destfile_name, sheet_name='vPartition')

      # Get a list of all valid column indices
      valid_column_indices1 = list(df1.columns)
      valid_column_indices2 = list(df2.columns)
      valid_column_indices3 = list(df3.columns)

      # Get the column indices to keep
      keep_cols_vInfo = [col_idx for col_idx in valid_column_indices1 if col_idx in ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'HasTools', 'Disks', 'Total disk capacity', 'Provisioned MiB', 'Provisioned GB', 'In Use MiB', 'In Use GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']]
      keep_cols_vDisk = [col_idx for col_idx in valid_column_indices2 if col_idx in ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'DiskCount', 'Disk', 'Capacity MiB', 'Capacity GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']]
      keep_cols_vPart = [col_idx for col_idx in valid_column_indices3 if col_idx in ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'Disk', 'Capacity MiB', 'Capacity GB', 'Consumed MiB', 'Consumed GB', 'Free MiB', 'Free GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']]

      # Create a new DataFrame with only the columns that you want to keep
      new_df1 = df1.loc[:, keep_cols_vInfo]
      new_df2 = df2.loc[:, keep_cols_vDisk]
      new_df3 = df3.loc[:, keep_cols_vPart]

      # Save the adjusted worksheets to the same workbook
      writer = pd.ExcelWriter(destfile_name, mode='w')
      new_df1.to_excel(writer, sheet_name='vInfo', index=False)
      new_df2.to_excel(writer, sheet_name='vDisk', index=False)
      new_df3.to_excel(writer, sheet_name='vPartition', index=False)
      writer.close()


    ## Truncate vPartition Storage Capacity Cols into vInfo
    def trunc_cols_vPart(destfile_name):
      # Load the Excel file into a Pandas DataFrame
      tcdf1 = pd.read_excel(destfile_name, sheet_name='vInfo')
      tcdf2 = pd.read_excel(destfile_name, sheet_name='vDisk')
      tcdf3 = pd.read_excel(destfile_name, sheet_name='vPartition')

      # Identify storage-related columns
      storage_columns = [col_idx for col_idx in tcdf3.columns if col_idx in ['Capacity GB', 'Consumed GB', 'Free GB']]
      
      # Aggregate storage data from vPartition
      aggregated_df = (
          tcdf3.groupby('VM')
          .agg({col: 'sum' for col in storage_columns})
          .reset_index()
          .fillna("0")  # Fill missing values with 0
      )

      # Merge aggregated data into vInfo
      vSummary_df = pd.merge(tcdf1, aggregated_df, on='VM', how='left')

      # Reorder columns to place aggregated columns after "In Use GB"
      in_use_gb_index = vSummary_df.columns.get_loc('In Use GB')
      new_column_order = (
          vSummary_df.columns[:in_use_gb_index + 1].tolist() + 
          aggregated_df.columns[1:].tolist() +  # Exclude 'VM' which is already in tcdf1
          vSummary_df.columns[in_use_gb_index + 1:].tolist()
      )
      vSummary_df = vSummary_df[new_column_order]

      # Save the merged DataFrame to a new sheet 'vSummary'
      with pd.ExcelWriter(destfile_name, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
          vSummary_df.to_excel(writer, sheet_name='vSummary', index=False)

      # Save or use the merged DataFrame
      writer = pd.ExcelWriter(destfile_name, mode='w')
      vSummary_df.to_excel(writer, sheet_name='vSummary', index=False)
      tcdf1.to_excel(writer, sheet_name='vInfo', index=False)
      tcdf2.to_excel(writer, sheet_name='vDisk', index=False)
      tcdf3.to_excel(writer, sheet_name='vPartition', index=False)
      writer.close()


    ## Remove Excess vSummary Cols
    def trimvSum1(destfile):
      vSum_ws = destfile['vSummary']
      vSum_ws.delete_cols(22, 3)
      vSum_ws.delete_cols(13, 1)
      vSum_ws.delete_cols(11, 1)


    ## Consolidate Consumed GB Column
    def consol_vSum(destfile):

      # Get the 'vSummary' worksheet
      vSum_ws = destfile['vSummary']

      # Find the column indices for 'Consumed GB' and 'In Use GB'
      consumed_gb_col_idx = None
      in_use_gb_col_idx = None
      for col_idx, col in enumerate(vSum_ws.iter_cols(1, vSum_ws.max_column), 1):  # 1-based indexing
          if col[0].value == 'Consumed GB':
              consumed_gb_col_idx = col_idx
          elif col[0].value == 'In Use GB':
              in_use_gb_col_idx = col_idx

      if consumed_gb_col_idx is None or in_use_gb_col_idx is None:
          print("Error: 'Consumed GB' or 'In Use GB' column not found.")
      else:
          # Iterate over rows, starting from the second row (assuming the first row is headers)
          for row_idx in range(2, vSum_ws.max_row + 1):
              consumed_gb_cell = vSum_ws.cell(row=row_idx, column=consumed_gb_col_idx)
              in_use_gb_cell = vSum_ws.cell(row=row_idx, column=in_use_gb_col_idx)

              # Check if 'Consumed GB' is empty or blank
              if consumed_gb_cell.value is None or consumed_gb_cell.value == "":
                  consumed_gb_cell.value = in_use_gb_cell.value


    ## Filter First Row
    def filter_rows(destfile):
      """Filters the first row on all sheets.

      Args:
        destfile: An xl workbook object.
      """

      for ws in destfile.worksheets:
        # Set the auto-filter range to the first row of each sheet.
        ws.auto_filter.ref = "A1:{}".format(ws.dimensions.split(':')[1])


################# RUNTIME #################


    ## Get the Excel file name from the user
    root = tk.Tk()
    root.withdraw()

    try:
      srcfile_name = filedialog.askopenfilename()

      try:
        ## Open the Excel file
        srcfile = xl.load_workbook(srcfile_name)
        #label2.config(text=f"Working... 5% Completed...")

        ## Create a new Excel file
        destfile_name = srcfile_name[:-5] + "-EDITED.xlsx"
        srcfile.save(destfile_name)
        #label2.config(text=f"Working... 12% Completed...")

        ## Open new Excel file and remove extra sheets
        destfile = xl.load_workbook(destfile_name)
        
        keep_sheets = ['vInfo', 'vDisk', 'vPartition']
        for sheetName in destfile.sheetnames:
            if sheetName not in keep_sheets:
                del destfile[sheetName]
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 18% Completed...")

        ## Remove Formatting
        destfile = xl.load_workbook(destfile_name)
        removeFormatting(destfile)
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 25% Completed...")

        ## Add Columns
        destfile = xl.load_workbook(destfile_name)
        for worksheet in destfile.worksheets:
          worksheet.insert_cols(3, 6)
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 35% Completed...")

        ## Rename columns C:H in all sheets to "IsFile", "IsSQL", "IsOrcl", "IsPGres", "IsExch", and "IsTestDev" 
        destfile = xl.load_workbook(destfile_name)
        for worksheet in destfile.worksheets:
          worksheet['C1'] = 'IsFile'
          worksheet['D1'] = 'IsSQL'
          worksheet['E1'] = 'IsOrcl'
          worksheet['F1'] = 'IsPGres'
          worksheet['G1'] = 'IsExch'
          worksheet['H1'] = 'IsTestDev'
        #label2.config(text=f"Working... 50% Completed...")

        # Insert "HasTools" Column in vInfo Sheet
        destfile["vInfo"].insert_cols(9, 1)
        destfile["vInfo"]['I1'] = 'HasTools'
        #label2.config(text=f"Working... 57% Completed...")

        # Insert "DiskCount" Column in vDisk Sheet
        destfile["vDisk"].insert_cols(9, 1)
        destfile["vDisk"]['I1'] = 'DiskCount'
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 63% Completed...")

        ## Match Workload Types
        destfile = xl.load_workbook(destfile_name)
        match_fs(destfile)
        match_sql(destfile)
        match_orcl(destfile)
        match_pgres(destfile)
        match_gendb(destfile)
        match_exch(destfile)
        match_tstdev(destfile)
        for sheetname in destfile.sheetnames:
          destfile.active = destfile[sheetname]
          used_range_a = get_used_range(destfile.active)
          set_no_values(destfile, used_range_a)
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 75% Completed...")

        ## Del Cols
        del_cols_vInfo(destfile_name)
        #label2.config(text=f"Working... 80% Completed...")

        ## Insert Columns for MiB to GB math
        destfile = xl.load_workbook(destfile_name)
        vinfo_ws = destfile["vInfo"]
        vdisk_ws = destfile["vDisk"]
        vpart_ws = destfile["vPartition"]
        for sheet in destfile:
            destfile[sheet.title].views.sheetView[0].tabSelected = False
        removeFormatting(destfile)
        #label2.config(text=f"Working... 83% Completed...")

        # Insert Columns in vInfo Sheet
        destfile.active = vinfo_ws
        vinfoprovmib_cell = vinfo_findprovmib(vinfo_ws)
        if vinfoprovmib_cell is not None:
          vinfo_provmib_inscol(vinfo_ws, vinfoprovmib_cell)
        vinfoinusemib_cell = vinfo_findinusemib(vinfo_ws)
        if vinfoinusemib_cell is not None:
          vinfo_inusemib_inscol(vinfo_ws, vinfoinusemib_cell)
        #label2.config(text=f"Working... 86% Completed...")
        
        # Insert Columns in vDisk Sheet
        destfile.active = vdisk_ws
        vdiskcapmib_cell = vdisk_findcapacitymib(vdisk_ws)
        if vdiskcapmib_cell is not None:
          vdisk_capacitymib_inscol(vdisk_ws, vdiskcapmib_cell)
        vdiskdiskcount_cell = vdisk_finddiskcount(vdisk_ws)
        if vdiskdiskcount_cell is not None:
          vdisk_diskcount_val(vdisk_ws, vdiskdiskcount_cell)
        #label2.config(text=f"Working... 89% Completed...")

        # Insert Columns in vPartition Sheet
        destfile.active = vpart_ws
        vpartcapmib_cell = vpart_findcapacitymib(vpart_ws)
        if vpartcapmib_cell is not None:
          vpart_capacitymib_inscol(vpart_ws, vpartcapmib_cell)
        vpartconsmib_cell = vpart_findconsumedmib(vpart_ws)
        if vpartconsmib_cell is not None:
          vpart_consumedmib_inscol(vpart_ws, vpartconsmib_cell)
        vpartfreemib_cell = vpart_findfreemib(vpart_ws)
        if vpartfreemib_cell is not None:
          vpart_freemib_inscol(vpart_ws, vpartfreemib_cell)

        destfile.save(destfile_name)
        #label2.config(text=f"Working... 92% Completed...")

        ## Compare VMs
        destfile = xl.load_workbook(destfile_name)
        vinfo_ws = destfile["vInfo"]
        vpart_ws = destfile["vPartition"]
        compare_vms(vinfo_ws, vpart_ws)
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 95% Completed...")

        ## Truncate vPartition Storage Capacity Cols into vInfo
        trunc_cols_vPart(destfile_name)
        #label2.config(text=f"Working... 98% Completed...")

        ## Remove Formatting and Excess Columns on vSummary sheet
        destfile = xl.load_workbook(destfile_name)
        removeFormatting(destfile)
        trimvSum1(destfile)
        consol_vSum(destfile)
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 99% Completed...")
        
        ## Filter First Row
        destfile = xl.load_workbook(destfile_name)
        filter_rows(destfile)
        destfile.save(destfile_name)
        #label2.config(text=f"Working... 100% Completed...")
      except Exception as e:
        label2.config(text=f"Source File import aborted. Analysis canceled.\n \nError: {e}")
      else:
        label2.config(text=f"Success!\n \nRVTools Analysis Completed\n \nAnalyzed file will be new file appended with -EDITED and saved in same directory as Source File\n \nYou may now click the Exit Button to close this Program, or choose another file to analyze")
    except Exception as e:
      label2.config(text=f"Error selecting file: {str(e)}")


if __name__ == "__main__":
    main()