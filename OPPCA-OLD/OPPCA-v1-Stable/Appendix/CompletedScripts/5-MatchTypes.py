import os
import sys
import openpyxl as xl
import tkinter as tk
from tkinter import filedialog

def main():
    # Get the Excel file name from the user
    root = tk.Tk()
    root.withdraw()

    srcfile_name = filedialog.askopenfilename()

    # Open the Excel file
    srcfile = xl.load_workbook(srcfile_name)
    #print(srcfile.sheetnames)

    # Create a new Excel file
    destfile_name = srcfile_name[:-5] + "-EDITED.xlsx"
    srcfile.save(destfile_name)

    # Open new Excel file and remove extra sheets
    destfile = xl.load_workbook(destfile_name)
    #print(destfile.sheetnames)
    
    keep_sheets = ['vInfo', 'vDisk', 'vPartition']
    for sheetName in destfile.sheetnames:
        if sheetName not in keep_sheets:
            del destfile[sheetName]
    destfile.save(destfile_name)

    # Remove Formatting
    def removeFormatting(destfile):
      """Removes all formatting from an excel workbook.

      Args:
        destfile: An xl workbook object.
      """

      for worksheet in destfile.worksheets:
        for cell in worksheet.iter_rows():
          for c in cell:
            c.style = 'Normal'

    destfile = xl.load_workbook(destfile_name)
    removeFormatting(destfile)
    destfile.save(destfile_name)

    # Add Columns
    destfile = xl.load_workbook(destfile_name)

    for worksheet in destfile.worksheets:
      worksheet.insert_cols(2, 6)

    destfile.save(destfile_name)

    # Rename columns B:G in all sheets to "IsFile", "IsSQL", "IsOrcl", "IsPGres", "IsExch", and "IsTestDev" 
    destfile = xl.load_workbook(destfile_name)

    for worksheet in destfile.worksheets:
      worksheet['B1'] = 'IsFile'
      worksheet['C1'] = 'IsSQL'
      worksheet['D1'] = 'IsOrcl'
      worksheet['E1'] = 'IsPGres'
      worksheet['F1'] = 'IsExch'
      worksheet['G1'] = 'IsTestDev'

    # Insert "HasTools" Column in vInfo Sheet
    destfile["vInfo"].insert_cols(8, 1)
    destfile["vInfo"]['H1'] = 'HasTools'

    # Insert "DiskCount" Column in vDisk Sheet
    destfile["vDisk"].insert_cols(8, 1)
    destfile["vDisk"]['H1'] = 'DiskCount'

    destfile.save(destfile_name)

    # Match File Servers
    def match_fs(destfile):
      """Matches file servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      fs_str = ["file", "fs", "nas", "share", "ftp"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in fs_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 1).value = "Yes"

    # Match SQL DBs
    def match_sql(destfile):
      """Matches SQL DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      sql_str = ["sql"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in sql_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 2).value = "Yes"

    # Match Oracle DBs
    def match_orcl(destfile):
      """Matches Oracle DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      orcl_str = ["orcl", "oracle"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in orcl_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 3).value = "Yes"

    # Match PostGres DBs
    def match_pgres(destfile):
      """Matches PostGres DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      pgres_str = ["pgres", "postgres"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in pgres_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 4).value = "Yes"

    # Match Possible DBs
    def match_gendb(destfile):
      """Matches Possible DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      gendb_str = ["db", "database"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in gendb_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 2).value = "Check"
            cell.offset(0, 3).value = "Check"
            cell.offset(0, 4).value = "Check"

    # Match Exchange Servers
    def match_exch(destfile):
      """Matches Exchange Servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      exch_str = ["exch", "exchange"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in exch_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 5).value = "Yes"

    # Match TestDev
    def match_tstdev(destfile):
      """Matches TestDev in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      tstdev_str = ["tst", "test", "dev"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in tstdev_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 6).value = "Yes"

    # Set Used Range of Column A
    def get_used_range(worksheet):
      """Gets the used range of a worksheet.

      Args:
        worksheet: An xl worksheet object.

      Returns:
        A tuple of two cell objects, representing the start and end of the used range of the worksheet.
      """

      # Get the first and last row numbers in the worksheet.
      first_row_number = worksheet.min_row
      last_row_number = worksheet.max_row

      # Iterate over the cells in column A and find the first non-empty cell.
      start_cell = None
      for row_index in range(first_row_number, last_row_number + 1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          start_cell = cell
          break

      # If the first non-empty cell in column A is None, then the used range of the worksheet is empty.
      if start_cell is None:
        return None

      # Find the last non-empty cell in column A.
      end_cell = None
      for row_index in range(last_row_number, first_row_number - 1, -1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          end_cell = cell
          break

      # Return the start and end cells of the used range of the worksheet.
      return start_cell, end_cell

    def set_no_values(destfile, used_range_a):
      """Sets the value of all empty cells in columns B to G to "No", based on the used range of column A.

      Args:
        destfile: An openpyxl workbook object.
        used_range_a: A tuple of two cell objects, representing the start and end of the used range of column A.
      """

      # Get the start and end row numbers of the used range of column A.
      start_row_number = used_range_a[0].row
      end_row_number = used_range_a[1].row

      # Iterate over the rows in the used range of column A.
      for row_index in range(start_row_number, end_row_number + 1):

        # Iterate over the cells in columns B to G.
        for column_index in range(2, 8):

          # Get the cell object.
          cell = destfile.active.cell(row=row_index, column=column_index)

          # If the cell value is empty, then set it to "No".
          if cell.value is None:
            cell.value = "No"

    destfile = xl.load_workbook(destfile_name)
    match_fs(destfile)
    match_sql(destfile)
    match_orcl(destfile)
    match_pgres(destfile)
    match_gendb(destfile)
    match_exch(destfile)
    match_tstdev(destfile)
    for sheetname in destfile.sheetnames:
      destfile.active = destfile[sheetname]
      used_range_a = get_used_range(destfile.active)
      set_no_values(destfile, used_range_a)
    destfile.save(destfile_name)

if __name__ == "__main__":
    main()