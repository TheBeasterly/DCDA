import os
import sys
import openpyxl as xl
import tkinter as tk
from tkinter import filedialog

def main():
    ## Get the Excel file name from the user
    root = tk.Tk()
    root.withdraw()

    srcfile_name = filedialog.askopenfilename()

    ## Open the Excel file
    srcfile = xl.load_workbook(srcfile_name)

    ## Create a new Excel file
    destfile_name = srcfile_name[:-5] + "-EDITED.xlsx"
    srcfile.save(destfile_name)

    ## Open new Excel file and remove extra sheets
    destfile = xl.load_workbook(destfile_name)
    
    keep_sheets = ['vInfo', 'vDisk', 'vPartition']
    for sheetName in destfile.sheetnames:
        if sheetName not in keep_sheets:
            del destfile[sheetName]
    destfile.save(destfile_name)

    ## Remove Formatting
    def removeFormatting(destfile):
      """Removes all formatting from an excel workbook.

      Args:
        destfile: An xl workbook object.
      """

      for worksheet in destfile.worksheets:
        for cell in worksheet.iter_rows():
          for c in cell:
            c.style = 'Normal'

    destfile = xl.load_workbook(destfile_name)
    removeFormatting(destfile)
    destfile.save(destfile_name)

    ## Add Columns
    destfile = xl.load_workbook(destfile_name)

    for worksheet in destfile.worksheets:
      worksheet.insert_cols(2, 6)

    destfile.save(destfile_name)

    ## Rename columns B:G in all sheets to "IsFile", "IsSQL", "IsOrcl", "IsPGres", "IsExch", and "IsTestDev" 
    destfile = xl.load_workbook(destfile_name)

    for worksheet in destfile.worksheets:
      worksheet['B1'] = 'IsFile'
      worksheet['C1'] = 'IsSQL'
      worksheet['D1'] = 'IsOrcl'
      worksheet['E1'] = 'IsPGres'
      worksheet['F1'] = 'IsExch'
      worksheet['G1'] = 'IsTestDev'

    # Insert "HasTools" Column in vInfo Sheet
    destfile["vInfo"].insert_cols(8, 1)
    destfile["vInfo"]['H1'] = 'HasTools'

    # Insert "DiskCount" Column in vDisk Sheet
    destfile["vDisk"].insert_cols(8, 1)
    destfile["vDisk"]['H1'] = 'DiskCount'

    destfile.save(destfile_name)

    ## Match File Servers
    def match_fs(destfile):
      """Matches file servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      fs_str = ["file", "fs", "nas", "share", "ftp"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in fs_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 1).value = "Yes"

    ## Match SQL DBs
    def match_sql(destfile):
      """Matches SQL DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      sql_str = ["sql"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in sql_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 2).value = "Yes"

    ## Match Oracle DBs
    def match_orcl(destfile):
      """Matches Oracle DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      orcl_str = ["orcl", "oracle"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in orcl_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 3).value = "Yes"

    ## Match PostGres DBs
    def match_pgres(destfile):
      """Matches PostGres DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      pgres_str = ["pgres", "postgres"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in pgres_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 4).value = "Yes"

    ## Match Possible DBs
    def match_gendb(destfile):
      """Matches Possible DBs in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      gendb_str = ["db", "database"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in gendb_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 2).value = "Check"
            cell.offset(0, 3).value = "Check"
            cell.offset(0, 4).value = "Check"

    ## Match Exchange Servers
    def match_exch(destfile):
      """Matches Exchange Servers in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      exch_str = ["exch", "exchange"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in exch_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 5).value = "Yes"

    ## Match TestDev
    def match_tstdev(destfile):
      """Matches TestDev in all sheets.

      Args:
        destfile: An xl workbook object.
      """

      tstdev_str = ["tst", "test", "dev"]

      for worksheet in destfile.worksheets:
        for cell in worksheet["A:A"]:
          match_found = False
          for search_string in tstdev_str:
            if search_string.lower() in cell.value.lower():
              match_found = True
              break

          if match_found:
            cell.offset(0, 6).value = "Yes"

    ## Set No Values
    # Set Used Range of Column A
    def get_used_range(worksheet):
      """Gets the used range of a worksheet.

      Args:
        worksheet: An xl worksheet object.

      Returns:
        A tuple of two cell objects, representing the start and end of the used range of the worksheet.
      """

      # Get the first and last row numbers in the worksheet.
      first_row_number = worksheet.min_row
      last_row_number = worksheet.max_row

      # Iterate over the cells in column A and find the first non-empty cell.
      start_cell = None
      for row_index in range(first_row_number, last_row_number + 1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          start_cell = cell
          break

      # If the first non-empty cell in column A is None, then the used range of the worksheet is empty.
      if start_cell is None:
        return None

      # Find the last non-empty cell in column A.
      end_cell = None
      for row_index in range(last_row_number, first_row_number - 1, -1):
        cell = worksheet["A{}".format(row_index)]
        if cell.value is not None:
          end_cell = cell
          break

      # Return the start and end cells of the used range of the worksheet.
      return start_cell, end_cell

    #Set No Values in columns B through G
    def set_no_values(destfile, used_range_a):
      """Sets the value of all empty cells in columns B to G to "No", based on the used range of column A.

      Args:
        destfile: An openpyxl workbook object.
        used_range_a: A tuple of two cell objects, representing the start and end of the used range of column A.
      """

      # Get the start and end row numbers of the used range of column A.
      start_row_number = used_range_a[0].row
      end_row_number = used_range_a[1].row

      # Iterate over the rows in the used range of column A.
      for row_index in range(start_row_number, end_row_number + 1):

        # Iterate over the cells in columns B to G.
        for column_index in range(2, 8):

          # Get the cell object.
          cell = destfile.active.cell(row=row_index, column=column_index)

          # If the cell value is empty, then set it to "No".
          if cell.value is None:
            cell.value = "No"

    ## Insert Columns for MiB to GB Math
    # Find "Provisioned MiB" on vInfo Sheet
    def vinfo_findprovmib(vinfo_ws):
      """Finds the cell in row A of the "vInfo" worksheet that contains the string "Provisioned MiB".

      Args:
        vinfo_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Provisioned MiB", or None if the string is not found.
      """

      for row in vinfo_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Provisioned MiB":
            return cell

      return None

    # Insert Column after "Provisioned MiB" on vInfo Sheet
    def vinfo_provmib_inscol(vinfo_ws, vinfoprovmib_cell):
      """Inserts a new column to the right of the specified cell in the "vInfo" worksheet.

      Args:
        vinfo_ws: An xl worksheet object.
        vinfoprovmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vinfoprovmib_cell.column
      vinfo_ws.insert_cols(column_index + 1)

      # Set the value of the cell in row A of the newly added column to "Provisioned GB".
      vinfo_ws.cell(row=1, column=column_index + 1).value = "Provisioned GB"

    # Find "In Use MiB" on vInfo Sheet
    def vinfo_findinusemib(vinfo_ws):
      """Finds the cell in row A of the "vInfo" worksheet that contains the string "In Use MiB".

      Args:
        vinfo_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "In Use MiB", or None if the string is not found.
      """

      for row in vinfo_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "In Use MiB":
            return cell

      return None

    # Insert Column after "In Use MiB" on vInfo Sheet
    def vinfo_inusemib_inscol(vinfo_ws, vinfoinusemib_cell):
      """Inserts a new column to the right of the specified cell in the "vInfo" worksheet.

      Args:
        vinfo_ws: An xl worksheet object.
        vinfoinusemib_cell: The cell to insert the new column to the right of.
      """

      column_index = vinfoinusemib_cell.column
      vinfo_ws.insert_cols(column_index + 1)

      # Set the value of the cell in row A of the newly added column to "In Use GB".
      vinfo_ws.cell(row=1, column=column_index + 1).value = "In Use GB"

    # Find "Capacity MiB" on vDisk Sheet
    def vdisk_findcapacitymib(vdisk_ws):
      """Finds the cell in row A of the "vDisk" worksheet that contains the string "Capacity MiB".

      Args:
        vdisk_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Capacity MiB", or None if the string is not found.
      """

      for row in vdisk_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Capacity MiB":
            return cell

      return None

    # Insert Column after "Capacity MiB" on vDisk Sheet
    def vdisk_capacitymib_inscol(vdisk_ws, vdiskcapmib_cell):
      """Inserts a new column to the right of the specified cell in the "vDisk" worksheet.

      Args:
        vdisk_ws: An xl worksheet object.
        vdiskcapmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vdiskcapmib_cell.column
      vdisk_ws.insert_cols(column_index + 1)

      # Set the value of the cell in row A of the newly added column to "Capacity GB".
      vdisk_ws.cell(row=1, column=column_index + 1).value = "Capacity GB"

    # Find "Capacity MiB" on vPartition Sheet
    def vpart_findcapacitymib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Capacity MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Capacity MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Capacity MiB":
            return cell

      return None

    # Insert Column after "Capacity MiB" on vPartition Sheet
    def vpart_capacitymib_inscol(vpart_ws, vpartcapmib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartcapmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartcapmib_cell.column
      vpart_ws.insert_cols(column_index + 1)

      # Set the value of the cell in row A of the newly added column to "Capacity GB".
      vpart_ws.cell(row=1, column=column_index + 1).value = "Capacity GB"

    # Find "Consumed MiB" on vPartition Sheet
    def vpart_findconsumedmib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Consumed MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Consumed MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Consumed MiB":
            return cell

      return None

    # Insert Column after "Consumed MiB" on vPartition Sheet
    def vpart_consumedmib_inscol(vpart_ws, vpartconsmib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartconsmib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartconsmib_cell.column
      vpart_ws.insert_cols(column_index + 1)

      # Set the value of the cell in row A of the newly added column to "Consumed GB".
      vpart_ws.cell(row=1, column=column_index + 1).value = "Consumed GB"

    # Find "Free MiB" on vPartition Sheet
    def vpart_findfreemib(vpart_ws):
      """Finds the cell in row A of the "vPartition" worksheet that contains the string "Free MiB".

      Args:
        vpart_ws: An xl worksheet object.

      Returns:
        The cell object in row A that contains the string "Free MiB", or None if the string is not found.
      """

      for row in vpart_ws.iter_cols(min_row=1, max_row=1):
        for cell in row:
          if cell.value == "Free MiB":
            return cell

      return None

    # Insert Column after "Free MiB" on vPartition Sheet
    def vpart_freemib_inscol(vpart_ws, vpartfreemib_cell):
      """Inserts a new column to the right of the specified cell in the "vPartition" worksheet.

      Args:
        vpart_ws: An xl worksheet object.
        vpartfreemib_cell: The cell to insert the new column to the right of.
      """

      column_index = vpartfreemib_cell.column
      vpart_ws.insert_cols(column_index + 1)

      # Set the value of the cell in row A of the newly added column to "Free GB".
      vpart_ws.cell(row=1, column=column_index + 1).value = "Free GB"

    ## Match Workload Types
    destfile = xl.load_workbook(destfile_name)
    match_fs(destfile)
    match_sql(destfile)
    match_orcl(destfile)
    match_pgres(destfile)
    match_gendb(destfile)
    match_exch(destfile)
    match_tstdev(destfile)
    for sheetname in destfile.sheetnames:
      destfile.active = destfile[sheetname]
      used_range_a = get_used_range(destfile.active)
      set_no_values(destfile, used_range_a)
    destfile.save(destfile_name)

    ## Insert Columns for MiB to GB math
    destfile = xl.load_workbook(destfile_name)
    vinfo_ws = destfile["vInfo"]
    vdisk_ws = destfile["vDisk"]
    vpart_ws = destfile["vPartition"]

    # Insert Columns in vInfo Sheet
    vinfoprovmib_cell = vinfo_findprovmib(vinfo_ws)
    if vinfoprovmib_cell is not None:
      vinfo_provmib_inscol(vinfo_ws, vinfoprovmib_cell)
    vinfoinusemib_cell = vinfo_findinusemib(vinfo_ws)
    if vinfoinusemib_cell is not None:
      vinfo_inusemib_inscol(vinfo_ws, vinfoinusemib_cell)

    # Insert Columns in vDisk Sheet
    vdiskcapmib_cell = vdisk_findcapacitymib(vdisk_ws)
    if vdiskcapmib_cell is not None:
      vdisk_capacitymib_inscol(vdisk_ws, vdiskcapmib_cell)

    # Insert Columns in vPartition Sheet
    vpartcapmib_cell = vpart_findcapacitymib(vpart_ws)
    if vpartcapmib_cell is not None:
      vpart_capacitymib_inscol(vpart_ws, vpartcapmib_cell)
    vpartconsmib_cell = vpart_findconsumedmib(vpart_ws)
    if vpartconsmib_cell is not None:
      vpart_consumedmib_inscol(vpart_ws, vpartconsmib_cell)
    vpartfreemib_cell = vpart_findfreemib(vpart_ws)
    if vpartfreemib_cell is not None:
      vpart_freemib_inscol(vpart_ws, vpartfreemib_cell)

    destfile.save(destfile_name)



if __name__ == "__main__":
    main()