# RVToolsAnalysis_web.py
import os
import sys # Keep for pandas/openpyxl interaction if needed
import openpyxl as xl
from openpyxl.styles import Font, Border, PatternFill, Protection, Alignment
from openpyxl.utils import get_column_letter # Used in filter_rows
import pandas as pd
import traceback
import time
import json

# =================================================================================================================================================================================================================================================
# ============================================================================================= HELPER FUNCTIONS (Based on User's Uploaded Code) ============================================================================================
# =================================================================================================================================================================================================================================================

# Note: Removed internal print statements from helpers as progress is yielded by main func

## Remove Formatting
def removeFormatting(destfile):
    """Removes all formatting from an excel workbook."""
    try:
        for worksheet in destfile.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.has_style:
                        cell.font = Font()
                        cell.border = Border()
                        cell.fill = PatternFill(fill_type=None) # Explicitly set fill_type
                        cell.number_format = 'General'
                        cell.protection = Protection()
                        cell.alignment = Alignment()
    except Exception as e:
        print(f"SERVER WARNING: Error during removeFormatting: {e}") # Log detailed error

## Match File Servers
def match_fs(destfile):
    """Matches file servers in all sheets based on VM name in Column A."""
    fs_str = ["file", "fs", "nas", "share", "ftp"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in fs_str):
                    if worksheet.max_column >= 3: worksheet.cell(row=cell.row, column=3).value = "Yes"
                    else: print(f"SERVER WARNING: Col C (IsFile) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: print(f"SERVER ERROR: match_fs: {e}"); traceback.print_exc()

## Match SQL DBs
def match_sql(destfile):
    """Matches SQL DBs in all sheets based on VM name in Column A."""
    sql_str = ["sql"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in sql_str):
                    if worksheet.max_column >= 4: worksheet.cell(row=cell.row, column=4).value = "Yes"
                    else: print(f"SERVER WARNING: Col D (IsSQL) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: print(f"SERVER ERROR: match_sql: {e}"); traceback.print_exc()

## Match Oracle DBs
def match_orcl(destfile):
    """Matches Oracle DBs in all sheets based on VM name in Column A."""
    orcl_str = ["orcl", "oracle"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in orcl_str):
                    if worksheet.max_column >= 5: worksheet.cell(row=cell.row, column=5).value = "Yes"
                    else: print(f"SERVER WARNING: Col E (IsOrcl) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: print(f"SERVER ERROR: match_orcl: {e}"); traceback.print_exc()

## Match PostGres DBs
def match_pgres(destfile):
    """Matches PostGres DBs in all sheets based on VM name in Column A."""
    pgres_str = ["pgres", "postgres"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in pgres_str):
                    if worksheet.max_column >= 6: worksheet.cell(row=cell.row, column=6).value = "Yes"
                    else: print(f"SERVER WARNING: Col F (IsPGres) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: print(f"SERVER ERROR: match_pgres: {e}"); traceback.print_exc()

## Match Possible DBs (General - Sets "Check")
def match_gendb(destfile):
    """Matches Possible DBs; sets 'Check' if not already 'Yes'."""
    gendb_str = ["db", "database"]
    db_check_cols = [4, 5, 6] # D=IsSQL, E=IsOrcl, F=IsPGres
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in gendb_str):
                    for col_idx in db_check_cols:
                        if worksheet.max_column >= col_idx:
                            target_cell = worksheet.cell(row=cell.row, column=col_idx)
                            if target_cell.value != "Yes": target_cell.value = "Check"
                        # else: print(f"SERVER WARNING: Col {col_idx} missing sheet '{worksheet.title}' row {cell.row} for gendb")
    except Exception as e: print(f"SERVER ERROR: match_gendb: {e}"); traceback.print_exc()

## Match Exchange Servers
def match_exch(destfile):
    """Matches Exchange Servers in all sheets based on VM name in Column A."""
    exch_str = ["exch", "exchange"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in exch_str):
                    if worksheet.max_column >= 7: worksheet.cell(row=cell.row, column=7).value = "Yes"
                    else: print(f"SERVER WARNING: Col G (IsExch) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: print(f"SERVER ERROR: match_exch: {e}"); traceback.print_exc()

## Match TestDev
def match_tstdev(destfile):
    """Matches Test/Dev systems in all sheets based on VM name in Column A."""
    tstdev_str = ["tst", "test", "dev"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in tstdev_str):
                    if worksheet.max_column >= 8: worksheet.cell(row=cell.row, column=8).value = "Yes"
                    else: print(f"SERVER WARNING: Col H (IsTestDev) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: print(f"SERVER ERROR: match_tstdev: {e}"); traceback.print_exc()

## Get Last Row Helper
def get_last_row_in_col_a(worksheet):
    """Finds the last row index in Column A that contains data."""
    last_row = 0
    try:
        # Iterate backwards from max_row if possible and seems reasonable
        start_check_row = min(worksheet.max_row, 1048576) # Avoid huge empty sheet checks
        for row_idx in range(start_check_row, 0, -1):
            cell = worksheet.cell(row=row_idx, column=1)
            if cell.value is not None and str(cell.value).strip() != "":
                last_row = row_idx; break
        # Fallback if max_row was 0 or misleadingly small
        if last_row == 0 and worksheet.calculate_dimension() != 'A1': # Check if sheet isn't just A1
             for row_idx in range(1, start_check_row + 1):
                 cell = worksheet.cell(row=row_idx, column=1)
                 if cell.value is not None and str(cell.value).strip() != "":
                     last_row = max(last_row, row_idx)
    except Exception as e: print(f"SERVER WARNING: Error getting last row A: {e}")
    return last_row if last_row > 0 else 1 # Return at least 1 if header exists

## Set No Values
def set_no_values(destfile):
    """Sets empty cells in columns C to H to 'No'."""
    try:
        for ws_name in ['vInfo', 'vDisk', 'vPartition']:
            if ws_name in destfile.sheetnames:
                worksheet = destfile[ws_name]
                if worksheet.max_row <= 1: continue
                last_row_a = get_last_row_in_col_a(worksheet)
                for row_index in range(2, last_row_a + 1):
                    for column_index in range(3, 9): # Columns C to H
                        if worksheet.max_column >= column_index:
                            cell = worksheet.cell(row=row_index, column=column_index)
                            if cell.value is None or str(cell.value).strip() == "":
                                cell.value = "No"
            else: print(f"SERVER WARNING: Sheet '{ws_name}' not found for set_no_values.")
    except Exception as e: print(f"SERVER ERROR: set_no_values: {e}"); traceback.print_exc()

## Insert GB Columns Helpers (vInfo, vDisk, vPartition)
def find_col_idx(ws, header_name):
    """Finds column index (1-based) for header."""
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == header_name: return col_idx
    return None

def insert_gb_col(ws, mib_header, gb_header, sheet_name_for_warn):
    """Generic helper to insert and calculate GB columns."""
    mib_col_idx = find_col_idx(ws, mib_header)
    if mib_col_idx is None: print(f"SERVER WARNING: Col '{mib_header}' not found in {sheet_name_for_warn}."); return False
    gb_col_idx = mib_col_idx + 1; ws.insert_cols(gb_col_idx)
    ws.cell(row=1, column=gb_col_idx).value = gb_header
    for i in range(2, ws.max_row + 1):
        mib_cell = ws.cell(row=i, column=mib_col_idx); gb_cell = ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = pd.to_numeric(mib_cell.value, errors='coerce') # Handle non-numeric gracefully
            if pd.notna(mib_value):
                gb_value = round(mib_value / 953.7, 2)
                gb_cell.value = gb_value; gb_cell.number_format = '0.00'
            else: gb_cell.value = None # Set to blank if source not numeric
        except Exception: gb_cell.value = None # Catch other potential errors
    return True

def vinfo_insert_gb_cols(ws):
    insert_gb_col(ws, "Provisioned MiB", "Provisioned GB", "vInfo")
    insert_gb_col(ws, "In Use MiB", "In Use GB", "vInfo")
def vdisk_insert_gb_cols(ws):
    insert_gb_col(ws, "Capacity MiB", "Capacity GB", "vDisk")
def vpart_insert_gb_cols(ws):
    insert_gb_col(ws, "Capacity MiB", "Capacity GB", "vPartition")
    insert_gb_col(ws, "Consumed MiB", "Consumed GB", "vPartition")
    insert_gb_col(ws, "Free MiB", "Free GB", "vPartition")

## Set Disk Count Value
def vdisk_diskcount_val(vdisk_ws):
    """Sets DiskCount column to 1."""
    diskcount_col_idx = find_col_idx(vdisk_ws, "DiskCount")
    if diskcount_col_idx is None: print("SERVER WARNING: 'DiskCount' column not found in vDisk."); return False
    for i in range(2, vdisk_ws.max_row + 1): vdisk_ws.cell(row=i, column=diskcount_col_idx).value = 1
    return True

## Compare VMs for HasTools
def compare_vms(vinfo_ws, vpart_ws):
  """Compares VMs in vInfo to vPartition, sets HasTools."""
  if vinfo_ws is None or vpart_ws is None: print("SERVER ERROR: Missing sheet for compare_vms."); return
  try:
      vpart_vms = set()
      if vpart_ws.max_row > 1:
          for cell in vpart_ws['A'][1:]: # Skip header
              if cell.value is not None and str(cell.value).strip(): vpart_vms.add(str(cell.value))
      print(f"SERVER INFO: Found {len(vpart_vms)} vPart VMs for compare.")
      hastools_col_idx = 9 # Assume Column I
      if vinfo_ws.max_column < hastools_col_idx: print(f"SERVER WARNING: 'HasTools' column missing in vInfo."); return
      if vinfo_ws.max_row > 1:
          for row_idx in range(2, vinfo_ws.max_row + 1):
              vinfo_vm_cell = vinfo_ws.cell(row=row_idx, column=1); hastools_cell = vinfo_ws.cell(row=row_idx, column=hastools_col_idx)
              vm_name = str(vinfo_vm_cell.value) if vinfo_vm_cell.value is not None else ""
              hastools_cell.value = "Yes" if vm_name and vm_name in vpart_vms else "No"
  except Exception as e: print(f"SERVER ERROR: compare_vms: {e}"); traceback.print_exc()

## Delete Unnecessary Columns (Pandas)
def del_cols_vInfo(destfile_path):
    """Deletes unnecessary columns using Pandas."""
    try:
        keep_cols = { # Same as before
            'vInfo': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'HasTools', 'Disks', 'Total disk capacity', 'Provisioned MiB', 'Provisioned GB', 'In Use MiB', 'In Use GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vDisk': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'DiskCount', 'Disk', 'Capacity MiB', 'Capacity GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vPartition': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'Disk', 'Capacity MiB', 'Capacity GB', 'Consumed MiB', 'Consumed GB', 'Free MiB', 'Free GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']
        }
        all_sheets_df = {}
        try: # Read only expected sheets
            xls = pd.ExcelFile(destfile_path, engine='openpyxl')
            for sheet_name in ['vInfo', 'vDisk', 'vPartition']:
                if sheet_name in xls.sheet_names: all_sheets_df[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
            xls.close()
        except Exception as ex_read_err: print(f"SERVER ERROR: Reading for del_cols: {ex_read_err}"); raise
        if not all_sheets_df: raise ValueError("No sheets found for del_cols.")

        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in all_sheets_df.items():
                if sheet_name in keep_cols:
                    cols_to_keep = [col for col in keep_cols[sheet_name] if col in df.columns]
                    if not cols_to_keep: print(f"SERVER WARNING: No valid cols found for sheet '{sheet_name}'. Skipping."); continue
                    df[cols_to_keep].to_excel(writer, sheet_name=sheet_name, index=False)

    except Exception as e: print(f"SERVER ERROR: del_cols_vInfo: {e}"); traceback.print_exc(); raise

## Truncate vPartition to vSummary (Pandas)
def trunc_cols_vPart(destfile_path):
    """Aggregates vPartition data into vSummary using Pandas."""
    try:
        dfs = {}
        try: # Read needed sheets
            xls = pd.ExcelFile(destfile_path, engine='openpyxl')
            for sheet_name in ['vInfo', 'vDisk', 'vPartition']: # Need all 3 to write back
                if sheet_name in xls.sheet_names: dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
            xls.close()
        except Exception as ex_read_err: print(f"SERVER ERROR: Reading for trunc_cols: {ex_read_err}"); raise
        if 'vInfo' not in dfs or 'vPartition' not in dfs: raise ValueError("vInfo or vPartition sheet missing for truncation.")
        tcdf1 = dfs['vInfo']; tcdf3 = dfs['vPartition']

        storage_cols = ['Capacity GB', 'Consumed GB', 'Free GB']
        existing_storage_cols = [col for col in storage_cols if col in tcdf3.columns]
        vSummary_df = tcdf1.copy() # Start with vInfo data

        if existing_storage_cols:
            if 'VM' not in tcdf3.columns: raise ValueError("'VM' col missing in vPartition.");
            if 'VM' not in tcdf1.columns: raise ValueError("'VM' col missing in vInfo.");
            for col in existing_storage_cols: tcdf3[col] = pd.to_numeric(tcdf3[col], errors='coerce').fillna(0)
            aggregated_df = tcdf3.groupby('VM', as_index=False)[existing_storage_cols].sum()
            vSummary_df = pd.merge(tcdf1, aggregated_df, on='VM', how='left')
            for col in existing_storage_cols: # Fill NaNs only for the merged columns
                 if col in vSummary_df.columns: vSummary_df[col].fillna(0, inplace=True)

            # Reorder columns if 'In Use GB' exists
            if 'In Use GB' in vSummary_df.columns:
                try:
                    idx = vSummary_df.columns.get_loc('In Use GB')
                    cols = vSummary_df.columns.tolist()
                    new_order = cols[:idx+1] + existing_storage_cols + [c for c in cols[idx+1:] if c not in existing_storage_cols]
                    vSummary_df = vSummary_df[new_order]
                except KeyError: print("SERVER WARNING: Cannot reorder vSummary columns.")
        else: print("SERVER WARNING: No storage cols found in vPartition to aggregate.")

        # Write all sheets back (vSummary, vInfo, vDisk if exists, vPartition)
        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            vSummary_df.to_excel(writer, sheet_name='vSummary', index=False)
            dfs['vInfo'].to_excel(writer, sheet_name='vInfo', index=False)
            if 'vDisk' in dfs: dfs['vDisk'].to_excel(writer, sheet_name='vDisk', index=False)
            dfs['vPartition'].to_excel(writer, sheet_name='vPartition', index=False)

    except Exception as e: print(f"SERVER ERROR: trunc_cols_vPart: {e}"); traceback.print_exc(); raise

## Trim vSummary Columns (Openpyxl)
def trimvSum1(destfile):
    """Removes specific columns by name from vSummary using openpyxl."""
    if 'vSummary' not in destfile.sheetnames: print("SERVER WARNING: vSummary sheet not found for trimvSum1."); return
    vSum_ws = destfile['vSummary']
    cols_to_delete_names = ['Provisioned MiB', 'In Use MiB']
    try:
        headers = [cell.value for cell in vSum_ws[1]]
        indices_to_delete = []
        for name in cols_to_delete_names:
            try: idx = headers.index(name) + 1; indices_to_delete.append(idx)
            except ValueError: print(f"SERVER WARNING: Column '{name}' not found for trimvSum1.")
        indices_to_delete.sort(reverse=True) # Delete from right to left
        for col_idx in indices_to_delete: vSum_ws.delete_cols(col_idx, 1)
    except Exception as e: print(f"SERVER ERROR: trimvSum1: {e}"); traceback.print_exc()

## Consolidate vSummary Consumed GB (Openpyxl) - Includes fix for 0
def consol_vSum(destfile):
    """Copies 'In Use GB' to 'Consumed GB' if 'Consumed GB' is empty, blank, or zero."""
    if 'vSummary' not in destfile.sheetnames: print("SERVER WARNING: vSummary sheet not found for consol_vSum."); return
    vSum_ws = destfile['vSummary']
    try:
        consumed_gb_col_idx, in_use_gb_col_idx = None, None
        headers = [cell.value for cell in vSum_ws[1]]
        for idx, header in enumerate(headers, 1):
            if header == 'Consumed GB': consumed_gb_col_idx = idx
            elif header == 'In Use GB': in_use_gb_col_idx = idx
        if consumed_gb_col_idx is None or in_use_gb_col_idx is None: print(f"SERVER ERROR: Consolidation columns missing in vSummary."); return

        for row_idx in range(2, vSum_ws.max_row + 1):
            consumed_cell = vSum_ws.cell(row=row_idx, column=consumed_gb_col_idx)
            in_use_cell = vSum_ws.cell(row=row_idx, column=in_use_gb_col_idx)
            consumed_val = consumed_cell.value
            # Check for None, empty string, or numerical 0
            if consumed_val is None or (isinstance(consumed_val, str) and consumed_val.strip() == "") or consumed_val == 0:
                consumed_cell.value = in_use_cell.value
                if in_use_cell.has_style and in_use_cell.number_format: consumed_cell.number_format = in_use_cell.number_format
    except Exception as e: print(f"SERVER ERROR: consol_vSum: {e}"); traceback.print_exc()

## Filter First Row (Openpyxl)
def filter_rows(destfile):
  """Applies auto-filter to the header row on all sheets."""
  try:
    for ws in destfile.worksheets:
      if ws.max_row > 0 and ws.max_column > 0:
          last_col_letter = get_column_letter(ws.max_column)
          ws.auto_filter.ref = ws.dimensions # Apply to data range might be better
          # ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}" # Or include total row? Let's stick to data range.
          print(f"Applied filter to sheet '{ws.title}' range {ws.dimensions}")
  except Exception as e: print(f"SERVER ERROR: filter_rows: {e}"); traceback.print_exc()

# ================================================================================================================================================================================================================================================
# ===================================================================================================== MAIN PROCESSING FUNCTION (SSE GENERATOR) =================================================================================================
# ================================================================================================================================================================================================================================================

def process_rvtools_file(input_filepath, output_folder, original_basename):
    """
    Processes RVTools, yields SSE updates, categorizes, adds summary sheets
    with dynamic SUBTOTAL rows, reorders, and applies filters.
    """
    # --- SSE Helper Functions ---
    def sse_message(data): data = str(data).replace('\n', ' '); return f"data: {data}\n\n"
    def yield_result(success, message): payload = json.dumps({"success": success, "message": message}); yield f"event: result\ndata: {payload}\n\n"

    # --- Categorization Helper Functions ---
    def assign_workload_category(row): # Includes General DB, Multi-Error
        is_values={'SQL':row.get('IsSQL')=='Yes','Exchange':row.get('IsExch')=='Yes','File':row.get('IsFile')=='Yes','Oracle':row.get('IsOrcl')=='Yes','Postgres':row.get('IsPGres')=='Yes'}
        db_check_flags={'SQL_Check':row.get('IsSQL')=='Check','Orcl_Check':row.get('IsOrcl')=='Check','PGres_Check':row.get('IsPGres')=='Check'}
        yes_count=sum(is_values.values())
        if yes_count > 1: return "Multi-Error"
        elif yes_count == 1:
            for w, is_y in is_values.items():
                if is_y: return w
        else: return "General DB" if any(db_check_flags.values()) else "Standard"
        return "Unknown"
    def assign_environment(row): return "Test/Dev" if row.get('IsTestDev') == 'Yes' else "Prod"

    # --- Main Processing Logic ---
    destfile = None; output_filename = None; writer = None; workbook = None
    try:
        # --- Setup Output Path ---
        yield sse_message(f"Starting analysis for: {original_basename}"); time.sleep(0.1)
        if original_basename: base_name_no_ext, ext = os.path.splitext(original_basename); output_filename = f"{base_name_no_ext}-ANALYZED{ext}"
        else: temp_base, ext = os.path.splitext(os.path.basename(input_filepath)); output_filename = f"{temp_base}-ORIGNAMEMISSING-ANALYZED{ext}"
        output_filepath = os.path.join(output_folder, output_filename)
        yield sse_message(f"Output file: {output_filename}"); print(f"Output file: {output_filepath}"); time.sleep(0.1)

        # --- Steps 1-8: Initial Data Prep (Openpyxl) ---
        # ... (Same as before: Load, Format, Add/Rename Cols, Match Workloads, Compare, Fill DiskCount, Calc GB) ...
        yield sse_message("Step 1-8: Performing initial data prep..."); time.sleep(0.1)
        destfile = xl.load_workbook(input_filepath); removeFormatting(destfile);
        keep_sheets=['vInfo', 'vDisk', 'vPartition']; sheets_to_delete=[s for s in destfile.sheetnames if s not in keep_sheets]
        if not any(s in destfile.sheetnames for s in keep_sheets): raise ValueError("Required sheets missing.")
        for sn in sheets_to_delete: del destfile[sn]
        for ws_name in keep_sheets:
             if ws_name in destfile:
                 ws=destfile[ws_name]; ws.insert_cols(3, 6); ws['C1']='IsFile'; ws['D1']='IsSQL'; ws['E1']='IsOrcl'; ws['F1']='IsPGres'; ws['G1']='IsExch'; ws['H1']='IsTestDev'
                 if ws_name=='vInfo': ws.insert_cols(9, 1); ws['I1']='HasTools'
                 elif ws_name=='vDisk': ws.insert_cols(9, 1); ws['I1']='DiskCount'
        match_fs(destfile); match_sql(destfile); match_orcl(destfile); match_pgres(destfile); match_exch(destfile); match_tstdev(destfile); match_gendb(destfile);
        set_no_values(destfile)
        vinfo_ws=destfile['vInfo'] if 'vInfo' in destfile else None; vpart_ws=destfile['vPartition'] if 'vPartition' in destfile else None
        if vinfo_ws and vpart_ws: compare_vms(vinfo_ws, vpart_ws)
        if 'vDisk' in destfile: vdisk_diskcount_val(destfile['vDisk'])
        if vinfo_ws: vinfo_insert_gb_cols(vinfo_ws)
        if 'vDisk' in destfile: vdisk_insert_gb_cols(destfile['vDisk'])
        if vpart_ws: vpart_insert_gb_cols(vpart_ws)
        destfile.save(output_filepath); destfile.close(); destfile=None # Save intermediate before pandas
        yield sse_message(" -> Initial prep complete."); time.sleep(0.1)

        # --- Steps 9-10: Pandas Operations ---
        yield sse_message("Step 9: Removing unused columns (pandas)..."); time.sleep(0.1)
        try: del_cols_vInfo(output_filepath)
        except Exception as e: print(f"Pandas del_cols error: {e}"); raise
        yield sse_message(" -> Unused columns removed."); time.sleep(0.1)
        yield sse_message("Step 10: Creating vSummary sheet (pandas)..."); time.sleep(0.1)
        try: trunc_cols_vPart(output_filepath) # This overwrites file with vSummary, vInfo, vDisk, vPartition
        except Exception as e: print(f"Pandas trunc_cols error: {e}"); raise
        yield sse_message(" -> vSummary sheet created/updated."); time.sleep(0.1)

        # --- Step 11: Openpyxl Final Cleanup (before summary generation) ---
        yield sse_message("Step 11: Reloading for final openpyxl cleanup..."); time.sleep(0.1)
        destfile = xl.load_workbook(output_filepath)
        removeFormatting(destfile) # Remove pandas formatting
        if 'vSummary' in destfile: trimvSum1(destfile); consol_vSum(destfile) # Trim/Consolidate vSummary
        else: print("SERVER WARNING: vSummary not found for final cleanup.")
        # NO filtering here
        destfile.save(output_filepath); # Save cleanup changes
        destfile.close(); destfile=None
        yield sse_message(" -> Intermediate save before summary generation."); time.sleep(0.1)

        # --- Step 13: Generate Summaries (Pandas) & Write Sheets (ExcelWriter) ---
        # (Data Only, NO Static Totals Row)
        yield sse_message("Step 13: Generating analysis summaries..."); time.sleep(0.1)
        summary_sheets_generated = [] # Keep track of sheets successfully generated
        try:
            yield sse_message(" -> Reading vSummary sheet for summaries..."); time.sleep(0.1)
            df_summary_source = pd.read_excel(output_filepath, sheet_name='vSummary', engine='openpyxl')
            yield sse_message(" -> vSummary loaded."); time.sleep(0.1)

            required_cols = ['VM','IsFile','IsSQL','IsOrcl','IsPGres','IsExch','IsTestDev','Disks','Consumed GB','Powerstate','Datacenter','Cluster']
            missing_cols = [c for c in required_cols if c not in df_summary_source.columns]
            if missing_cols: raise ValueError(f"Missing columns in vSummary: {missing_cols}")
            df_summary_source['VM Count'] = 1

            yield sse_message(" -> Categorizing workloads..."); time.sleep(0.1)
            df_summary_source['Workload'] = df_summary_source.apply(assign_workload_category, axis=1)
            df_summary_source['Environment'] = df_summary_source.apply(assign_environment, axis=1)

            # Position new columns before 'IsFile'
            try:
                 isfile_idx = df_summary_source.columns.get_loc('IsFile'); cols = df_summary_source.columns.tolist()
                 if 'Workload' in cols: cols.remove('Workload'); cols.insert(isfile_idx, 'Workload')
                 if 'Environment' in cols: cols.remove('Environment'); cols.insert(isfile_idx + 1, 'Environment')
                 df_summary_source = df_summary_source[cols]
            except KeyError: yield sse_message(" -> Warn: Cannot position category columns.")

            # Filter for PoweredOn VMs
            df_powered_on = df_summary_source[df_summary_source['Powerstate'] == 'poweredOn'].copy()

            # Aggregation definitions
            agg_total = {'VM Count': 'sum', 'Disks': 'sum', 'Consumed GB': 'sum'}
            agg_pon_counts = {'VM Count': 'sum'}
            agg_dc_pon_metrics = {'PoweredOn VMs': ('VM Count', 'sum'), 'PoweredOn Disks': ('Disks', 'sum'), 'PoweredOn Consumed GB': ('Consumed GB', 'sum')}

            # Calculate Summaries
            yield sse_message(" -> Calculating summaries..."); time.sleep(0.1)
            summary_overall_total = df_summary_source.groupby(['Workload', 'Environment'], observed=False).agg(agg_total).reset_index()
            summary_overall_pon_counts = df_summary_source.groupby(['Workload', 'Environment', 'Powerstate'], observed=False).agg(agg_pon_counts).reset_index()
            summary_dc_base = df_summary_source.groupby(['Datacenter', 'Cluster', 'Workload', 'Environment'], observed=False).agg(agg_total).reset_index()
            summary_dc_pon_metrics = df_powered_on.groupby(['Datacenter', 'Cluster', 'Workload', 'Environment'], observed=False).agg(**agg_dc_pon_metrics).reset_index()
            summary_dc_combined = pd.merge(summary_dc_base, summary_dc_pon_metrics, on=['Datacenter', 'Cluster', 'Workload', 'Environment'], how='left').fillna(0)

            # --- Write DataFrames (WITHOUT total rows) to sheets ---
            yield sse_message(" -> Writing data to summary sheets..."); time.sleep(0.1)
            summary_sheets_to_write = {
                "Overall Summary Totals": summary_overall_total,
                "Overall Powerstate Counts": summary_overall_pon_counts,
                "Datacenter Summary Combined": summary_dc_combined,
                "vSummary": df_summary_source # Overwrite vSummary with categorized version
            }
            # Use append mode to add/replace sheets without disturbing others (like vInfo)
            with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                for sheet_name, df_to_write in summary_sheets_to_write.items():
                     if not df_to_write.empty:
                         df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
                         summary_sheets_generated.append(sheet_name) # Track generated sheets
            yield sse_message(f" -> {len(summary_sheets_generated)} summary sheets written."); time.sleep(0.1)

        except Exception as summary_err:
            # Log error and re-raise to stop processing if summary gen fails critically
            print(f"SERVER ERROR: Generating summary dataframes/writing sheets: {summary_err}")
            traceback.print_exc()
            yield sse_message(f"ERROR: Failed to generate summary sheets: {summary_err}"); time.sleep(0.1)
            raise # Stop processing

        # --- *** NEW Step 14: Add Formulas, Reorder Sheets & Apply Filters (Openpyxl) *** ---
        yield sse_message("Step 14: Adding totals, reordering sheets, applying filters..."); time.sleep(0.1)
        try:
            workbook = xl.load_workbook(output_filepath) # Load final workbook

            # --- Add SUBTOTAL Formulas ---
            sheets_for_subtotals = { # Map sheet name to numeric columns and label column index
                'Overall Summary Totals': {'cols': [3, 4, 5], 'label_col': 1}, # VM Count, Disks, Consumed GB; Label in Workload col
                'Overall Powerstate Counts': {'cols': [4], 'label_col': 1}, # VM Count; Label in Workload col
                'Datacenter Summary Combined': {'cols': [5, 6, 7, 8, 9, 10], 'label_col': 1} # VM Count -> PON Consumed GB; Label in DC col
            }
            total_row_font = Font(bold=True) # Define bold font for total row

            for sheet_name, config in sheets_for_subtotals.items():
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    max_row = ws.max_row
                    # Only add total row if there's data beyond the header
                    if max_row > 1:
                        total_row_idx = max_row + 1
                        # Add "Grand Total" label
                        label_cell = ws.cell(row=total_row_idx, column=config['label_col'])
                        label_cell.value = "Grand Total"
                        label_cell.font = total_row_font
                        # Add SUBTOTAL formulas
                        for col_idx in config['cols']:
                             col_letter = get_column_letter(col_idx)
                             # Formula: =SUBTOTAL(9, C2:C<max_row>) where 9=SUM
                             formula = f"=SUBTOTAL(9,{col_letter}2:{col_letter}{max_row})"
                             total_cell = ws.cell(row=total_row_idx, column=col_idx)
                             total_cell.value = formula
                             total_cell.font = total_row_font
                             # Try to copy number format from cell above if possible
                             if max_row >= 2:
                                 prev_cell = ws.cell(row=max_row, column=col_idx)
                                 if prev_cell.number_format:
                                     total_cell.number_format = prev_cell.number_format
                        yield sse_message(f" -> Added SUBTOTAL row to '{sheet_name}'."); time.sleep(0.1)
                else:
                     yield sse_message(f" -> Sheet '{sheet_name}' not found, skipping SUBTOTAL row."); time.sleep(0.1)

            # --- Reorder Sheets ---
            desired_order = [ # Sheets to appear first
                'Overall Summary Totals', 'Overall Powerstate Counts', 'Datacenter Summary Combined',
                'vSummary', 'vInfo', 'vDisk', 'vPartition' # Original sheets after
            ]
            current_sheets = workbook.sheetnames
            final_order = [sheet for sheet in desired_order if sheet in current_sheets]
            # Add any sheets not in desired_order (e.g., if vDisk was missing) to the end
            final_order += [sheet for sheet in current_sheets if sheet not in final_order]
            workbook._sheets = sorted(workbook._sheets, key=lambda ws: final_order.index(ws.title))
            yield sse_message(" -> Sheets reordered."); time.sleep(0.1)

            # --- Apply Filters ---
            filter_rows(workbook) # Apply to headers of all sheets
            yield sse_message(" -> Filters applied."); time.sleep(0.1)

            # --- Final Save ---
            workbook.save(output_filepath)
            yield sse_message(" -> Final workbook saved."); time.sleep(0.1)

        except Exception as final_step_err:
             print(f"SERVER ERROR: During final SUBTOTAL/reorder/filter/save: {final_step_err}")
             traceback.print_exc()
             yield sse_message(f"WARNING: Failed during final workbook adjustments: {final_step_err}"); time.sleep(0.1)
             # Allow to proceed to final yield, main data might still be okay
        finally:
             if workbook: workbook.close() # Ensure workbook is closed

        # --- Final Success Yield ---
        yield sse_message("--- Analysis File Generation Complete ---")
        print(f"--- Analysis Completed: {original_basename} -> {output_filename} ---")
        yield from yield_result(True, output_filename)

    # --- Global Error Handling ---
    except Exception as e:
        # Ensure workbook handles are closed on error
        if destfile:
            try: destfile.close()
            except Exception as close_err: print(f"SERVER WARNING: Error closing destfile: {close_err}")
        if 'writer' in locals() and writer is not None and writer.book is not None: # Close pandas writer if open
             try: writer.close()
             except Exception as close_err: print(f"SERVER WARNING: Error closing ExcelWriter: {close_err}")
        if workbook: # Close openpyxl workbook if open
            try: workbook.close()
            except Exception as close_err: print(f"SERVER WARNING: Error closing workbook: {close_err}")

        print(f"\n--- ERROR processing {original_basename} ---"); traceback.print_exc()
        error_message = f"An error occurred: {type(e).__name__} - {e}"
        yield sse_message(f"ERROR: {error_message} - See server logs.")
        time.sleep(0.1)
        yield from yield_result(False, error_message)

# ================================================================
# == END OF SCRIPT ==
# ================================================================