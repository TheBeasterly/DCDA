# RVToolsAnalysis_web.py
import os
import sys # Keep for pandas/openpyxl interaction if needed
import openpyxl as xl
from openpyxl.styles import Font, Border, PatternFill, Protection, Alignment
from openpyxl.utils import get_column_letter # Used in filter_rows
import pandas as pd
import traceback
import time
import json
import logging # Added logging for consistency

# Configure logging (adjust level as needed)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# =================================================================================================================================================================================================================================================
# ============================================================================================= HELPER FUNCTIONS (From Original Uploaded Code) ================================================================================================
# =================================================================================================================================================================================================================================================

## Remove Formatting
def removeFormatting(destfile):
    """Removes all formatting from an excel workbook."""
    logging.info("Removing formatting from workbook...")
    try:
        for worksheet in destfile.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.has_style:
                        cell.font = Font()
                        cell.border = Border()
                        cell.fill = PatternFill(fill_type=None) # Explicitly set fill_type
                        cell.number_format = 'General'
                        cell.protection = Protection()
                        cell.alignment = Alignment()
        logging.info("Formatting removal complete.")
    except Exception as e:
        logging.error(f"Error during removeFormatting: {e}", exc_info=True)
        # print(f"SERVER WARNING: Error during removeFormatting: {e}") # Keep print if needed for SSE

## Match File Servers
def match_fs(destfile):
    """Matches file servers in all sheets based on VM name in Column A."""
    logging.info("Matching file servers...")
    fs_str = ["file", "fs", "nas", "share", "ftp"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in fs_str):
                    if worksheet.max_column >= 3: worksheet.cell(row=cell.row, column=3).value = "Yes"
                    else: logging.warning(f"Col C (IsFile) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: logging.error(f"ERROR in match_fs: {e}", exc_info=True)

## Match SQL DBs
def match_sql(destfile):
    """Matches SQL DBs in all sheets based on VM name in Column A."""
    logging.info("Matching SQL DBs...")
    sql_str = ["sql"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in sql_str):
                    if worksheet.max_column >= 4: worksheet.cell(row=cell.row, column=4).value = "Yes"
                    else: logging.warning(f"Col D (IsSQL) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: logging.error(f"ERROR in match_sql: {e}", exc_info=True)

## Match Oracle DBs
def match_orcl(destfile):
    """Matches Oracle DBs in all sheets based on VM name in Column A."""
    logging.info("Matching Oracle DBs...")
    orcl_str = ["orcl", "oracle"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in orcl_str):
                    if worksheet.max_column >= 5: worksheet.cell(row=cell.row, column=5).value = "Yes"
                    else: logging.warning(f"Col E (IsOrcl) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: logging.error(f"ERROR in match_orcl: {e}", exc_info=True)

## Match PostGres DBs
def match_pgres(destfile):
    """Matches PostGres DBs in all sheets based on VM name in Column A."""
    logging.info("Matching Postgres DBs...")
    pgres_str = ["pgres", "postgres"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in pgres_str):
                    if worksheet.max_column >= 6: worksheet.cell(row=cell.row, column=6).value = "Yes"
                    else: logging.warning(f"Col F (IsPGres) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: logging.error(f"ERROR in match_pgres: {e}", exc_info=True)

## Match Possible DBs (General - Sets "Check")
def match_gendb(destfile):
    """Matches Possible DBs; sets 'Check' if not already 'Yes'."""
    logging.info("Matching general DBs (Check)...")
    gendb_str = ["db", "database"]
    db_check_cols = [4, 5, 6] # D=IsSQL, E=IsOrcl, F=IsPGres
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in gendb_str):
                    # Apply 'Check' only if none of the specific DB cols are already 'Yes'
                    is_specific_db = False
                    for col_idx_check in db_check_cols:
                        if worksheet.max_column >= col_idx_check and worksheet.cell(row=cell.row, column=col_idx_check).value == "Yes":
                            is_specific_db = True
                            break
                    if not is_specific_db:
                        for col_idx_set in db_check_cols:
                            if worksheet.max_column >= col_idx_set:
                                target_cell = worksheet.cell(row=cell.row, column=col_idx_set)
                                # Only set Check if it's currently not Yes (redundant due to outer check, but safe)
                                if target_cell.value != "Yes":
                                    target_cell.value = "Check"
    except Exception as e: logging.error(f"ERROR in match_gendb: {e}", exc_info=True)

## Match Exchange Servers
def match_exch(destfile):
    """Matches Exchange Servers in all sheets based on VM name in Column A."""
    logging.info("Matching Exchange servers...")
    exch_str = ["exch", "exchange"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in exch_str):
                    if worksheet.max_column >= 7: worksheet.cell(row=cell.row, column=7).value = "Yes"
                    else: logging.warning(f"Col G (IsExch) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: logging.error(f"ERROR in match_exch: {e}", exc_info=True)

## Match TestDev
def match_tstdev(destfile):
    """Matches Test/Dev systems in all sheets based on VM name in Column A."""
    logging.info("Matching Test/Dev systems...")
    tstdev_str = ["tst", "test", "dev"]
    try:
        for worksheet in destfile.worksheets:
            if worksheet.max_column < 1: continue
            for cell in worksheet["A"]:
                if cell.row == 1 or cell.value is None or not isinstance(cell.value, str): continue
                cell_value_lower = cell.value.lower()
                if any(search_string in cell_value_lower for search_string in tstdev_str):
                    if worksheet.max_column >= 8: worksheet.cell(row=cell.row, column=8).value = "Yes"
                    else: logging.warning(f"Col H (IsTestDev) missing sheet '{worksheet.title}' row {cell.row}")
    except Exception as e: logging.error(f"ERROR in match_tstdev: {e}", exc_info=True)

## Get Last Row Helper
def get_last_row_in_col_a(worksheet):
    """Finds the last row index in Column A that contains data."""
    last_row = 0
    try:
        # Iterate backwards from max_row which should be reliable after openpyxl load
        start_check_row = worksheet.max_row
        for row_idx in range(start_check_row, 0, -1):
            cell = worksheet.cell(row=row_idx, column=1) # Column A
            if cell.value is not None and str(cell.value).strip() != "":
                last_row = row_idx; break
    except Exception as e: logging.warning(f"Error getting last row in Col A for sheet '{worksheet.title}': {e}")
    return last_row if last_row > 0 else 1 # Return at least 1 if header exists

## Set No Values
def set_no_values(destfile):
    """Sets empty cells in columns C to H to 'No' up to the last VM row."""
    logging.info("Setting default 'No' values for classification flags...")
    try:
        for ws_name in ['vInfo', 'vDisk', 'vPartition']:
            if ws_name in destfile.sheetnames:
                worksheet = destfile[ws_name]
                if worksheet.max_row <= 1: continue
                last_row_a = get_last_row_in_col_a(worksheet) # Find last actual VM row
                logging.info(f"Processing 'set_no_values' for {ws_name} up to row {last_row_a}...")
                for row_index in range(2, last_row_a + 1): # Iterate from row 2 to last VM row
                    for column_index in range(3, 9): # Columns C (IsFile) to H (IsTestDev)
                        if worksheet.max_column >= column_index:
                            cell = worksheet.cell(row=row_index, column=column_index)
                            # Set to 'No' only if cell is truly empty/None
                            if cell.value is None or str(cell.value).strip() == "":
                                cell.value = "No"
            else: logging.warning(f"Sheet '{ws_name}' not found for set_no_values.")
    except Exception as e: logging.error(f"ERROR in set_no_values: {e}", exc_info=True)

## Insert GB Columns Helpers (vInfo, vDisk, vPartition)
def find_col_idx(ws, header_name):
    """Finds column index (1-based) for header."""
    if ws.max_row == 0: return None # Handle empty sheet
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        # Compare lower case for robustness
        if cell.value is not None and str(cell.value).strip().lower() == header_name.lower():
            return col_idx
    return None

def insert_gb_col(ws, mib_header, gb_header, sheet_name_for_warn):
    """Generic helper to insert and calculate GB columns using MiB/953.7 factor."""
    mib_col_idx = find_col_idx(ws, mib_header)
    if mib_col_idx is None:
        logging.warning(f"Column '{mib_header}' not found in {sheet_name_for_warn}. Cannot insert {gb_header}.")
        return False
    gb_col_idx = mib_col_idx + 1
    ws.insert_cols(gb_col_idx)
    ws.cell(row=1, column=gb_col_idx).value = gb_header
    ws.cell(row=1, column=gb_col_idx).font = Font(bold=True) # Bold header
    last_row_a = get_last_row_in_col_a(ws) # Calculate only up to last VM row
    logging.info(f"Calculating {gb_header} for {sheet_name_for_warn} up to row {last_row_a}...")
    for i in range(2, last_row_a + 1):
        mib_cell = ws.cell(row=i, column=mib_col_idx)
        gb_cell = ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = pd.to_numeric(mib_cell.value, errors='coerce') # Handle non-numeric gracefully
            if pd.notna(mib_value) and mib_value != 0: # Calculate if numeric and not zero
                # Using the specific factor from original script
                gb_value = round(mib_value / 953.7, 2)
                gb_cell.value = gb_value
                gb_cell.number_format = '0.00'
            elif pd.notna(mib_value) and mib_value == 0:
                gb_cell.value = 0 # Set GB explicitly to 0 if MiB is 0
                gb_cell.number_format = '0.00'
            else: gb_cell.value = None # Set to blank if source not numeric/coerced to NaN
        except Exception as calc_err:
            logging.warning(f"Could not calculate {gb_header} for row {i} in {sheet_name_for_warn}: {calc_err}")
            gb_cell.value = None # Set to blank on error
    logging.info(f"Finished calculating {gb_header} for {sheet_name_for_warn}.")
    return True

def vinfo_insert_gb_cols(ws):
    insert_gb_col(ws, "Provisioned MiB", "Provisioned GB", "vInfo")
    insert_gb_col(ws, "In Use MiB", "In Use GB", "vInfo")
def vdisk_insert_gb_cols(ws):
    insert_gb_col(ws, "Capacity MiB", "Capacity GB", "vDisk")
def vpart_insert_gb_cols(ws):
    insert_gb_col(ws, "Capacity MiB", "Capacity GB", "vPartition")
    insert_gb_col(ws, "Consumed MiB", "Consumed GB", "vPartition")
    insert_gb_col(ws, "Free MiB", "Free GB", "vPartition")

## Set Disk Count Value
def vdisk_diskcount_val(vdisk_ws):
    """Sets DiskCount column to 1 for all data rows."""
    logging.info("Setting DiskCount value in vDisk...")
    diskcount_col_idx = find_col_idx(vdisk_ws, "DiskCount")
    if diskcount_col_idx is None:
        logging.warning("'DiskCount' column not found in vDisk. Cannot set values.")
        return False
    last_row_a = get_last_row_in_col_a(vdisk_ws)
    for i in range(2, last_row_a + 1):
        vdisk_ws.cell(row=i, column=diskcount_col_idx).value = 1
    logging.info(f"DiskCount set to 1 up to row {last_row_a}.")
    return True

## Compare VMs for HasTools
def compare_vms(vinfo_ws, vpart_ws):
  """Compares VMs in vInfo to vPartition, sets HasTools in vInfo (Col I)."""
  logging.info("Comparing vInfo VMs to vPartition to set HasTools flag...")
  if vinfo_ws is None or vpart_ws is None:
      logging.error("Missing vInfo or vPartition sheet for compare_vms.")
      return
  try:
      vpart_vms = set()
      if vpart_ws.max_row > 1:
          # Read VM names from vPartition efficiently
          for cell in vpart_ws['A'][1:]: # Skip header row
              if cell.value is not None and str(cell.value).strip():
                  vpart_vms.add(str(cell.value))
      logging.info(f"Found {len(vpart_vms)} unique non-empty VM names in vPartition.")

      # Find HasTools column index (should be I = 9 after inserts)
      hastools_col_idx = find_col_idx(vinfo_ws, "HasTools") # Dynamically find it
      if hastools_col_idx is None:
           logging.warning("'HasTools' column missing in vInfo. Cannot set flag.")
           return

      last_row_a_vinfo = get_last_row_in_col_a(vinfo_ws)
      logging.info(f"Setting HasTools in vInfo up to row {last_row_a_vinfo}...")
      match_count = 0
      for row_idx in range(2, last_row_a_vinfo + 1):
          vinfo_vm_cell = vinfo_ws.cell(row=row_idx, column=1) # VM name in Col A
          hastools_cell = vinfo_ws.cell(row=row_idx, column=hastools_col_idx)
          vm_name = str(vinfo_vm_cell.value) if vinfo_vm_cell.value is not None else ""
          if vm_name and vm_name in vpart_vms:
              hastools_cell.value = "Yes"
              match_count += 1
          else:
              hastools_cell.value = "No"
      logging.info(f"Comparison complete. Set 'Yes' for {match_count} VMs.")

  except Exception as e: logging.error(f"ERROR in compare_vms: {e}", exc_info=True)

## Delete Unnecessary Columns (Pandas)
def del_cols_vInfo(destfile_path):
    """Deletes unnecessary columns using Pandas, keeping specific columns."""
    logging.info(f"Running Pandas del_cols_vInfo on {destfile_path}...")
    try:
        # Define columns to keep for each sheet (matches original script)
        keep_cols_map = {
            'vInfo': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'HasTools', 'Disks', 'Total disk capacity', 'Provisioned MiB', 'Provisioned GB', 'In Use MiB', 'In Use GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vDisk': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'DiskCount', 'Disk', 'Capacity MiB', 'Capacity GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vPartition': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'Disk', 'Capacity MiB', 'Capacity GB', 'Consumed MiB', 'Consumed GB', 'Free MiB', 'Free GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']
        }
        all_sheets_df = {}
        original_sheets = []
        try: # Read all existing sheets first
            xls = pd.ExcelFile(destfile_path, engine='openpyxl')
            original_sheets = xls.sheet_names
            for sheet_name in original_sheets:
                 # Only process sheets we intend to modify
                 if sheet_name in keep_cols_map:
                      all_sheets_df[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                 # Read others to write them back unchanged (if needed, currently overwrites)
                 # else:
                 #      all_sheets_df[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
            xls.close()
        except Exception as ex_read_err:
            logging.error(f"Pandas read error in del_cols: {ex_read_err}", exc_info=True); raise
        if not all_sheets_df: raise ValueError("No relevant sheets found for del_cols.")

        # Overwrite the file with modified sheets
        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            for sheet_name in original_sheets: # Iterate in original order
                 if sheet_name in all_sheets_df and sheet_name in keep_cols_map:
                     df = all_sheets_df[sheet_name]
                     # Filter columns to keep only those existing in the dataframe
                     cols_to_keep_in_df = [col for col in keep_cols_map[sheet_name] if col in df.columns]
                     if not cols_to_keep_in_df:
                          logging.warning(f"No valid columns to keep found for sheet '{sheet_name}'. Skipping.")
                          continue
                     logging.info(f"Writing sheet '{sheet_name}' with columns: {cols_to_keep_in_df}")
                     df[cols_to_keep_in_df].to_excel(writer, sheet_name=sheet_name, index=False)
                 # else: # Write back other sheets unchanged if read earlier
                 #      if sheet_name in all_sheets_df:
                 #           all_sheets_df[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

        logging.info("Pandas del_cols_vInfo completed successfully.")

    except Exception as e:
        logging.error(f"ERROR in del_cols_vInfo: {e}", exc_info=True); raise

## Truncate vPartition to vSummary (Pandas)
def trunc_cols_vPart(destfile_path):
    """Aggregates vPartition data into vSummary using Pandas, overwriting the file."""
    logging.info(f"Running Pandas trunc_cols_vPart (vSummary creation) on {destfile_path}...")
    try:
        dfs = {}
        required_sheets = ['vInfo', 'vPartition']
        sheets_to_keep_after = ['vSummary', 'vInfo', 'vDisk', 'vPartition'] # Sheets expected in output
        try: # Read only needed sheets
            xls = pd.ExcelFile(destfile_path, engine='openpyxl')
            # Read required for calculation
            for sheet_name in required_sheets:
                 if sheet_name in xls.sheet_names: dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                 else: raise ValueError(f"Required sheet '{sheet_name}' missing for vSummary creation.")
            # Read vDisk if present to write it back
            if 'vDisk' in xls.sheet_names: dfs['vDisk'] = pd.read_excel(xls, sheet_name='vDisk')
            xls.close()
        except Exception as ex_read_err:
            logging.error(f"Pandas read error in trunc_cols: {ex_read_err}", exc_info=True); raise

        # --- Perform Aggregation ---
        df_vinfo = dfs['vInfo']; df_vpartition = dfs['vPartition']

        # Columns to aggregate from vPartition
        storage_cols = ['Capacity GB', 'Consumed GB', 'Free GB']
        existing_storage_cols = [col for col in storage_cols if col in df_vpartition.columns]
        if not existing_storage_cols:
             logging.warning("No storage columns (Capacity GB, Consumed GB, Free GB) found in vPartition to aggregate. vSummary will lack these sums.")
             aggregated_df = pd.DataFrame(columns=['VM'] + existing_storage_cols) # Empty agg result
        else:
             if 'VM' not in df_vpartition.columns: raise ValueError("'VM' column missing in vPartition.");
             # Ensure numeric for aggregation, fill NaN with 0
             for col in existing_storage_cols:
                 df_vpartition[col] = pd.to_numeric(df_vpartition[col], errors='coerce').fillna(0)
             aggregated_df = df_vpartition.groupby('VM', as_index=False)[existing_storage_cols].sum()

        # --- Merge with vInfo ---
        if 'VM' not in df_vinfo.columns: raise ValueError("'VM' column missing in vInfo.");
        # Start vSummary with all columns from vInfo
        vSummary_df = df_vinfo.copy()
        # Merge aggregated storage if available
        if not aggregated_df.empty:
            vSummary_df = pd.merge(vSummary_df, aggregated_df, on='VM', how='left')
            # Fill NaNs introduced by merge (for VMs in vInfo but not vPart) with 0
            for col in existing_storage_cols:
                 if col in vSummary_df.columns: vSummary_df[col] = pd.to_numeric(vSummary_df[col], errors='coerce').fillna(0)
        else: # Add empty columns if no aggregation happened
             for col in storage_cols: vSummary_df[col] = 0

        # --- Reorder Columns --- (Place aggregated cols after 'In Use GB')
        target_col_after = 'In Use GB'
        cols = vSummary_df.columns.tolist()
        if target_col_after in cols and existing_storage_cols:
            try:
                 idx = cols.index(target_col_after)
                 # Remove aggregated columns from their current positions (likely end)
                 current_agg_cols = [c for c in existing_storage_cols if c in cols]
                 for col in current_agg_cols: cols.remove(col)
                 # Insert them after the target column
                 cols = cols[:idx+1] + current_agg_cols + cols[idx+1:]
                 vSummary_df = vSummary_df[cols] # Apply new order
            except Exception as reorder_err:
                 logging.warning(f"Could not reorder vSummary columns: {reorder_err}")
        elif not existing_storage_cols:
             logging.info("No aggregated columns to reorder.")
        else:
             logging.warning(f"Target column '{target_col_after}' not found for reordering vSummary.")

        # --- Write all expected sheets back ---
        logging.info("Writing vSummary and other sheets back...")
        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            vSummary_df.to_excel(writer, sheet_name='vSummary', index=False)
            dfs['vInfo'].to_excel(writer, sheet_name='vInfo', index=False)
            if 'vDisk' in dfs: dfs['vDisk'].to_excel(writer, sheet_name='vDisk', index=False)
            # Write original vPartition back (not the aggregated one)
            dfs['vPartition'].to_excel(writer, sheet_name='vPartition', index=False)

        logging.info("Pandas trunc_cols_vPart (vSummary creation) completed successfully.")

    except Exception as e:
        logging.error(f"ERROR in trunc_cols_vPart: {e}", exc_info=True); raise

## Trim vSummary Columns (Openpyxl)
def trimvSum1(destfile):
    """Removes 'Provisioned MiB' and 'In Use MiB' columns from vSummary using openpyxl."""
    if 'vSummary' not in destfile.sheetnames:
        logging.warning("vSummary sheet not found for trimvSum1.")
        return
    vSum_ws = destfile['vSummary']
    logging.info("Trimming MiB columns from vSummary...")
    cols_to_delete_names = ['Provisioned MiB', 'In Use MiB']
    deleted_count = 0
    try:
        # Find column indices from right to left to avoid shifting issues
        headers = [cell.value for cell in vSum_ws[1]]
        indices_to_delete = []
        for name in cols_to_delete_names:
            try:
                 # Find index (1-based)
                 idx = headers.index(name) + 1
                 indices_to_delete.append(idx)
            except ValueError:
                 logging.warning(f"Column '{name}' not found for trimming in vSummary.")

        # Sort indices in descending order before deleting
        indices_to_delete.sort(reverse=True)
        for col_idx in indices_to_delete:
             vSum_ws.delete_cols(col_idx, 1)
             deleted_count += 1
        logging.info(f"Trimmed {deleted_count} columns from vSummary.")

    except Exception as e: logging.error(f"ERROR in trimvSum1: {e}", exc_info=True)

## Consolidate vSummary Consumed GB (Openpyxl)
def consol_vSum(destfile):
    """Copies 'In Use GB' to 'Consumed GB' if 'Consumed GB' is empty, blank, or zero."""
    if 'vSummary' not in destfile.sheetnames:
        logging.warning("vSummary sheet not found for consol_vSum.")
        return
    vSum_ws = destfile['vSummary']
    logging.info("Consolidating vSummary Consumed GB from In Use GB where needed...")
    updated_count = 0
    try:
        consumed_gb_col_idx, in_use_gb_col_idx = None, None
        headers = [cell.value for cell in vSum_ws[1]]
        for idx, header in enumerate(headers, 1):
            if header == 'Consumed GB': consumed_gb_col_idx = idx
            elif header == 'In Use GB': in_use_gb_col_idx = idx

        if consumed_gb_col_idx is None: logging.warning("'Consumed GB' column not found."); return
        if in_use_gb_col_idx is None: logging.warning("'In Use GB' column not found."); return

        last_row_a = get_last_row_in_col_a(vSum_ws)
        for row_idx in range(2, last_row_a + 1):
            consumed_cell = vSum_ws.cell(row=row_idx, column=consumed_gb_col_idx)
            in_use_cell = vSum_ws.cell(row=row_idx, column=in_use_gb_col_idx)
            consumed_val = consumed_cell.value
            # Check for None, empty string, or numerical 0
            is_empty_or_zero = False
            if consumed_val is None or (isinstance(consumed_val, str) and consumed_val.strip() == ""):
                is_empty_or_zero = True
            else:
                try: # Check for numerical zero
                    if pd.to_numeric(consumed_val, errors='coerce') == 0:
                        is_empty_or_zero = True
                except Exception: pass # Ignore conversion errors

            if is_empty_or_zero:
                # Copy value and format if source has value
                if in_use_cell.value is not None:
                    consumed_cell.value = in_use_cell.value
                    if in_use_cell.has_style and in_use_cell.number_format:
                         consumed_cell.number_format = in_use_cell.number_format
                    updated_count +=1

        logging.info(f"Consolidated Consumed GB for {updated_count} rows.")

    except Exception as e: logging.error(f"ERROR in consol_vSum: {e}", exc_info=True)

## Filter First Row (Openpyxl)
def filter_rows(workbook):
  """Applies auto-filter to the header row on all sheets."""
  logging.info("Applying filters to all sheets...")
  applied_count = 0
  try:
    for ws in workbook.worksheets:
      # Apply filter only if sheet has more than just a header row
      if ws.max_row > 1 and ws.max_column > 0:
          # Reference the data range including header for filtering
          ws.auto_filter.ref = ws.dimensions
          applied_count += 1
          # Optional: Freeze top row for better usability
          ws.freeze_panes = 'A2'
    logging.info(f"Filters applied to {applied_count} sheets.")
  except Exception as e: logging.error(f"ERROR in filter_rows: {e}", exc_info=True)

# ================================================
# --- Categorization Helper Functions ---
# ================================================
# (Defined outside process_rvtools_file for potential reuse)
def assign_workload_category(row):
    # Using .get() for safety if columns might be missing in edge cases
    is_values = {
        'SQL': row.get('IsSQL') == 'Yes',
        'Exchange': row.get('IsExch') == 'Yes',
        'File': row.get('IsFile') == 'Yes',
        'Oracle': row.get('IsOrcl') == 'Yes',
        'Postgres': row.get('IsPGres') == 'Yes'
    }
    db_check_flags = {
        'SQL_Check': row.get('IsSQL') == 'Check',
        'Orcl_Check': row.get('IsOrcl') == 'Check',
        'PGres_Check': row.get('IsPGres') == 'Check'
    }
    yes_count = sum(is_values.values())
    if yes_count > 1: return "Multi-Error"
    elif yes_count == 1:
        for w, is_y in is_values.items():
            if is_y: return w
    # Check for 'Check' flags only if no 'Yes' flags were found
    elif any(db_check_flags.values()): return "General DB"
    else: return "Standard" # Default if no 'Yes' or 'Check'
    # Fallback, should not be reached with above logic
    # return "Unknown"

def assign_environment(row):
    return "Test/Dev" if row.get('IsTestDev') == 'Yes' else "Prod"

# ================================================
# --- *** Reusable Summary Calculation Function *** ---
# ================================================
def calculate_summaries(df_categorized_summary):
    """
    Calculates summary DataFrames from a categorized vSummary DataFrame.
    Returns a dictionary of {sheet_name: dataframe}.
    Uses column names from the specific refactored version.
    """
    logging.info("Calculating summaries...")
    summary_dfs = {}

    # Define required columns based on the aggregations needed
    required_cols = ['Workload', 'Environment', 'VM Count', 'Disks', 'Consumed GB', 'Powerstate', 'Datacenter', 'Cluster']

    # Input validation
    if df_categorized_summary is None or df_categorized_summary.empty:
        logging.error("Input DataFrame for calculate_summaries is None or empty.")
        return {} # Return empty dict on error
    if not all(c in df_categorized_summary.columns for c in required_cols):
        missing = [c for c in required_cols if c not in df_categorized_summary.columns]
        logging.error(f"Missing required columns for calculate_summaries: {missing}")
        return {} # Return empty dict on error

    try:
        # Ensure numeric columns are numeric, fillna if necessary before aggregation
        numeric_agg_cols = ['VM Count', 'Disks', 'Consumed GB']
        for col in numeric_agg_cols:
             # Coerce errors, then fill NaN with 0 for aggregation
             df_categorized_summary[col] = pd.to_numeric(df_categorized_summary[col], errors='coerce').fillna(0)

        # Filter for PoweredOn VMs AFTER ensuring numeric conversion
        df_powered_on = df_categorized_summary[df_categorized_summary['Powerstate'] == 'poweredOn'].copy()

        # Define Aggregations
        agg_total = {'VM Count': 'sum', 'Disks': 'sum', 'Consumed GB': 'sum'}
        agg_pon_counts = {'VM Count': 'sum'} # Used for Overall Powerstate Counts logic
        agg_dc_pon_metrics = { # Used for merging into DC Combined
             'PoweredOn VMs': ('VM Count', 'sum'),
             'PoweredOn Disks': ('Disks', 'sum'),
             'PoweredOn Consumed GB': ('Consumed GB', 'sum')
             }

        # --- Calculate Summaries ---
        logging.info("Calculating Overall Summary Totals...")
        summary_overall_total = df_categorized_summary.groupby(['Workload', 'Environment'], observed=False, dropna=False).agg(agg_total).reset_index()
        summary_dfs['Overall Summary Totals'] = summary_overall_total

        logging.info("Calculating Overall Powerstate Counts...")
        summary_overall_pon_counts = df_categorized_summary.groupby(['Workload', 'Environment', 'Powerstate'], observed=False, dropna=False).agg(agg_pon_counts).reset_index()
        summary_dfs['Overall Powerstate Counts'] = summary_overall_pon_counts

        logging.info("Calculating Datacenter Summary Combined...")
        summary_dc_base = df_categorized_summary.groupby(['Datacenter', 'Cluster', 'Workload', 'Environment'], observed=False, dropna=False).agg(agg_total).reset_index()
        summary_dc_pon_metrics = df_powered_on.groupby(['Datacenter', 'Cluster', 'Workload', 'Environment'], observed=False, dropna=False).agg(**agg_dc_pon_metrics).reset_index()
        summary_dc_combined = pd.merge(summary_dc_base, summary_dc_pon_metrics, on=['Datacenter', 'Cluster', 'Workload', 'Environment'], how='left').fillna(0)
        # Ensure specific numeric types after merge/fillna if needed downstream, e.g., int for counts
        int_cols = ['VM Count', 'Disks', 'PoweredOn VMs', 'PoweredOn Disks']
        for col in int_cols:
             if col in summary_dc_combined.columns:
                   # Use astype(int) which might be faster if no NaNs expected after fillna(0)
                   summary_dc_combined[col] = summary_dc_combined[col].astype(int)

        summary_dfs['Datacenter Summary Combined'] = summary_dc_combined

        logging.info("Summaries calculated successfully.")

    except Exception as e:
        logging.error(f"Failed during summary calculation: {e}", exc_info=True)
        # traceback.print_exc() # Keep print if needed for SSE
        return {} # Return empty dict if calculation fails

    return summary_dfs

# ================================================================================================================================================================================================================================================
# ===================================================================================================== MAIN PROCESSING FUNCTION (SSE GENERATOR) =================================================================================================
# ================================================================================================================================================================================================================================================

def process_rvtools_file(input_filepath, output_folder, original_basename):
    """
    Processes RVTools, yields SSE updates, categorizes, calls calculate_summaries,
    adds dynamic SUBTOTAL rows using openpyxl, reorders sheets, and applies filters.
    Refactored version incorporating original helper functions.
    """
    # --- SSE Helper Functions ---
    def sse_message(data): data = str(data).replace('\n', ' '); return f"data: {data}\n\n"
    def yield_result(success, message): payload = json.dumps({"success": success, "message": message}); yield f"event: result\ndata: {payload}\n\n"

    # --- Main Processing Logic ---
    destfile = None; output_filename = None; writer = None; workbook = None
    start_time = time.time() # Start timer
    try:
        # --- Setup Output Path ---
        yield sse_message(f"Starting analysis for: {original_basename}"); time.sleep(0.1)
        if not os.path.isdir(output_folder): os.makedirs(output_folder); yield sse_message(f"Created output folder: {output_folder}")
        if original_basename: base_name_no_ext, ext = os.path.splitext(original_basename); output_filename = f"{base_name_no_ext}-ANALYZED{ext}"
        else: temp_base, ext = os.path.splitext(os.path.basename(input_filepath)); output_filename = f"{temp_base}-ORIGNAMEMISSING-ANALYZED{ext}"
        output_filepath = os.path.join(output_folder, output_filename)
        yield sse_message(f"Output file will be: {output_filename}"); logging.info(f"Output file path: {output_filepath}"); time.sleep(0.1)

        # --- Steps 1-8: Initial Data Prep (Openpyxl - using original helpers) ---
        yield sse_message("Step 1-8: Running Pre-processing..."); time.sleep(0.1)
        intermediate_filepath = output_filepath + ".prep.xlsx" # Use intermediate file
        try:
            logging.info(f"Loading workbook: {input_filepath}")
            destfile = xl.load_workbook(input_filepath)
            removeFormatting(destfile)

            logging.info("Deleting non-essential sheets...")
            keep_sheets=['vInfo', 'vDisk', 'vPartition'] # Sheets needed for processing
            sheets_to_delete=[s for s in destfile.sheetnames if s not in keep_sheets]
            if not any(s in destfile.sheetnames for s in keep_sheets):
                raise ValueError("Required sheets (vInfo, vDisk, vPartition) missing.")
            for sn in sheets_to_delete:
                 if sn in destfile: del destfile[sn]
            logging.info(f"Kept sheets: {destfile.sheetnames}")

            logging.info("Inserting classification/helper columns...")
            for ws_name in keep_sheets:
                 if ws_name in destfile:
                     ws=destfile[ws_name]
                     # Insert classification flag columns (C-H)
                     ws.insert_cols(3, 6)
                     ws['C1']='IsFile'; ws['D1']='IsSQL'; ws['E1']='IsOrcl'; ws['F1']='IsPGres'; ws['G1']='IsExch'; ws['H1']='IsTestDev'
                     for col_letter in ['C','D','E','F','G','H']: ws[f'{col_letter}1'].font = Font(bold=True)
                     # Insert specific helper columns
                     if ws_name=='vInfo':
                          ws.insert_cols(9, 1); ws['I1']='HasTools'; ws['I1'].font = Font(bold=True)
                     elif ws_name=='vDisk':
                          ws.insert_cols(9, 1); ws['I1']='DiskCount'; ws['I1'].font = Font(bold=True)

            logging.info("Matching workload/environment keywords...")
            match_fs(destfile); match_sql(destfile); match_orcl(destfile); match_pgres(destfile); match_exch(destfile); match_tstdev(destfile); match_gendb(destfile);
            set_no_values(destfile) # Set blanks to 'No'

            logging.info("Setting HasTools and DiskCount values...")
            vinfo_ws=destfile['vInfo'] if 'vInfo' in destfile else None
            vpart_ws=destfile['vPartition'] if 'vPartition' in destfile else None
            vdisk_ws=destfile['vDisk'] if 'vDisk' in destfile else None
            if vinfo_ws and vpart_ws: compare_vms(vinfo_ws, vpart_ws)
            if vdisk_ws: vdisk_diskcount_val(vdisk_ws)

            logging.info("Inserting and calculating GB columns...")
            if vinfo_ws: vinfo_insert_gb_cols(vinfo_ws)
            if vdisk_ws: vdisk_insert_gb_cols(vdisk_ws)
            if vpart_ws: vpart_insert_gb_cols(vpart_ws)

            logging.info(f"Saving intermediate file to: {intermediate_filepath}")
            destfile.save(intermediate_filepath)
        except Exception as prep_err:
             logging.error(f"Openpyxl pre-processing failed: {prep_err}", exc_info=True)
             yield sse_message(f"ERROR during Step 1-8: {prep_err}"); time.sleep(0.1)
             raise # Stop if critical prep fails
        finally:
             if destfile: destfile.close(); destfile=None
        yield sse_message(" -> Pre-processing complete."); time.sleep(0.1)

        # --- Steps 9-10: Pandas Operations (using original helpers) ---
        yield sse_message("Step 9: Removing unused columns..."); time.sleep(0.1)
        try: del_cols_vInfo(intermediate_filepath) # Operate on intermediate file
        except Exception as e: logging.error(f"Pandas del_cols error: {e}", exc_info=True); raise
        yield sse_message(" -> Unused columns removed."); time.sleep(0.1)

        yield sse_message("Step 10: Creating vSummary sheet..."); time.sleep(0.1)
        try: trunc_cols_vPart(intermediate_filepath) # This overwrites intermediate file
        except Exception as e: logging.error(f"Pandas trunc_cols error: {e}", exc_info=True); raise
        yield sse_message(" -> vSummary sheet created/updated."); time.sleep(0.1)

        # --- Step 11: Openpyxl Final Cleanup (before summary generation - using original helpers) ---
        yield sse_message("Step 11: Performing cleanup..."); time.sleep(0.1)
        try:
            destfile = xl.load_workbook(intermediate_filepath) # Load the file pandas wrote
            removeFormatting(destfile) # Remove pandas formatting artefacts
            if 'vSummary' in destfile.sheetnames:
                 trimvSum1(destfile) # Trim MiB cols
                 consol_vSum(destfile) # Consolidate Consumed GB
            else: logging.warning("vSummary not found for cleanup (trim/consol).")
            # NO filtering here
            logging.info(f"Saving final cleaned data to: {output_filepath}")
            destfile.save(output_filepath); # Save final cleaned version to the actual output path
        except Exception as clean_err:
             logging.error(f"Cleanup failed: {clean_err}", exc_info=True)
             yield sse_message(f"ERROR during Step 11 cleanup: {clean_err}"); time.sleep(0.1)
             # If cleanup fails, maybe still try to generate summaries from intermediate? Or stop? Stopping seems safer.
             raise
        finally:
             if destfile: destfile.close(); destfile=None
             # Clean up intermediate file
             try:
                  if os.path.exists(intermediate_filepath): os.remove(intermediate_filepath)
             except OSError as rm_err: logging.warning(f"Could not remove intermediate file {intermediate_filepath}: {rm_err}")
        yield sse_message(" -> Cleanup complete."); time.sleep(0.1)

        # --- Step 13: Categorize, Generate Summaries (using calculate_summaries) & Write Sheets ---
        yield sse_message("Step 13: Categorizing and generating analysis summaries..."); time.sleep(0.1)
        summary_sheets_generated = []
        try:
            yield sse_message(" -> Reading cleaned vSummary sheet..."); time.sleep(0.1)
            # Read from the final cleaned output file
            df_summary_source = pd.read_excel(output_filepath, sheet_name='vSummary', engine='openpyxl')
            if df_summary_source.empty: raise ValueError("vSummary sheet is empty after cleanup.")
            yield sse_message(" -> vSummary loaded."); time.sleep(0.1)

            # --- Add VM Count column (essential for calculate_summaries) ---
            if 'VM Count' not in df_summary_source.columns:
                 if 'VM' in df_summary_source.columns:
                      logging.info("Adding 'VM Count' column with value 1.")
                      df_summary_source['VM Count'] = 1
                 else: raise ValueError("Cannot add 'VM Count', missing 'VM' column.")

            # Check required columns for categorization & summaries are present AFTER cleanup
            required_cols_cat = ['VM','IsFile','IsSQL','IsOrcl','IsPGres','IsExch','IsTestDev','Disks','Consumed GB','Powerstate','Datacenter','Cluster', 'VM Count']
            missing_cols = [c for c in required_cols_cat if c not in df_summary_source.columns]
            if missing_cols: raise ValueError(f"Missing columns in final vSummary needed for categorization/summaries: {missing_cols}")

            # --- Categorize ---
            yield sse_message(" -> Categorizing workloads..."); time.sleep(0.1)
            df_summary_source['Workload'] = df_summary_source.apply(assign_workload_category, axis=1)
            df_summary_source['Environment'] = df_summary_source.apply(assign_environment, axis=1)

            # --- Position new columns --- (relative to 'IsFile')
            yield sse_message(" -> Positioning category columns..."); time.sleep(0.1)
            try:
                if 'IsFile' in df_summary_source.columns:
                    isfile_idx = df_summary_source.columns.get_loc('IsFile')
                    cols = df_summary_source.columns.tolist()
                    # Remove first if present, then insert
                    if 'Workload' in cols: cols.remove('Workload')
                    if 'Environment' in cols: cols.remove('Environment')
                    # Insert Environment then Workload before IsFile index
                    cols.insert(isfile_idx, 'Environment')
                    cols.insert(isfile_idx, 'Workload')
                    df_summary_source = df_summary_source[cols]
                    yield sse_message(" -> Category columns positioned."); time.sleep(0.05)
                else: yield sse_message(" -> Warn: 'IsFile' not found, cannot position category columns.")
            except Exception as pos_err: yield sse_message(f" -> Warn: Error positioning category columns: {pos_err}")

            # --- Call calculate_summaries function ---
            yield sse_message(" -> Calculating summaries..."); time.sleep(0.1)
            summary_dfs = calculate_summaries(df_summary_source.copy()) # Pass copy
            if not summary_dfs: raise ValueError("Summary calculation function returned empty results.")
            yield sse_message(" -> Summaries calculated."); time.sleep(0.05)

            # --- Write sheets using append/replace mode ---
            yield sse_message(" -> Writing categorized vSummary and Overall Summary sheets..."); time.sleep(0.1)
            with pd.ExcelWriter(output_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                # Write categorized vSummary back (replacing the version from step 11)
                df_summary_source.to_excel(writer, sheet_name='vSummary', index=False)
                logging.info("Overwrote vSummary with categorized data.")
                summary_sheets_generated.append('vSummary (Categorized)')
                # Write summary tables from the dictionary
                for sheet_name, df_to_write in summary_dfs.items():
                        if not df_to_write.empty:
                            df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
                            logging.info(f"Wrote Overall Summary sheet: {sheet_name}")
                            summary_sheets_generated.append(sheet_name)
                        else: logging.warning(f"Skipping empty Overall Summary sheet: {sheet_name}")
            yield sse_message(f" -> {len(summary_sheets_generated)} sheets written/updated in Step 13."); time.sleep(0.1)

        except Exception as summary_err:
            logging.error(f"Error during categorization/summary generation (Step 13): {summary_err}", exc_info=True)
            yield sse_message(f"ERROR: Failed Step 13: {summary_err}"); time.sleep(0.1)
            raise # Stop processing if this critical step fails

        # --- Step 14: Add Formulas, Reorder Sheets & Apply Filters (Openpyxl - using original helpers) ---
        yield sse_message("Step 14: Adding subtotals, reordering sheets, and applying first row filtering..."); time.sleep(0.1)
        try:
            logging.info("Loading workbook for final adjustments...")
            workbook = xl.load_workbook(output_filepath)

            # --- Add SUBTOTAL Formulas ---
            logging.info("Adding subtotal rows...")
            sheets_for_subtotals = { # Map sheet name to numeric columns and label column index
                'Overall Summary Totals': {'cols': [3, 4, 5], 'label_col': 1}, # VM Count, Disks, Consumed GB; Label Col A (Workload)
                'Overall Powerstate Counts': {'cols': [4], 'label_col': 1}, # VM Count; Label Col A (Workload)
                'Datacenter Summary Combined': {'cols': [5, 6, 7, 8, 9, 10], 'label_col': 1} # VM Count -> PON Consumed GB; Label Col A (DC)
            }
            total_row_font = Font(bold=True)
            subtotal_added_count = 0
            for sheet_name, config in sheets_for_subtotals.items():
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]; max_row = ws.max_row
                    if max_row > 1: # Only add if there's data rows
                        total_row_idx = max_row + 1
                        label_cell = ws.cell(row=total_row_idx, column=config['label_col']); label_cell.value = "Grand Total"; label_cell.font = total_row_font
                        for col_idx in config['cols']:
                            col_letter = get_column_letter(col_idx); formula = f"=SUBTOTAL(9,{col_letter}2:{col_letter}{max_row})"
                            total_cell = ws.cell(row=total_row_idx, column=col_idx); total_cell.value = formula; total_cell.font = total_row_font
                            if max_row >= 2: # Try copy number format
                                prev_cell = ws.cell(row=max_row, column=col_idx)
                                if prev_cell.number_format: total_cell.number_format = prev_cell.number_format
                        subtotal_added_count += 1
                        logging.info(f"Added subtotal row to '{sheet_name}'.")
                else: logging.warning(f"Sheet '{sheet_name}' not found for subtotal.")
            yield sse_message(f" -> Added subtotal rows to {subtotal_added_count} sheets."); time.sleep(0.1)

            # --- Reorder Sheets ---
            logging.info("Reordering sheets...")
            desired_order = [ 'Overall Summary Totals', 'Overall Powerstate Counts', 'Datacenter Summary Combined', 'vSummary', 'vInfo', 'vDisk', 'vPartition' ] # Add others if they exist
            current_sheets = workbook.sheetnames
            final_order = [s for s in desired_order if s in current_sheets] + [s for s in current_sheets if s not in desired_order]
            workbook._sheets = sorted(workbook._sheets, key=lambda ws_sort: final_order.index(ws_sort.title))
            yield sse_message(" -> Sheets reordered."); time.sleep(0.1)

            # --- Apply Filters (using original helper) ---
            filter_rows(workbook) # Applies filters and freezes panes
            yield sse_message(" -> First row filters applied."); time.sleep(0.1)

            # --- Final Save ---
            logging.info("Saving final adjusted workbook...")
            workbook.save(output_filepath)
            yield sse_message(" -> Final workbook saved."); time.sleep(0.1)

        except Exception as final_step_err:
              logging.error(f"Final Step 14 adjustments failed: {final_step_err}", exc_info=True)
              yield sse_message(f"WARNING: Failed final adjustments (Step 14): {final_step_err}"); time.sleep(0.1)
        finally:
             if workbook: workbook.close()

        # --- Final Success Yield ---
        total_time = time.time() - start_time
        yield sse_message(f"--- Analysis File Generation Complete ({total_time:.2f}s) ---")
        logging.info(f"--- Analysis Completed: {original_basename} -> {output_filename} in {total_time:.2f} seconds ---")
        yield from yield_result(True, output_filename)

    # --- Global Error Handling ---
    except Exception as e:
        # Ensure workbook handles are closed on error
        if destfile:
            try: destfile.close()
            except Exception as close_err: print(f"SERVER WARNING: Error closing destfile: {close_err}")
        if 'writer' in locals() and writer is not None and writer.book is not None: # Close pandas writer if open
             try: writer.close()
             except Exception as close_err: print(f"SERVER WARNING: Error closing ExcelWriter: {close_err}")
        if workbook: # Close openpyxl workbook if open
            try: workbook.close()
            except Exception as close_err: print(f"SERVER WARNING: Error closing workbook: {close_err}")

        print(f"\n--- ERROR processing {original_basename} ---"); traceback.print_exc()
        error_message = f"An error occurred: {type(e).__name__} - {e}"
        yield sse_message(f"ERROR: {error_message} - See server logs.")
        time.sleep(0.1)
        yield from yield_result(False, error_message)

# ================================================================
# == END OF SCRIPT ==
# ================================================================