# app.py
# Corrected version 10: Removing url_for from stream route

import os
import uuid
import time
import json
import html
import traceback
import pandas as pd
import openpyxl as xl
import logging
from flask import (Flask, request, render_template, redirect,
                   url_for, send_from_directory, flash, session,
                   Response, jsonify, abort, current_app) # Keep current_app import for now
from werkzeug.utils import secure_filename

# --- Import backend processing functions ---
try:
    from RVToolsAnalysis_web import process_rvtools_file, calculate_summaries, filter_rows
except ImportError as import_err:
    logging.error(f"Failed to import required functions from RVToolsAnalysis_web: {import_err}", exc_info=True)
    def process_rvtools_file(*args, **kwargs): yield f"event: result\ndata: {json.dumps({'success': False, 'message': 'Error: Backend processing script (RVToolsAnalysis_web.py) not found or missing functions.'})}\n\n"
    def calculate_summaries(*args, **kwargs): logging.error("Dummy calculate_summaries called."); return {}
    def filter_rows(*args, **kwargs): logging.error("Dummy filter_rows called."); pass

# --- Configuration ---
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_FOLDER = os.path.join(BASE_DIR, 'client_data')
DEFAULT_PROJECT_NAME = "General"
ALLOWED_EXTENSIONS = {'xlsx'}
os.makedirs(DATA_FOLDER, exist_ok=True)

# Initialize Flask App
app = Flask(__name__)
app.config['DATA_FOLDER'] = DATA_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-secret-key-placeholder')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(name)s - %(message)s')

# --- Temporary Task Storage ---
TASK_INFO = {}

# --- Helper Function ---
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ================================================================
# --- Core Application Routes (Based on Original File) ---
# ================================================================
@app.route('/', methods=['GET'])
def index():
    """Renders the main landing/welcome page."""
    return render_template('landing.html')

@app.route('/rvtools_analysis', methods=['GET'])
def rvtools_analysis():
    """Renders the upload form page for RVTools analysis."""
    client_list = []; data_folder_path = app.config['DATA_FOLDER']
    try:
        if os.path.exists(data_folder_path): items = os.listdir(data_folder_path); client_list = [i for i in items if os.path.isdir(os.path.join(data_folder_path, i)) and not i.startswith('.')]; client_list.sort(key=str.lower)
    except Exception as e: logging.error(f"Error scanning clients in '{data_folder_path}': {e}", exc_info=True); flash("Error retrieving client list.", "error")
    return render_template('rvtools_upload.html', client_list=client_list)

@app.route('/get_projects/<client_name>')
def get_projects(client_name):
    """API endpoint (called by JS) to get project folders for a client."""
    safe_client_name = secure_filename(client_name)
    if not safe_client_name: logging.warning(f"Invalid client name requested in API: {client_name}"); return jsonify({"error": "Invalid client name"}), 400
    client_dir = os.path.join(app.config['DATA_FOLDER'], safe_client_name); project_list = []
    try:
        if not os.path.isdir(client_dir): return jsonify({"projects": []})
        items = os.listdir(client_dir); project_list = [item for item in items if os.path.isdir(os.path.join(client_dir, item)) and not item.startswith('.') and item not in ['Originals', 'Analyzed']]; project_list.sort(key=str.lower); logging.info(f"Found projects for client '{safe_client_name}': {project_list}"); return jsonify({"projects": project_list})
    except Exception as e: logging.error(f"API Error scanning projects for client '{safe_client_name}': {e}", exc_info=True); return jsonify({"error": "Server error scanning projects."}), 500

@app.route('/check_file_exists/<client_name>/<project_name>/<path:original_filename>')
def check_file_exists(client_name, project_name, original_filename):
    """API endpoint to check if original or analyzed file exists."""
    safe_client_name = secure_filename(client_name); safe_project_name = secure_filename(project_name)
    if not safe_client_name or not safe_project_name or not original_filename: return jsonify({"error": "Invalid input"}), 400
    try:
        safe_original_filename = secure_filename(original_filename); client_project_path = os.path.join(app.config['DATA_FOLDER'], safe_client_name, safe_project_name); original_filepath = os.path.join(client_project_path, 'Originals', safe_original_filename); base_name, ext = os.path.splitext(safe_original_filename); analyzed_filename = f"{base_name}-ANALYZED{ext}"; analyzed_filepath = os.path.join(client_project_path, 'Analyzed', analyzed_filename); original_exists = os.path.exists(original_filepath) and os.path.isfile(original_filepath); analyzed_exists = os.path.exists(analyzed_filepath) and os.path.isfile(analyzed_filepath); logging.info(f"API Check: C='{safe_client_name}', P='{safe_project_name}', F='{safe_original_filename}' -> OrigExists: {original_exists}, AnalyzedExists: {analyzed_exists}"); return jsonify({"original_exists": original_exists, "analyzed_exists": analyzed_exists})
    except Exception as e: logging.error(f"API file check error: C={safe_client_name}, P={safe_project_name}, F={original_filename}: {e}", exc_info=True); return jsonify({"error": "Server error checking file."}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handles upload, checks collision action, saves file, starts task."""
    form_redirect_target = 'rvtools_analysis'
    if 'rvtools_file' not in request.files: flash('No file part selected.', 'error'); return redirect(url_for(form_redirect_target))
    file = request.files['rvtools_file']; original_filename = file.filename
    if not original_filename or not allowed_file(original_filename): flash('No file selected or invalid file type (must be .xlsx).', 'error'); return redirect(url_for(form_redirect_target))
    selected_client = request.form.get('selected_client'); new_client_name = request.form.get('new_client_name', '').strip(); client_name_to_use = ""
    if selected_client == '_new' and new_client_name: client_name_to_use = new_client_name
    elif selected_client and selected_client != '_new': client_name_to_use = selected_client
    else: flash('Please select an existing client or provide a name for a new client.', 'error'); return redirect(url_for(form_redirect_target))
    safe_client_name = secure_filename(client_name_to_use)
    if not safe_client_name: flash(f'Invalid client name provided: "{html.escape(client_name_to_use)}". Please use valid characters.', 'error'); return redirect(url_for(form_redirect_target))
    selected_project = request.form.get('selected_project'); new_project_name = request.form.get('new_project_name', '').strip(); project_name_to_use = DEFAULT_PROJECT_NAME # Default
    if selected_project == '_new' and new_project_name: project_name_to_use = new_project_name
    elif selected_project and selected_project not in ['_new', '', '_none_selected', '_general_select']: project_name_to_use = selected_project
    elif selected_client == '_new' or selected_project == DEFAULT_PROJECT_NAME or selected_project == '_general_select': project_name_to_use = DEFAULT_PROJECT_NAME
    safe_project_name = secure_filename(project_name_to_use)
    if not safe_project_name: safe_project_name = DEFAULT_PROJECT_NAME; logging.warning(f"Project name '{project_name_to_use}' was invalid, using default '{DEFAULT_PROJECT_NAME}'.")
    collision_action = request.form.get('collision_action'); logging.info(f"Upload attempt: Client='{safe_client_name}', Project='{safe_project_name}', File='{original_filename}', CollisionAction='{collision_action}'")
    try: # Construct paths
        client_project_path = os.path.join(app.config['DATA_FOLDER'], safe_client_name, safe_project_name); upload_dir = os.path.join(client_project_path, 'Originals'); analyzed_dir = os.path.join(client_project_path, 'Analyzed'); safe_original_filename = secure_filename(original_filename); input_filepath = os.path.join(upload_dir, safe_original_filename); analyzed_base, analyzed_ext = os.path.splitext(safe_original_filename); analyzed_filename = f"{analyzed_base}-ANALYZED{analyzed_ext}"; analyzed_filepath = os.path.join(analyzed_dir, analyzed_filename); os.makedirs(upload_dir, exist_ok=True); os.makedirs(analyzed_dir, exist_ok=True)
    except Exception as e: logging.error(f"Error constructing paths for upload: {e}", exc_info=True); flash(f"Server error creating directories: {e}", 'error'); return redirect(url_for(form_redirect_target))
    original_exists = os.path.exists(input_filepath) and os.path.isfile(input_filepath); analyzed_exists = os.path.exists(analyzed_filepath) and os.path.isfile(analyzed_filepath); collision_detected = original_exists or analyzed_exists
    if collision_detected and collision_action != 'overwrite': flash(f'File "{html.escape(original_filename)}" or its analyzed version already exists. Upload cancelled as overwrite was not selected.', 'error'); logging.warning(f"Upload cancelled: Collision detected for '{safe_original_filename}' in {safe_client_name}/{safe_project_name} and overwrite denied."); return redirect(url_for(form_redirect_target))
    elif collision_detected and collision_action == 'overwrite': logging.warning(f"Upload Info: Collision detected for '{safe_original_filename}' in {safe_client_name}/{safe_project_name}. Proceeding with OVERWRITE.")
    try: # Save File & Store Task
        file.save(input_filepath); task_id = uuid.uuid4().hex[:16]; TASK_INFO[task_id] = {'original_filename': safe_original_filename, 'input_filepath': input_filepath,'output_folder': analyzed_dir, 'client_name': safe_client_name,'project_name': safe_project_name,'timestamp': time.time()}; logging.info(f"File '{safe_original_filename}' saved successfully. Task {task_id} created for {safe_client_name}/{safe_project_name}."); return redirect(url_for('processing_page', task_id=task_id))
    except Exception as e: logging.error(f"Error saving file '{safe_original_filename}' or storing task: {e}", exc_info=True); flash(f'Error saving file: {e}', 'error'); return redirect(url_for(form_redirect_target))

@app.route('/processing/<task_id>')
def processing_page(task_id):
    """Renders the page that will display SSE updates for a task."""
    task_details = TASK_INFO.get(task_id);
    if not task_details: flash("Task details not found or task already completed.", "error"); return redirect(url_for('index'))
    return render_template('processing.html', task_id=task_id, display_filename=task_details.get('original_filename', 'Unknown File'), client_name=task_details.get('client_name', 'Unknown Client'), project_name=task_details.get('project_name', 'Unknown Project'))

@app.route('/stream/<task_id>')
def stream(task_id):
    """Endpoint for the SSE stream, calls the backend processing function."""
    logging.info(f"SSE Stream requested for task_id: {task_id}")
    task_details = TASK_INFO.get(task_id)
    if not task_details:
        logging.error(f"SSE Stream Error: Task details not found for task_id {task_id}.")
        def error_stream_task_not_found(): yield f"event: result\ndata: {json.dumps({'success': False, 'message': 'Task details not found. It might have already completed or failed.'})}\n\n";
        return Response(error_stream_task_not_found(), mimetype='text/event-stream')
    input_filepath = task_details['input_filepath']; original_basename = task_details['original_filename']
    output_folder = task_details['output_folder']; safe_client_name = task_details['client_name']
    safe_project_name = task_details['project_name']; logging.info(f"SSE Stream: Task {task_id} details retrieved. Input: '{input_filepath}', Output Folder: '{output_folder}'")
    if not os.path.exists(input_filepath) or not os.path.isfile(input_filepath):
        logging.error(f"SSE Stream Error: Input file '{input_filepath}' missing for task {task_id}.")
        def error_stream_file_missing(): yield f"event: result\ndata: {json.dumps({'success': False, 'message': f'Input file {html.escape(original_basename)} is missing.'})}\n\n";
        if task_id in TASK_INFO: TASK_INFO.pop(task_id);
        return Response(error_stream_file_missing(), mimetype='text/event-stream')

    # --- Event Stream Generator (Restructured V3 - No url_for) ---
    def event_stream():
        logging.info(f"SSE Stream: Starting processing generator for task {task_id}")
        final_payload_str = None
        processed_event_data_str = None
        generator_error = None

        # --- Phase 1: Execute generator and yield progress ---
        try:
            generator = process_rvtools_file(input_filepath, output_folder, original_basename)
            for msg in generator:
                if msg.startswith("event: result"):
                    try: final_payload_str = msg.split("data: ", 1)[1].rsplit("\n\n", 1)[0]; logging.info(f"SSE Stream: Intercepted final result event string for task {task_id}")
                    except IndexError: logging.error(f"SSE Stream Error: Malformed result event received for task {task_id}: {msg}"); generator_error = Exception("Malformed result event received from backend processing."); final_payload_str = None; break
                else: yield msg
        except Exception as e: logging.error(f"SSE Stream FATAL Error during processing task {task_id}: {e}", exc_info=True); generator_error = e

        # --- Phase 2: Process result/error ---
        try:
            if generator_error:
                error_payload = {"success": False, "message": f"Fatal server error during processing: {generator_error}"}
                processed_event_data_str = json.dumps(error_payload)
            elif final_payload_str:
                try: # Try parsing the final payload from backend
                    payload = json.loads(final_payload_str)
                    if payload.get("success"):
                        analyzed_filename = payload.get("message")
                        if not analyzed_filename: raise ValueError("Analyzed filename missing from successful result payload.")
                        # *** REMOVED url_for call ***
                        # Just send back the necessary info for JS to build the URL
                        success_data = {
                            "success": True,
                            "client_name": safe_client_name,
                            "project_name": safe_project_name,
                            "analyzed_filename": analyzed_filename
                            # "results_url" key is removed
                        }
                        processed_event_data_str = json.dumps(success_data)
                        logging.info(f"SSE Stream: Processing successful for task {task_id}. Analyzed file: '{analyzed_filename}'")
                    else:
                        # Backend failure, forward payload string
                        processed_event_data_str = final_payload_str
                        logging.warning(f"SSE Stream: Processing failed for task {task_id} (reported by backend): {payload.get('message')}")
                except Exception as e: # Error processing final payload
                    logging.error(f"SSE Stream Error: Error processing final result payload for task {task_id}: {e}", exc_info=True)
                    processed_event_data_str = json.dumps({"success": False, "message": f"Error processing final result data: {e}"})
            else: # No error and no payload
                if processed_event_data_str is None:
                    logging.warning(f"SSE Stream Warning: Backend generator finished for task {task_id} without yielding a valid final result event.")
                    processed_event_data_str = json.dumps({'success': False, 'message': 'Processing script finished unexpectedly without a final result.'})

            # --- Phase 3: Yield final event ---
            yield f"event: result\ndata: {processed_event_data_str}\n\n"

        except Exception as final_proc_err: # Catch errors in phase 2 logic
            logging.error(f"SSE Stream Error: Error during final event processing phase for task {task_id}: {final_proc_err}", exc_info=True)
            fallback_error = json.dumps({"success": False, "message": f"Server error preparing final result: {final_proc_err}"})
            yield f"event: result\ndata: {fallback_error}\n\n"
        # --- Phase 4: Cleanup ---
        finally:
             if task_id in TASK_INFO:
                 task_removed = TASK_INFO.pop(task_id, None);
                 if task_removed: logging.info(f"SSE Stream: Cleaned up task info for completed/failed task {task_id}.")
                 else: logging.warning(f"SSE Stream: Task info for {task_id} was already cleaned up.")

    # Configure and return the Response object for SSE
    response=Response(event_stream(),mimetype='text/event-stream');
    response.headers['Cache-Control']='no-cache';
    response.headers['X-Accel-Buffering'] = 'no';
    response.headers['Connection'] = 'keep-alive';
    return response

# Adjusted Download Route
@app.route('/download/<path:filepath>')
def download_file_relpath(filepath):
    """Downloads a file given a path relative to the DATA_FOLDER."""
    data_folder_abs = os.path.abspath(app.config['DATA_FOLDER']); requested_path_abs = os.path.normpath(os.path.join(data_folder_abs, filepath)); logging.info(f"Download request for relative path: {filepath}"); logging.info(f"Absolute data folder: {data_folder_abs}"); logging.info(f"Requested absolute path: {requested_path_abs}")
    if not requested_path_abs.startswith(data_folder_abs): logging.error(f"Download Forbidden: Path traversal attempt detected. Requested: '{requested_path_abs}', Base: '{data_folder_abs}'"); abort(403)
    try: directory = os.path.dirname(filepath); filename = os.path.basename(filepath); logging.info(f"Serving file '{filename}' from directory relative to DATA_FOLDER: '{directory}'"); return send_from_directory(directory=os.path.join(data_folder_abs, directory), path=filename, as_attachment=True)
    except FileNotFoundError: logging.error(f"Download Error: File not found at '{requested_path_abs}'"); abort(404)
    except Exception as e: logging.error(f"Download Error: {e}", exc_info=True); abort(500)


# ================================================================
# --- Results & Browse Routes (Based on Original File) ---
# ================================================================

@app.route('/results/<client_name>/<project_name>/<path:filename_in_url>')
def view_results(client_name, project_name, filename_in_url):
    """
    Displays analysis results.
    Determines if -ANALYZED and -EDITED versions exist for download buttons.
    """
    logging.info(f"Results Request: C='{client_name}', P='{project_name}', Displaying File='{filename_in_url}'")
    safe_client_name = secure_filename(client_name)
    safe_project_name = secure_filename(project_name)
    safe_filename_in_url = secure_filename(filename_in_url)

    if not safe_client_name or not safe_project_name or not safe_filename_in_url:
        flash("Invalid client, project, or filename specified.", "error")
        return redirect(url_for('browse_existing'))

    summary_data_json = {}
    expected_sheets = ['Overall Summary Totals', 'Overall Powerstate Counts', 'Datacenter Summary Combined']
    
    data_folder_abs = os.path.abspath(app.config['DATA_FOLDER'])
    analyzed_dir_abs = os.path.abspath(os.path.join(data_folder_abs, safe_client_name, safe_project_name, 'Analyzed'))

    # Determine the true original base name
    current_file_base = ""
    if safe_filename_in_url.endswith("-ANALYZED.xlsx"):
        current_file_base = safe_filename_in_url[:-len("-ANALYZED.xlsx")]
    elif safe_filename_in_url.endswith("-EDITED.xlsx"):
        current_file_base = safe_filename_in_url[:-len("-EDITED.xlsx")]
    else:
        # This case should ideally not happen if files are always -ANALYZED or -EDITED
        # If it's an original file name, we might not have an "analyzed" or "edited" version yet.
        # For robustness, let's try to derive a base, assuming it might be an original upload name.
        current_file_base, _ = os.path.splitext(safe_filename_in_url)
        logging.warning(f"Filename '{safe_filename_in_url}' does not have a recognized suffix. Assuming base is '{current_file_base}'.")


    # Define specific filenames for -ANALYZED and -EDITED versions
    analyzed_version_name_on_disk = f"{current_file_base}-ANALYZED.xlsx"
    edited_version_name_on_disk = f"{current_file_base}-EDITED.xlsx"

    # Path to the file whose data is actually being displayed (from the URL)
    path_to_file_being_viewed = os.path.normpath(os.path.join(analyzed_dir_abs, safe_filename_in_url))

    # Check existence of the file specified in the URL for viewing
    if not os.path.isfile(path_to_file_being_viewed) or not os.path.abspath(path_to_file_being_viewed).startswith(analyzed_dir_abs):
        logging.error(f"Results Error: File to view not found or invalid path: {path_to_file_being_viewed}")
        flash(f"File '{html.escape(safe_filename_in_url)}' not found.", "error")
        return redirect(url_for('browse_existing'))

    logging.info(f"Results: Displaying data from: {path_to_file_being_viewed}")

    # Check for the presence of the -ANALYZED version for its download button
    path_to_analyzed_on_disk = os.path.join(analyzed_dir_abs, analyzed_version_name_on_disk)
    show_download_analyzed_button = os.path.isfile(path_to_analyzed_on_disk)

    # Check for the presence of the -EDITED version for its download button
    path_to_edited_on_disk = os.path.join(analyzed_dir_abs, edited_version_name_on_disk)
    show_download_edited_button = os.path.isfile(path_to_edited_on_disk)

    # Read data from the file specified in the URL
    try:
        excel_data = pd.read_excel(path_to_file_being_viewed, sheet_name=expected_sheets, engine='openpyxl', skipfooter=1)
        for sheet_name in expected_sheets:
            if sheet_name in excel_data:
                df = excel_data[sheet_name]
                summary_data_json[sheet_name] = df.fillna('').astype(str).to_json(orient='records', date_format='iso')
            else:
                summary_data_json[sheet_name] = "[]"; logging.warning(f"Sheet '{sheet_name}' not found in {safe_filename_in_url}.")
    except Exception as e:
        logging.error(f"Results Error: Reading/converting Excel '{path_to_file_being_viewed}': {e}", exc_info=True)
        flash(f"Error reading results from '{html.escape(safe_filename_in_url)}'.", "error")
        for sheet_name in expected_sheets: summary_data_json[sheet_name] = "[]"

    return render_template('results.html',
                           client_name=safe_client_name,
                           project_name=safe_project_name,
                           displayed_filename=safe_filename_in_url, # The file whose data is shown
                           analyzed_version_filename_for_download=analyzed_version_name_on_disk if show_download_analyzed_button else None,
                           edited_version_filename_for_download=edited_version_name_on_disk if show_download_edited_button else None,
                           summary_data_json=summary_data_json)


@app.route('/browse')
def browse_existing():
    """
    Renders the single-page browse interface.
    Can now accept 'client' and 'project' query parameters for pre-selection.
    """
    logging.info("Browse Request: Loading single browse page.")
    
    # Get pre-selection parameters from URL query if they exist
    preselected_client = request.args.get('client')
    preselected_project = request.args.get('project')
    
    logging.info(f"Browse page pre-selection: Client='{preselected_client}', Project='{preselected_project}'")

    client_list = []
    data_folder_path = app.config['DATA_FOLDER']
    try:
        if os.path.exists(data_folder_path):
            items = os.listdir(data_folder_path)
            client_list = [ i for i in items if os.path.isdir(os.path.join(data_folder_path, i)) and not i.startswith('.') ]
            client_list.sort(key=str.lower)
    except Exception as e:
        logging.error(f"Browse Error getting client list: {e}", exc_info=True)
        flash("Error listing clients.", "error")
        client_list = []
        
    return render_template('browse_existing.html', 
                           client_list=client_list,
                           preselected_client=preselected_client,
                           preselected_project=preselected_project)

@app.route('/get_files/<client_name>/<project_name>') # API
def get_files(client_name, project_name):
    """API endpoint called by JS to populate files for a selected project.
       Now includes both -ANALYZED.xlsx and -EDITED.xlsx files.
    """
    safe_client_name = secure_filename(client_name)
    safe_project_name = secure_filename(project_name)

    if not safe_client_name or not safe_project_name:
        logging.warning(f"Invalid client/project name for get_files API: C={client_name}, P={project_name}")
        return jsonify({"error": "Invalid client/project name"}), 400

    # Construct paths
    project_dir = os.path.join(app.config['DATA_FOLDER'], safe_client_name, safe_project_name)
    originals_dir = os.path.join(project_dir, 'Originals')
    analyzed_dir = os.path.join(project_dir, 'Analyzed') # This directory holds both -ANALYZED and -EDITED files

    originals_list = []
    processed_files_list = [] # Renamed from analyzed_list for clarity

    # Scan Originals
    try:
        if os.path.isdir(originals_dir):
            items = os.listdir(originals_dir)
            originals_list = [f for f in items if os.path.isfile(os.path.join(originals_dir, f)) and not f.startswith('.')]
            originals_list.sort(key=str.lower)
    except Exception as e:
        logging.error(f"API Error scanning Originals for {safe_client_name}/{safe_project_name}: {e}", exc_info=True)

    # Scan Analyzed directory for both -ANALYZED.xlsx and -EDITED.xlsx files
    try:
        if os.path.isdir(analyzed_dir):
            items = os.listdir(analyzed_dir)
            processed_files_list = [
                f for f in items if
                os.path.isfile(os.path.join(analyzed_dir, f)) and
                not f.startswith('.') and
                (f.lower().endswith('-analyzed.xlsx') or f.lower().endswith('-edited.xlsx')) # Include both
            ]
            processed_files_list.sort(key=str.lower)
    except Exception as e:
        logging.error(f"API Error scanning Processed files (Analyzed/Edited) for {safe_client_name}/{safe_project_name}: {e}", exc_info=True)

    logging.info(f"API get_files for {safe_client_name}/{safe_project_name}: Found {len(originals_list)} Originals, {len(processed_files_list)} Processed (Analyzed/Edited).")
    # Return under the key 'analyzed' for compatibility with existing JS, or change JS if preferred
    return jsonify({"originals": originals_list, "analyzed": processed_files_list})


# ================================================================
# --- Edit and Update Routes ---
# ================================================================

@app.route('/edit/<client>/<project>/<filename_from_url>') # Renamed last param for clarity
def edit_summary(client, project, filename_from_url):
    """Displays the editable summary data grid."""
    logging.info(f"EDIT_SUMMARY Route: Received raw client='{client}', project='{project}', filename_from_url='{filename_from_url}'")
    
    safe_client = secure_filename(client)
    safe_project = secure_filename(project)
    safe_filename = secure_filename(filename_from_url) # Filename from URL after securing
    
    logging.info(f"EDIT_SUMMARY Route: Secured params: client='{safe_client}', project='{safe_project}', filename='{safe_filename}'")

    # Check the secured filename for suffixes
    is_analyzed = safe_filename.lower().endswith('-analyzed.xlsx')
    is_edited = safe_filename.lower().endswith('-edited.xlsx')
    logging.info(f"EDIT_SUMMARY Route: Suffix checks: is_analyzed={is_analyzed}, is_edited={is_edited} for '{safe_filename}'")


    if not safe_client or not safe_project or not safe_filename or not (is_analyzed or is_edited):
        flash("Invalid client, project, or filename for editing.", "error")
        logging.warning(f"EDIT_SUMMARY Route: Invalid edit request. Raw: C='{client}', P='{project}', F='{filename_from_url}'. Secured: C='{safe_client}', P='{safe_project}', F='{safe_filename}'. Condition (is_analyzed or is_edited) is False.")
        return redirect(url_for('browse_existing'))
    
    analyzed_folder_abs = os.path.abspath(os.path.join(app.config['DATA_FOLDER'], safe_client, safe_project, 'Analyzed'))
    filepath = os.path.join(analyzed_folder_abs, safe_filename) # Use the secured filename to build the path
    data_folder_abs = os.path.abspath(app.config['DATA_FOLDER'])

    if not os.path.abspath(filepath).startswith(data_folder_abs):
        logging.error(f"Edit Forbidden: Path traversal attempt detected. Path: '{filepath}'")
        abort(403) # Path traversal
    if not os.path.exists(filepath) or not os.path.isfile(filepath):
        logging.error(f"Analyzed file not found for editing at {filepath}")
        flash(f"Error: File '{html.escape(safe_filename)}' not found for editing.", "error")
        return redirect(url_for('browse_existing'))
    
    vsummary_data_json = "[]" # Default
    try:
        logging.info(f"Reading vSummary sheet from {filepath} for editing...")
        df_summary = pd.read_excel(filepath, sheet_name='vSummary', engine='openpyxl')
        if df_summary.empty:
            logging.warning(f"vSummary sheet in {safe_filename} is empty.")
            flash(f"Warning: 'vSummary' sheet in '{html.escape(safe_filename)}' is empty. Cannot edit.", "warning")
            return redirect(url_for('view_results', client_name=safe_client, project_name=safe_project, filename_in_url=safe_filename))
        else:
            df_summary.fillna('', inplace=True)
            for col in df_summary.select_dtypes(include=['datetime64[ns]']).columns:
                if col in df_summary.columns: df_summary[col] = df_summary[col].astype(str)
            vsummary_data_json = df_summary.to_json(orient='records', date_format='iso')
            logging.info(f"Successfully read and converted vSummary sheet ({len(df_summary)} rows) for editing.")
    except Exception as e:
        logging.error(f"Error reading vSummary sheet from {filepath} for editing: {e}", exc_info=True)
        flash(f"Error reading data from '{html.escape(safe_filename)}' for editing: {e}", "error")
        return redirect(url_for('view_results', client_name=safe_client, project_name=safe_project, filename_in_url=safe_filename))
    
    # Pass safe_filename as 'analyzed_filename' to the template, as that's what the template expects for context
    return render_template('edit_summary.html',
                           client_name=safe_client,
                           project_name=safe_project,
                           analyzed_filename=safe_filename,
                           vsummary_data_json=vsummary_data_json)


# --- Helper function for Update Route ---
def update_flags_from_workload(edited_workload, edited_environment):
    """
    Determines the correct 'Is...' flag values based on the edited Workload and Environment.
    Aligns with the logic expected by the provided RVToolsAnalysis_web.py categorization helpers.
    """
    flags = {'IsFile': 'No', 'IsSQL': 'No', 'IsOrcl': 'No', 'IsPGres': 'No', 'IsExch': 'No', 'IsTestDev': 'No'}
    if edited_workload == 'File': flags['IsFile'] = 'Yes'
    elif edited_workload == 'SQL': flags['IsSQL'] = 'Yes'
    elif edited_workload == 'Oracle': flags['IsOrcl'] = 'Yes'
    elif edited_workload == 'Postgres': flags['IsPGres'] = 'Yes'
    elif edited_workload == 'Exchange': flags['IsExch'] = 'Yes'
    flags['IsTestDev'] = 'Yes' if edited_environment == 'Test/Dev' else 'No'
    logging.debug(f"Updated flags based on Workload='{edited_workload}', Env='{edited_environment}': {flags}")
    return flags


# --- Update Route ---
@app.route('/update/<client>/<project>/<source_filename_param>', methods=['POST'])
def update_summary(client, project, source_filename_param):
    safe_client = secure_filename(client)
    safe_project = secure_filename(project)
    safe_source_filename = secure_filename(source_filename_param) # This is the file being EDITED
    log_prefix = f"Update Request [{safe_client}/{safe_project}/{safe_source_filename}]:"
    logging.info(f"{log_prefix} Received.")

    if not safe_client or not safe_project or not safe_source_filename or \
       not (safe_source_filename.endswith('-ANALYZED.xlsx') or safe_source_filename.endswith('-EDITED.xlsx')):
        logging.error(f"{log_prefix} Invalid request parameters. Source filename: {safe_source_filename}")
        return jsonify({"success": False, "message": "Invalid client, project, or source filename format."}), 400

    analyzed_folder_abs = os.path.abspath(os.path.join(app.config['DATA_FOLDER'], safe_client, safe_project, 'Analyzed'))
    source_filepath = os.path.join(analyzed_folder_abs, safe_source_filename) # Path to the file we are reading from
    data_folder_abs = os.path.abspath(app.config['DATA_FOLDER'])

    if not os.path.abspath(source_filepath).startswith(data_folder_abs):
        logging.error(f"{log_prefix} Path traversal attempt. Source: '{source_filepath}'")
        return jsonify({"success": False, "message": "Invalid source file path."}), 403
    if not os.path.isfile(source_filepath):
        logging.error(f"{log_prefix} Source file not found at {source_filepath}")
        return jsonify({"success": False, "message": f"Source file '{html.escape(safe_source_filename)}' not found."}), 404

    # --- Determine the output filename (always ...-EDITED.xlsx) ---
    base_name_for_output = ""
    if safe_source_filename.endswith("-ANALYZED.xlsx"):
        base_name_for_output = safe_source_filename[:-len("-ANALYZED.xlsx")]
    elif safe_source_filename.endswith("-EDITED.xlsx"):
        base_name_for_output = safe_source_filename[:-len("-EDITED.xlsx")]
    # No 'else' needed due to the check above, but ensure base_name_for_output is not empty
    if not base_name_for_output: # Should ideally not be reached
        logging.error(f"{log_prefix} Could not determine base name from {safe_source_filename}")
        return jsonify({"success": False, "message": "Could not determine output filename base."}), 500

    output_filename = f"{base_name_for_output}-EDITED.xlsx"
    output_filepath = os.path.join(analyzed_folder_abs, output_filename) # Path to the file we are writing to
    logging.info(f"{log_prefix} Source: {safe_source_filename}, Determined Output: {output_filename}")

    try:
        edited_data = request.get_json()
        if not isinstance(edited_data, list): raise ValueError("Invalid data format received.")
        logging.info(f"{log_prefix} Received {len(edited_data)} rows of edited data.")
    except Exception as e:
        logging.error(f"{log_prefix} Failed to parse JSON request body: {e}", exc_info=True)
        return jsonify({"success": False, "message": f"Error parsing request data: {e}"}), 400

    workbook = None
    try:
        logging.info(f"{log_prefix} Loading data from source workbook: {source_filepath}")
        all_sheets_dict = pd.read_excel(source_filepath, sheet_name=None, engine='openpyxl')
        if 'vSummary' not in all_sheets_dict:
            logging.error(f"{log_prefix} 'vSummary' sheet missing in source file.");
            return jsonify({"success": False, "message": "'vSummary' sheet missing in the source file."}), 500
        
        original_vsummary_df = all_sheets_dict['vSummary']
        # Preserve other sheets from the source file being edited
        other_sheets_to_preserve = {name: df for name, df in all_sheets_dict.items() if name != 'vSummary'}
        logging.info(f"{log_prefix} Loaded {len(original_vsummary_df)} rows from vSummary and {len(other_sheets_to_preserve)} other sheets from {safe_source_filename}.")

        if not edited_data:
            logging.warning(f"{log_prefix} Received empty data array. No updates to apply to vSummary. File will be saved as {output_filename} with original vSummary.")
            # If we still want to create/overwrite the -EDITED file even with no vSummary changes:
            updated_vsummary_df = original_vsummary_df.copy()
        else:
            edited_df = pd.DataFrame(edited_data)
            required_edit_cols = ['VM', 'Workload', 'Environment']
            if not all(col in edited_df.columns for col in required_edit_cols):
                missing = [c for c in required_edit_cols if c not in edited_df.columns]
                logging.error(f"{log_prefix} Missing required columns in submitted data: {missing}")
                return jsonify({"success": False, "message": f"Submitted data is missing columns: {missing}"}), 400
            
            updated_vsummary_df = original_vsummary_df.copy()
            if 'VM' not in updated_vsummary_df.columns:
                logging.error(f"{log_prefix} Original vSummary is missing 'VM' column.")
                return jsonify({"success": False, "message": "vSummary data integrity issue: missing VM column."}), 500
            
            updated_vsummary_df.set_index('VM', inplace=True)
            edited_df.set_index('VM', inplace=True)
            updated_vms_count = 0
            flag_cols_to_update = ['IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev']
            
            for vm_name, row_data_from_js in edited_df.iterrows():
                if vm_name in updated_vsummary_df.index:
                    new_workload = row_data_from_js['Workload']
                    new_environment = row_data_from_js['Environment']
                    updated_vsummary_df.loc[vm_name, 'Workload'] = new_workload
                    updated_vsummary_df.loc[vm_name, 'Environment'] = new_environment
                    updated_flags = update_flags_from_workload(new_workload, new_environment)
                    for flag_col in flag_cols_to_update:
                        if flag_col in updated_vsummary_df.columns:
                            updated_vsummary_df.loc[vm_name, flag_col] = updated_flags.get(flag_col, 'No')
                    updated_vms_count += 1
                else:
                    logging.warning(f"{log_prefix} VM '{html.escape(str(vm_name))}' from JS data not found in vSummary. Skipping update.")
            updated_vsummary_df.reset_index(inplace=True)
            logging.info(f"{log_prefix} Applied updates to {updated_vms_count} VMs in DataFrame.")

        logging.info(f"{log_prefix} Recalculating summaries for file: {output_filename}")
        if 'VM Count' not in updated_vsummary_df.columns and 'VM' in updated_vsummary_df.columns:
            updated_vsummary_df['VM Count'] = 1
        recalculated_summary_dfs = calculate_summaries(updated_vsummary_df.copy())
        if not recalculated_summary_dfs:
            logging.error(f"{log_prefix} Summary recalculation failed.")
            return jsonify({"success": False, "message": "Failed to recalculate summaries after update."}), 500
        logging.info(f"{log_prefix} Summaries recalculated.")
        
        logging.info(f"{log_prefix} Writing updated data to: {output_filepath}")
        with pd.ExcelWriter(output_filepath, engine='openpyxl', mode='w') as writer:
            updated_vsummary_df.to_excel(writer, sheet_name='vSummary', index=False)
            for sheet_name, df_summary in recalculated_summary_dfs.items():
                df_summary.to_excel(writer, sheet_name=sheet_name, index=False)
            # Preserve other sheets
            summary_sheet_names = list(recalculated_summary_dfs.keys())
            for sheet_name, df_other in other_sheets_to_preserve.items():
                 if sheet_name not in ['vSummary'] + summary_sheet_names: # Avoid overwriting vSummary or new summaries
                      df_other.to_excel(writer, sheet_name=sheet_name, index=False)
        logging.info(f"{log_prefix} Data written to {output_filename}")
        
        logging.info(f"{log_prefix} Applying Openpyxl formatting to {output_filename}...")
        try:
            workbook = xl.load_workbook(output_filepath)
            sheets_for_subtotals = {'Overall Summary Totals': {'cols': [3, 4, 5], 'label_col': 1}, 'Overall Powerstate Counts': {'cols': [4], 'label_col': 1}, 'Datacenter Summary Combined': {'cols': [5, 6, 7, 8, 9, 10], 'label_col': 1}}; total_row_font = xl.styles.Font(bold=True); subtotal_added_count = 0
            for sheet_name, config in sheets_for_subtotals.items():
                 if sheet_name in workbook.sheetnames:
                      ws = workbook[sheet_name]; max_row = ws.max_row
                      if max_row > 1:
                           total_row_idx = max_row + 1; label_cell = ws.cell(row=total_row_idx, column=config['label_col']); label_cell.value = "Grand Total"; label_cell.font = total_row_font
                           for col_idx in config['cols']:
                                col_letter = xl.utils.get_column_letter(col_idx); formula = f"=SUBTOTAL(9,{col_letter}2:{col_letter}{max_row})"; total_cell = ws.cell(row=total_row_idx, column=col_idx); total_cell.value = formula; total_cell.font = total_row_font
                                if max_row >= 2: prev_cell = ws.cell(row=max_row, column=col_idx);
                                if prev_cell.number_format: total_cell.number_format = prev_cell.number_format
                           subtotal_added_count += 1
            logging.info(f"{log_prefix} Added SUBTOTAL rows to {subtotal_added_count} sheets in {output_filename}.")
            
            # Reorder sheets: Summary sheets, then vSummary, then others from original
            final_sheet_order = ['Overall Summary Totals', 'Overall Powerstate Counts', 'Datacenter Summary Combined', 'vSummary']
            current_workbook_sheets = workbook.sheetnames
            for sheet_name in other_sheets_to_preserve.keys():
                if sheet_name not in final_sheet_order and sheet_name in current_workbook_sheets:
                    final_sheet_order.append(sheet_name)
            # Append any other sheets that might be in workbook but not in others_to_preserve (should not happen with mode='w')
            for s_name in current_workbook_sheets:
                if s_name not in final_sheet_order: final_sheet_order.append(s_name)

            workbook._sheets = sorted(workbook._sheets, key=lambda ws_sort: final_sheet_order.index(ws_sort.title))
            logging.info(f"{log_prefix} Sheets reordered in {output_filename}.")

            filter_rows(workbook); logging.info(f"{log_prefix} Filters applied to {output_filename}.")
            workbook.save(output_filepath); logging.info(f"{log_prefix} Final formatting applied and {output_filename} saved.")
        except Exception as fmt_err:
            logging.error(f"{log_prefix} Error during openpyxl formatting for {output_filename}: {fmt_err}", exc_info=True)
            return jsonify({"success": True, "filename": output_filename, "message": "Data saved to new -EDITED file, but final formatting failed. Please check the file."})
        finally:
             if workbook: workbook.close()
        
        logging.info(f"{log_prefix} Update process completed successfully. New file: {output_filename}")
        return jsonify({"success": True, "filename": output_filename, "message": f"Data saved to {output_filename} and summaries recalculated."})

    except Exception as e:
        logging.error(f"{log_prefix} Unhandled error during update process: {e}", exc_info=True)
        return jsonify({"success": False, "message": f"An unexpected server error occurred: {e}"}), 500


# ================================================================
# --- Run the App ---
# ================================================================
if __name__ == '__main__':
    logging.info("--- Starting Datacenter Data Analyzer ---")
    logging.info(f"Flask App Name: {app.name}")
    logging.info(f"Base Directory: {BASE_DIR}")
    logging.info(f"Data Folder: {app.config['DATA_FOLDER']}")
    # Set debug=False for production deployments! Use host='0.0.0.0' for external access.
    app.run(host='0.0.0.0', port=5000, debug=True) # Adjust host/port/debug as needed