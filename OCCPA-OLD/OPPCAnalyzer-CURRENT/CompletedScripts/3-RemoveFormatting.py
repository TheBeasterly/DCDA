import os
import sys
import openpyxl as xl
import tkinter as tk
from tkinter import filedialog

def main():
    # Get the Excel file name from the user
    root = tk.Tk()
    root.withdraw()

    srcfile_name = filedialog.askopenfilename()

    # Open the Excel file
    srcfile = xl.load_workbook(srcfile_name)
    #print(srcfile.sheetnames)

    # Create a new Excel file
    destfile_name = srcfile_name[:-5] + "-EDITED.xlsx"
    srcfile.save(destfile_name)

    # Open new Excel file and remove extra sheets
    destfile = xl.load_workbook(destfile_name)
    #print(destfile.sheetnames)
    
    keep_sheets = ['vInfo', 'vDisk', 'vPartition']
    for sheetName in destfile.sheetnames:
        if sheetName not in keep_sheets:
            del destfile[sheetName]
    destfile.save(destfile_name)

    # Remove Formatting
    def removeFormatting(destfile):
      """Removes all formatting from an excel workbook.

      Args:
        destfile: An xl workbook object.
      """

      for worksheet in destfile.worksheets:
        for cell in worksheet.iter_rows():
          for c in cell:
            c.style = 'Normal'

    destfile = xl.load_workbook(destfile_name)
    removeFormatting(destfile)
    destfile.save(destfile_name)

if __name__ == "__main__":
    main()