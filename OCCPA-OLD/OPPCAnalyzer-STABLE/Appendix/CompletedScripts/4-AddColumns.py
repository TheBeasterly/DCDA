import os
import sys
import openpyxl as xl
import tkinter as tk
from tkinter import filedialog

def main():
    # Get the Excel file name from the user
    root = tk.Tk()
    root.withdraw()

    srcfile_name = filedialog.askopenfilename()

    # Open the Excel file
    srcfile = xl.load_workbook(srcfile_name)
    #print(srcfile.sheetnames)

    # Create a new Excel file
    destfile_name = srcfile_name[:-5] + "-EDITED.xlsx"
    srcfile.save(destfile_name)

    # Open new Excel file and remove extra sheets
    destfile = xl.load_workbook(destfile_name)
    #print(destfile.sheetnames)
    
    keep_sheets = ['vInfo', 'vDisk', 'vPartition']
    for sheetName in destfile.sheetnames:
        if sheetName not in keep_sheets:
            del destfile[sheetName]
    destfile.save(destfile_name)

    # Remove Formatting
    def removeFormatting(destfile):
      """Removes all formatting from an excel workbook.

      Args:
        destfile: An xl workbook object.
      """

      for worksheet in destfile.worksheets:
        for cell in worksheet.iter_rows():
          for c in cell:
            c.style = 'Normal'

    destfile = xl.load_workbook(destfile_name)
    removeFormatting(destfile)
    destfile.save(destfile_name)

    # Add Columns
    destfile = xl.load_workbook(destfile_name)

    for worksheet in destfile.worksheets:
      worksheet.insert_cols(2, 6)

    destfile.save(destfile_name)

    # Rename columns B:G in all sheets to "IsFile", "IsSQL", "IsOrcl", "IsPGres", "IsExch", and "IsTestDev" 
    destfile = xl.load_workbook(destfile_name)

    for worksheet in destfile.worksheets:
      worksheet['B1'] = 'IsFile'
      worksheet['C1'] = 'IsSQL'
      worksheet['D1'] = 'IsOrcl'
      worksheet['E1'] = 'IsPGres'
      worksheet['F1'] = 'IsExch'
      worksheet['G1'] = 'IsTestDev'

    # Insert "HasTools" Column in vInfo Sheet
    destfile["vInfo"].insert_cols(8, 1)
    destfile["vInfo"]['H1'] = 'HasTools'

    # Insert "DiskCount" Column in vDisk Sheet
    destfile["vDisk"].insert_cols(8, 1)
    destfile["vDisk"]['H1'] = 'DiskCount'

    destfile.save(destfile_name)

if __name__ == "__main__":
    main()