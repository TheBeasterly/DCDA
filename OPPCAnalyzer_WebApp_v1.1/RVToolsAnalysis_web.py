# RVToolsAnalysis_web.py
import os
import sys
import openpyxl as xl
from openpyxl.styles import PatternFill # Keep this if removeFormatting needs it
import pandas as pd
import traceback
import time # Added for SSE delays
import json # Added for SSE final message

# =================================================================================================================================================================================================================================================
# ============================================================================================= ALL ORIGINAL HELPER FUNCTIONS FROM RVToolsAnalysis.py =============================================================================================
# =================================================================================================================================================================================================================================================

## Remove Formatting
def removeFormatting(destfile):
  """Removes all formatting from an excel workbook.

  Args:
    destfile: An xl workbook object.
  """
  try:
    for worksheet in destfile.worksheets:
      # Iterate over all cells that have data or formatting
      for row in worksheet.iter_rows():
        for cell in row:
            # Resetting style explicitly might be needed depending on openpyxl version
            # Using 'Normal' style assumes it exists. If not, other methods might be needed.
            # A safer approach might be to reset individual properties if 'Normal' fails.
             if cell.has_style:
                cell.font = xl.styles.Font()
                cell.border = xl.styles.Border()
                cell.fill = xl.styles.PatternFill()
                cell.number_format = 'General' # Reset number format
                cell.protection = xl.styles.Protection()
                cell.alignment = xl.styles.Alignment()
                # Explicitly set style attribute if available and necessary
                # cell.style = 'Normal' # This can sometimes cause issues if 'Normal' isn't defined
  except Exception as e:
    print(f"Warning: Error during removeFormatting: {e}")
    # Continue processing even if formatting removal fails partially

## Match File Servers
def match_fs(destfile):
  """Matches file servers in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  fs_str = ["file", "fs", "nas", "share", "ftp"]
  try:
    for worksheet in destfile.worksheets:
       # Check if column A exists and has values
       if worksheet.max_column < 1: continue # Skip empty sheets

       for cell in worksheet["A"]: # Iterate through cells in column A
          if cell.row == 1: continue # Skip header row
          if cell.value is None or not isinstance(cell.value, str): continue # Skip empty/non-string cells

          cell_value_lower = cell.value.lower()
          match_found = False
          for search_string in fs_str:
            if search_string in cell_value_lower: # Simple substring check
              match_found = True
              break

          if match_found:
            # Assuming 'IsFile' is now in Column C (index 3)
            if worksheet.max_column >= 3:
                worksheet.cell(row=cell.row, column=3).value = "Yes"
            else:
                print(f"Warning: Column C (IsFile) not found in sheet '{worksheet.title}' for row {cell.row}")

  except Exception as e:
      print(f"Error during match_fs: {e}")
      traceback.print_exc()


## Match SQL DBs
def match_sql(destfile):
  """Matches SQL DBs in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  sql_str = ["sql"]
  try:
    for worksheet in destfile.worksheets:
       if worksheet.max_column < 1: continue

       for cell in worksheet["A"]:
          if cell.row == 1: continue
          if cell.value is None or not isinstance(cell.value, str): continue

          cell_value_lower = cell.value.lower()
          match_found = False
          for search_string in sql_str:
            if search_string in cell_value_lower:
              match_found = True
              break

          if match_found:
            # Assuming 'IsSQL' is now in Column D (index 4)
            if worksheet.max_column >= 4:
                worksheet.cell(row=cell.row, column=4).value = "Yes"
            else:
                 print(f"Warning: Column D (IsSQL) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_sql: {e}")
      traceback.print_exc()


## Match Oracle DBs
def match_orcl(destfile):
  """Matches Oracle DBs in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  orcl_str = ["orcl", "oracle"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in orcl_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsOrcl' is now in Column E (index 5)
          if worksheet.max_column >= 5:
              worksheet.cell(row=cell.row, column=5).value = "Yes"
          else:
               print(f"Warning: Column E (IsOrcl) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_orcl: {e}")
      traceback.print_exc()


## Match PostGres DBs
def match_pgres(destfile):
  """Matches PostGres DBs in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  pgres_str = ["pgres", "postgres"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in pgres_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsPGres' is now in Column F (index 6)
          if worksheet.max_column >= 6:
              worksheet.cell(row=cell.row, column=6).value = "Yes"
          else:
              print(f"Warning: Column F (IsPGres) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_pgres: {e}")
      traceback.print_exc()


## Match Possible DBs (General)
def match_gendb(destfile):
  """Matches Possible (generic) DBs in all sheets based on VM name in Column A.
     Sets 'Check' in SQL, Oracle, PGres columns if a generic term is found
     and the specific match wasn't already 'Yes'.

  Args:
    destfile: An xl workbook object.
  """
  gendb_str = ["db", "database"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      # Define columns for DB checks (assuming they are D, E, F)
      db_check_cols = [4, 5, 6] # D=IsSQL, E=IsOrcl, F=IsPGres

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in gendb_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Set 'Check' only if the specific DB columns are not already 'Yes'
          for col_idx in db_check_cols:
             if worksheet.max_column >= col_idx:
                 target_cell = worksheet.cell(row=cell.row, column=col_idx)
                 if target_cell.value != "Yes":
                     target_cell.value = "Check"
             else:
                  print(f"Warning: Column {col_idx} not found in sheet '{worksheet.title}' for row {cell.row} during gendb check.")
  except Exception as e:
      print(f"Error during match_gendb: {e}")
      traceback.print_exc()


## Match Exchange Servers
def match_exch(destfile):
  """Matches Exchange Servers in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  exch_str = ["exch", "exchange"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in exch_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsExch' is now in Column G (index 7)
          if worksheet.max_column >= 7:
              worksheet.cell(row=cell.row, column=7).value = "Yes"
          else:
              print(f"Warning: Column G (IsExch) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_exch: {e}")
      traceback.print_exc()


## Match TestDev
def match_tstdev(destfile):
  """Matches Test/Dev systems in all sheets based on VM name in Column A.

  Args:
    destfile: An xl workbook object.
  """
  tstdev_str = ["tst", "test", "dev"]
  try:
    for worksheet in destfile.worksheets:
      if worksheet.max_column < 1: continue

      for cell in worksheet["A"]:
        if cell.row == 1: continue
        if cell.value is None or not isinstance(cell.value, str): continue

        cell_value_lower = cell.value.lower()
        match_found = False
        for search_string in tstdev_str:
          if search_string in cell_value_lower:
            match_found = True
            break

        if match_found:
          # Assuming 'IsTestDev' is now in Column H (index 8)
          if worksheet.max_column >= 8:
              worksheet.cell(row=cell.row, column=8).value = "Yes"
          else:
              print(f"Warning: Column H (IsTestDev) not found in sheet '{worksheet.title}' for row {cell.row}")
  except Exception as e:
      print(f"Error during match_tstdev: {e}")
      traceback.print_exc()


## Set No Values
# Helper to get used range of Column A (more robustly)
def get_last_row_in_col_a(worksheet):
    """Finds the last row index in Column A that contains data."""
    last_row = 0
    # Iterate backwards from max_row is safer if there are gaps
    for row_idx in range(worksheet.max_row, 0, -1):
        cell = worksheet.cell(row=row_idx, column=1) # Column A
        if cell.value is not None and str(cell.value).strip() != "":
            last_row = row_idx
            break
    # If sheet seems empty based on max_row but has data, check dimensions?
    # For simplicity, we rely on max_row being somewhat accurate.
    # Or iterate from 1 up to max_row if needed.
    if last_row == 0 and worksheet.max_row > 0: # Fallback check if max_row is misleading
        for row_idx in range(1, worksheet.max_row + 1):
             cell = worksheet.cell(row=row_idx, column=1)
             if cell.value is not None and str(cell.value).strip() != "":
                 last_row = max(last_row, row_idx)
    return last_row

# Set No Values in columns C through H
def set_no_values(destfile):
    """Sets the value of all empty cells in columns C to H to "No",
       up to the last row containing data in column A for each sheet.

    Args:
      destfile: An xl workbook object.
    """
    try:
        # Iterate through the relevant sheets
        for ws_name in ['vInfo', 'vDisk', 'vPartition']:
            if ws_name in destfile.sheetnames:
                worksheet = destfile[ws_name]
                if worksheet.max_row <= 1: continue # Skip header-only or empty sheets

                last_row_a = get_last_row_in_col_a(worksheet)
                if last_row_a == 0: # If column A seems empty, maybe use worksheet.max_row cautiously
                    last_row_a = worksheet.max_row if worksheet.max_row > 1 else 1
                    print(f"Warning: Could not determine last row in Col A for sheet '{ws_name}'. Using worksheet max_row {last_row_a}.")


                # Iterate over the rows determined by column A's content.
                # Start from row 2 to skip header
                for row_index in range(2, last_row_a + 1):
                    # Iterate over the columns C to H (indices 3 to 8).
                    for column_index in range(3, 9):
                        # Check if column exists before accessing
                        if worksheet.max_column >= column_index:
                            cell = worksheet.cell(row=row_index, column=column_index)
                            # If the cell value is empty (None), set it to "No".
                            if cell.value is None or str(cell.value).strip() == "":
                                cell.value = "No"
                        # else: No need to warn for every cell, assume column exists if header was set
            else:
                 print(f"Warning: Sheet '{ws_name}' not found in workbook for set_no_values.")

    except Exception as e:
        print(f"Error during set_no_values: {e}")
        traceback.print_exc()

## Insert Columns for MiB to GB Math

# --- vInfo Sheet ---
def vinfo_find_col(vinfo_ws, header_name):
    """Finds the column index (1-based) for a given header name in the first row."""
    for col_idx in range(1, vinfo_ws.max_column + 1):
        if vinfo_ws.cell(row=1, column=col_idx).value == header_name:
            return col_idx
    return None

def vinfo_insert_gb_col(vinfo_ws, mib_header, gb_header):
    """Inserts a GB column after the MiB column and calculates values."""
    mib_col_idx = vinfo_find_col(vinfo_ws, mib_header)
    if mib_col_idx is None:
        print(f"Warning: Column '{mib_header}' not found in vInfo sheet.")
        return False # Indicate failure

    gb_col_idx = mib_col_idx + 1
    vinfo_ws.insert_cols(gb_col_idx)
    vinfo_ws.cell(row=1, column=gb_col_idx).value = gb_header

    for i in range(2, vinfo_ws.max_row + 1):
        mib_cell = vinfo_ws.cell(row=i, column=mib_col_idx)
        gb_cell = vinfo_ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = mib_cell.value
            if mib_value is not None and isinstance(mib_value, (int, float)):
                # ********** Use 953.7 conversion **********
                gb_value = round(mib_value / 953.7, 2)
                # ******************************************
                gb_cell.value = gb_value
                gb_cell.number_format = '0.00' # Apply number format
            # else: leave gb_cell blank if MiB is not a number
        except (ValueError, TypeError):
             # Handle potential errors if value is string that can't be converted
             gb_cell.value = None # Or set to 0 or an error string if preferred
    # Don't print progress here, let the main function yield
    # print(f"Processed '{gb_header}' column in vInfo.")
    return True # Indicate success

# --- vDisk Sheet ---
def vdisk_find_col(vdisk_ws, header_name):
    """Finds the column index (1-based) for a given header name in the first row."""
    for col_idx in range(1, vdisk_ws.max_column + 1):
        if vdisk_ws.cell(row=1, column=col_idx).value == header_name:
            return col_idx
    return None

def vdisk_insert_gb_col(vdisk_ws, mib_header, gb_header):
    """Inserts a GB column after the MiB column and calculates values."""
    mib_col_idx = vdisk_find_col(vdisk_ws, mib_header)
    if mib_col_idx is None:
        print(f"Warning: Column '{mib_header}' not found in vDisk sheet.")
        return False

    gb_col_idx = mib_col_idx + 1
    vdisk_ws.insert_cols(gb_col_idx)
    vdisk_ws.cell(row=1, column=gb_col_idx).value = gb_header

    for i in range(2, vdisk_ws.max_row + 1):
        mib_cell = vdisk_ws.cell(row=i, column=mib_col_idx)
        gb_cell = vdisk_ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = mib_cell.value
            if mib_value is not None and isinstance(mib_value, (int, float)):
                # ********** Use 953.7 conversion **********
                gb_value = round(mib_value / 953.7, 2)
                # ******************************************
                gb_cell.value = gb_value
                gb_cell.number_format = '0.00'
            # else: leave gb_cell blank
        except (ValueError, TypeError):
             gb_cell.value = None
    # Don't print progress here
    # print(f"Processed '{gb_header}' column in vDisk.")
    return True

# --- vPartition Sheet ---
def vpart_find_col(vpart_ws, header_name):
    """Finds the column index (1-based) for a given header name in the first row."""
    for col_idx in range(1, vpart_ws.max_column + 1):
        if vpart_ws.cell(row=1, column=col_idx).value == header_name:
            return col_idx
    return None

def vpart_insert_gb_col(vpart_ws, mib_header, gb_header):
    """Inserts a GB column after the MiB column and calculates values."""
    mib_col_idx = vpart_find_col(vpart_ws, mib_header)
    if mib_col_idx is None:
        print(f"Warning: Column '{mib_header}' not found in vPartition sheet.")
        return False

    gb_col_idx = mib_col_idx + 1
    vpart_ws.insert_cols(gb_col_idx)
    vpart_ws.cell(row=1, column=gb_col_idx).value = gb_header

    for i in range(2, vpart_ws.max_row + 1):
        mib_cell = vpart_ws.cell(row=i, column=mib_col_idx)
        gb_cell = vpart_ws.cell(row=i, column=gb_col_idx)
        try:
            mib_value = mib_cell.value
            if mib_value is not None and isinstance(mib_value, (int, float)):
                # ********** Use 953.7 conversion **********
                gb_value = round(mib_value / 953.7, 2)
                # ******************************************
                gb_cell.value = gb_value
                gb_cell.number_format = '0.00'
            # else: leave gb_cell blank
        except (ValueError, TypeError):
             gb_cell.value = None
    # Don't print progress here
    # print(f"Processed '{gb_header}' column in vPartition.")
    return True

# -- Disk Count Value --
def vdisk_diskcount_val(vdisk_ws):
    """Inserts a value of '1' in the 'DiskCount' Column on the vDisk worksheet.
       Assumes 'DiskCount' header is already present."""
    diskcount_col_idx = vdisk_find_col(vdisk_ws, "DiskCount")
    if diskcount_col_idx is None:
        print("Warning: 'DiskCount' column header not found in vDisk sheet.")
        return False

    # Set the value to 1 for all data rows
    for i in range(2, vdisk_ws.max_row + 1):
        vdisk_ws.cell(row=i, column=diskcount_col_idx).value = 1
    # Don't print progress here
    # print("Processed 'DiskCount' values in vDisk.")
    return True


## Compare VMs on vInfo Sheet to vPart Sheet
def compare_vms(vinfo_ws, vpart_ws):
  """Compares VMs in vInfo (Col A) to vPartition (Col A) and sets 'HasTools'
     (assumed Col I / index 9 in vInfo) to 'Yes' or 'No'.

  Args:
    vinfo_ws: An xl Worksheet object for the vInfo worksheet.
    vpart_ws: An xl Worksheet object for the vPartition worksheet.
  """
  if vinfo_ws is None or vpart_ws is None:
      print("Error: vInfo or vPartition worksheet not provided for compare_vms.")
      return

  # --- Pre-load vPartition VM names into a set for faster lookup ---
  vpart_vms = set()
  if vpart_ws.max_row > 1: # Check if there's data beyond the header
      # Assuming VM names are in the first column (A)
      for cell in vpart_ws['A'][1:]: # Skip header row A1
          if cell.value is not None and str(cell.value).strip() != "":
              vpart_vms.add(str(cell.value)) # Add VM name as string
  print(f"Found {len(vpart_vms)} unique VM names in vPartition for comparison.") # Server log

  # --- Iterate through vInfo and check against the set ---
  hastools_col_idx = 9 # Assuming 'HasTools' is Column I

  # Ensure the HasTools column exists (at least header)
  if vinfo_ws.max_column < hastools_col_idx:
      print(f"Warning: 'HasTools' column (expected index {hastools_col_idx}) not found in vInfo sheet.")
      return # Cannot proceed if column is missing

  # Check header just in case
  if vinfo_ws.cell(row=1, column=hastools_col_idx).value != "HasTools":
       print(f"Warning: Column {hastools_col_idx} in vInfo is not named 'HasTools'. Proceeding anyway.")


  if vinfo_ws.max_row > 1:
      # Iterate through rows in vInfo (Column A for VM name)
      for row_idx in range(2, vinfo_ws.max_row + 1):
          vinfo_vm_cell = vinfo_ws.cell(row=row_idx, column=1) # Col A
          hastools_cell = vinfo_ws.cell(row=row_idx, column=hastools_col_idx) # Col I

          if vinfo_vm_cell.value is not None and str(vinfo_vm_cell.value).strip() != "":
              vm_name = str(vinfo_vm_cell.value)
              # Check if the VM name exists in the set from vPartition
              if vm_name in vpart_vms:
                  hastools_cell.value = "Yes"
              else:
                  hastools_cell.value = "No"
          else:
              # Handle empty VM name cell in vInfo if necessary
              hastools_cell.value = "No" # Or None, or skip


## Delete Unnecessary Columns (Using Pandas)
def del_cols_vInfo(destfile_path):
    """Deletes unnecessary columns using Pandas for efficiency.
       Reads from and writes back to the specified Excel file path.
    """
    try:
        # Define columns to keep for each sheet
        keep_cols = {
            'vInfo': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'HasTools', 'Disks', 'Total disk capacity', 'Provisioned MiB', 'Provisioned GB', 'In Use MiB', 'In Use GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vDisk': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'DiskCount', 'Disk', 'Capacity MiB', 'Capacity GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools'],
            'vPartition': ['VM', 'Powerstate', 'IsFile', 'IsSQL', 'IsOrcl', 'IsPGres', 'IsExch', 'IsTestDev', 'Disk', 'Capacity MiB', 'Capacity GB', 'Consumed MiB', 'Consumed GB', 'Free MiB', 'Free GB', 'Datacenter', 'Cluster', 'OS according to the configuration file', 'OS according to the VMware Tools']
        }

        # Read all relevant sheets into a dictionary of DataFrames
        try:
             all_sheets_df = pd.read_excel(destfile_path, sheet_name=['vInfo', 'vDisk', 'vPartition'], engine='openpyxl')
        except ValueError as e:
             print(f"Warning reading sheets with pandas: {e}. Trying to read existing sheets individually.")
             all_sheets_df = {}
             try:
                 xls = pd.ExcelFile(destfile_path, engine='openpyxl')
                 for sheet_name in ['vInfo', 'vDisk', 'vPartition']:
                     if sheet_name in xls.sheet_names:
                         all_sheets_df[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                 xls.close() # Close the ExcelFile object
             except Exception as ex_read_err:
                 print(f"Error trying to read sheets individually: {ex_read_err}")
                 raise # Re-raise the error if individual read also fails

        if not all_sheets_df: # Check if any sheets were successfully read
             raise ValueError("No valid sheets (vInfo, vDisk, vPartition) found to process.")

        # Create a Pandas Excel writer object using openpyxl engine to overwrite
        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in all_sheets_df.items():
                if sheet_name in keep_cols:
                    # Find which columns to keep ACTUALLY exist in the DataFrame
                    cols_to_keep_in_df = [col for col in keep_cols[sheet_name] if col in df.columns]
                    if len(cols_to_keep_in_df) < len(keep_cols[sheet_name]):
                         missing_cols = set(keep_cols[sheet_name]) - set(cols_to_keep_in_df)
                         print(f"Warning: Columns missing in '{sheet_name}' for deletion step: {missing_cols}")

                    if not cols_to_keep_in_df: # If no columns to keep are found, skip writing sheet
                         print(f"Warning: No valid columns found to keep for sheet '{sheet_name}'. Skipping.")
                         continue

                    new_df = df[cols_to_keep_in_df]
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    # Don't print here, let main function yield
                    # print(f"Processed columns for sheet: {sheet_name}")

    except FileNotFoundError:
        print(f"Error: File not found for del_cols_vInfo: {destfile_path}")
        raise
    except Exception as e:
        print(f"Error during del_cols_vInfo (Pandas): {e}")
        traceback.print_exc()
        raise


## Truncate vPartition Storage Capacity Cols into vInfo (Using Pandas)
def trunc_cols_vPart(destfile_path):
    """Aggregates storage data from vPartition and merges it into a new vSummary sheet.
       Reads from and writes back to the specified Excel file path.
    """
    try:
        # Read necessary sheets
        sheets_to_read = ['vInfo', 'vPartition']
        dfs = {}
        try:
            xls = pd.ExcelFile(destfile_path, engine='openpyxl')
            for sheet_name in sheets_to_read:
                 if sheet_name in xls.sheet_names:
                     dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                 else:
                     raise ValueError(f"Required sheet '{sheet_name}' not found in '{destfile_path}' for truncation.")
            # Also read vDisk if it exists, to preserve it
            if 'vDisk' in xls.sheet_names:
                 dfs['vDisk'] = pd.read_excel(xls, sheet_name='vDisk')
            xls.close() # Close the ExcelFile object

        except FileNotFoundError:
             print(f"Error: File not found for trunc_cols_vPart: {destfile_path}")
             raise
        except ValueError as e:
             print(f"Error: {e}")
             raise

        tcdf1 = dfs['vInfo']
        tcdf3 = dfs['vPartition']

        # Identify storage-related columns in vPartition
        storage_columns = ['Capacity GB', 'Consumed GB', 'Free GB']
        existing_storage_columns = [col for col in storage_columns if col in tcdf3.columns]
        if not existing_storage_columns:
             print("Warning: No storage columns found in vPartition. Skipping aggregation.")
             vSummary_df = tcdf1.copy() # Create vSummary based on vInfo only

        else:
            # print(f"Aggregating vPartition columns: {existing_storage_columns}") # Server log
            if 'VM' not in tcdf3.columns:
                 raise ValueError("'VM' column not found in vPartition sheet, cannot aggregate.")

            # Convert non-numeric to 0 before sum
            for col in existing_storage_columns:
                 tcdf3[col] = pd.to_numeric(tcdf3[col], errors='coerce').fillna(0)

            aggregated_df = (
                tcdf3.groupby('VM', as_index=False)
                [existing_storage_columns].sum()
            )

            if 'VM' not in tcdf1.columns:
                 raise ValueError("'VM' column not found in vInfo sheet, cannot merge.")

            vSummary_df = pd.merge(tcdf1, aggregated_df, on='VM', how='left')

            # Fill NaN for newly merged columns only
            for col in existing_storage_columns:
                 if col in vSummary_df.columns:
                      vSummary_df[col].fillna(0, inplace=True)

            # --- Reorder columns (Optional) ---
            if 'In Use GB' in vSummary_df.columns:
                try:
                    in_use_gb_index = vSummary_df.columns.get_loc('In Use GB')
                    cols_before = vSummary_df.columns[:in_use_gb_index + 1].tolist()
                    cols_after = [col for col in vSummary_df.columns[in_use_gb_index + 1:] if col not in existing_storage_columns]
                    new_column_order = cols_before + existing_storage_columns + cols_after
                    vSummary_df = vSummary_df[new_column_order]
                except KeyError:
                     print("Warning: 'In Use GB' column not found for reordering vSummary.")
            else:
                 print("Warning: 'In Use GB' column not found in vSummary base data.")

        # --- Save the results ---
        with pd.ExcelWriter(destfile_path, engine='openpyxl', mode='w') as writer:
            vSummary_df.to_excel(writer, sheet_name='vSummary', index=False)
            # print("Written vSummary sheet.") # Server log

            # Write back other preserved sheets
            if 'vInfo' in dfs: dfs['vInfo'].to_excel(writer, sheet_name='vInfo', index=False)
            if 'vDisk' in dfs: dfs['vDisk'].to_excel(writer, sheet_name='vDisk', index=False)
            if 'vPartition' in dfs: dfs['vPartition'].to_excel(writer, sheet_name='vPartition', index=False)

    except FileNotFoundError:
         print(f"Error: File not found for trunc_cols_vPart: {destfile_path}")
         raise
    except ValueError as e:
         print(f"Error during trunc_cols_vPart (Pandas Preprocessing): {e}")
         raise
    except Exception as e:
        print(f"Error during trunc_cols_vPart (Pandas): {e}")
        traceback.print_exc()
        raise


## Remove Excess vSummary Cols (Using openpyxl AFTER Pandas)
def trimvSum1(destfile):
    """Removes specific columns by name from the vSummary sheet using openpyxl."""
    if 'vSummary' not in destfile.sheetnames:
        print("Warning: vSummary sheet not found for trimming.")
        return

    vSum_ws = destfile['vSummary']
    cols_to_delete_by_name = ['Provisioned MiB', 'In Use MiB'] # Define columns to remove

    try:
        headers = [cell.value for cell in vSum_ws[1]]
        indices_to_delete = []
        for col_name in cols_to_delete_by_name:
            try:
                # Find index (1-based), add to list if found
                col_idx = headers.index(col_name) + 1
                if col_idx <= vSum_ws.max_column:
                    indices_to_delete.append(col_idx)
                else:
                     print(f"Warning: Found '{col_name}' but index {col_idx} > max_column {vSum_ws.max_column}?")
            except ValueError:
                print(f"Warning: Column '{col_name}' not found for deletion in trimvSum1.")

        # Sort indices descending for safe deletion
        indices_to_delete.sort(reverse=True)

        if not indices_to_delete:
             print("Info: No columns found to delete in trimvSum1 based on names.")
             return

        # print(f"Attempting to delete columns at indices: {indices_to_delete}") # Server log
        for col_idx in indices_to_delete:
             vSum_ws.delete_cols(col_idx, 1)
             # print(f"Deleted column at index {col_idx}") # Server log

    except Exception as e:
        print(f"Error during trimvSum1 (openpyxl column deletion): {e}")
        traceback.print_exc()


## Consolidate Consumed GB Column (Using openpyxl AFTER Pandas)
def consol_vSum(destfile):
    """Copies 'In Use GB' to 'Consumed GB' if 'Consumed GB' is empty/blank on vSummary sheet."""
    if 'vSummary' not in destfile.sheetnames:
        print("Warning: vSummary sheet not found for consolidation.")
        return

    vSum_ws = destfile['vSummary']
    try:
        consumed_gb_col_idx = None
        in_use_gb_col_idx = None
        headers = [cell.value for cell in vSum_ws[1]]
        for idx, header in enumerate(headers, 1):
            if header == 'Consumed GB': consumed_gb_col_idx = idx
            elif header == 'In Use GB': in_use_gb_col_idx = idx

        if consumed_gb_col_idx is None or in_use_gb_col_idx is None:
            missing = [("'" + h + "'") for h, i in [('Consumed GB', consumed_gb_col_idx), ('In Use GB', in_use_gb_col_idx)] if i is None]
            print(f"Error: {', '.join(missing)} column(s) not found in vSummary for consolidation.")
            return

        for row_idx in range(2, vSum_ws.max_row + 1):
            consumed_gb_cell = vSum_ws.cell(row=row_idx, column=consumed_gb_col_idx)
            in_use_gb_cell = vSum_ws.cell(row=row_idx, column=in_use_gb_col_idx)

            consumed_value = consumed_gb_cell.value
            # Check if None or empty string after stripping whitespace
            if consumed_value is None or (isinstance(consumed_value, str) and consumed_value.strip() == ""):
                 consumed_gb_cell.value = in_use_gb_cell.value
                 # Copy number format if source cell has one
                 if in_use_gb_cell.has_style and in_use_gb_cell.number_format:
                      consumed_gb_cell.number_format = in_use_gb_cell.number_format

    except Exception as e:
        print(f"Error during consol_vSum (openpyxl): {e}")
        traceback.print_exc()


## Filter First Row
def filter_rows(destfile):
  """Applies auto-filter to the header row on all sheets."""
  try:
    for ws in destfile.worksheets:
      if ws.max_row > 0 and ws.max_column > 0:
          # Use utility to get column letter
          last_col_letter = xl.utils.get_column_letter(ws.max_column)
          filter_range = f"A1:{last_col_letter}1"
          ws.auto_filter.ref = filter_range
  except Exception as e:
      print(f"Error applying filters: {e}")
      traceback.print_exc()


# ================================================================================================================================================================================================================================================
# ===================================================================================================== MAIN PROCESSING FUNCTION (SSE GENERATOR) =================================================================================================
# ================================================================================================================================================================================================================================================

def process_rvtools_file(input_filepath, output_folder, original_basename):
    """
    Processes the RVTools Excel file, yielding SSE-formatted progress updates.

    Args:
        input_filepath (str): The full path to the *saved* RVTools Excel file.
        output_folder (str): The folder where the processed file should be saved.
        original_basename (str): The original filename as uploaded by the user.

    Yields:
        str: Server-Sent Event formatted strings (e.g., "data: Processing step...\n\n").
             Includes a final 'result' event with JSON payload indicating outcome.
    """
    # Helper to format SSE messages for the client
    def sse_message(data):
        data = str(data).replace('\n', ' ') # Replace newlines in message
        return f"data: {data}\n\n"

    # Helper to yield a final result event to the client
    def yield_result(success, message):
        payload = json.dumps({"success": success, "message": message})
        yield f"event: result\ndata: {payload}\n\n"

    # --- Start Processing ---
    yield sse_message(f"Starting analysis for: {original_basename}")
    time.sleep(0.1)

    destfile = None # Initialize for finally block
    output_filename = None # Initialize for finally block and result payload
    try:
        # --- Basic File Setup ---
        if not os.path.exists(input_filepath):
             print(f"ERROR: Input file not found: {input_filepath}") # Server log
             yield from yield_result(False, "Server Error: Input file missing.")
             return

        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
                yield sse_message(f"Created output folder: {output_folder}")
                print(f"Created output folder: {output_folder}")
            except OSError as e:
                print(f"ERROR: Failed to create output folder '{output_folder}': {e}")
                yield from yield_result(False, f"Server Error: Cannot create folder.")
                return

        # --- Construct Output Filename (with -ANALYZED suffix) ---
        # !!! --- THIS IS THE SECTION TO MODIFY --- !!!
        if original_basename:
            base_name_no_ext, ext = os.path.splitext(original_basename)
            # CHANGE SUFFIX HERE:
            output_filename = f"{base_name_no_ext}-ANALYZED{ext}"
        else:
            # Fallback logic (shouldn't be needed if app.py always provides original_basename)
            temp_base, ext = os.path.splitext(os.path.basename(input_filepath))
             # CHANGE SUFFIX HERE:
            output_filename = f"{temp_base}-ORIGINAL_NAME_MISSING-ANALYZED{ext}"
            yield sse_message("Warning: Original filename missing; using temporary name.")
            print("Warning: Original filename not provided; using temp name as base.")

        # !!! --- END OF SECTION TO MODIFY --- !!!
        output_filepath = os.path.join(output_folder, output_filename)
        yield sse_message(f"Output file will be: {output_filename}")
        print(f"Output file will be: {output_filepath} (Derived from original: '{original_basename}')")
        time.sleep(0.1)

        # --- Core Processing Logic (using helper functions) ---

        # Step 1: Load source workbook & save copy (or load directly if helpers modify in place)
        # Option A: Load fresh copy to work on (safer if helpers modify heavily)
        yield sse_message("Step 1: Loading workbook for processing...")
        time.sleep(0.1)
        destfile = xl.load_workbook(input_filepath) # Load the original input
        # Option B: If helpers modify a file, copy it first like original logic
        # srcfile = xl.load_workbook(input_filepath)
        # srcfile.save(output_filepath) # Save copy with FINAL name early?
        # srcfile.close()
        # destfile = xl.load_workbook(output_filepath) # Work on the copy

        # For now, assume helpers work on the loaded 'destfile' object in memory
        yield sse_message(" -> Workbook loaded.")
        time.sleep(0.1)

        # Step 2: Remove extra sheets (if needed - depends if input has them)
        yield sse_message("Step 2: Checking sheets...")
        time.sleep(0.1)
        keep_sheets = ['vInfo', 'vDisk', 'vPartition']
        sheets_to_delete = [s for s in destfile.sheetnames if s not in keep_sheets]
        if sheets_to_delete:
             yield sse_message(f" -> Removing unnecessary sheets: {', '.join(sheets_to_delete)}...")
             if not any(s in destfile.sheetnames for s in keep_sheets):
                  print("ERROR: Required sheets missing.")
                  if destfile: destfile.close()
                  yield from yield_result(False, "Error: Required sheets (vInfo, vDisk, vPartition) not found.")
                  return
             for sheetName in sheets_to_delete: del destfile[sheetName]
             yield sse_message(f" -> Kept sheets: {destfile.sheetnames}")
        else:
             yield sse_message(" -> All required sheets present.")
        time.sleep(0.1)


        # Step 3: Remove Formatting
        yield sse_message("Step 3: Removing formatting...")
        time.sleep(0.1)
        removeFormatting(destfile)
        yield sse_message(" -> Formatting removed.")
        time.sleep(0.1)

        # Step 4: Add and Rename Columns
        yield sse_message("Step 4: Adding category columns...")
        time.sleep(0.1)
        target_sheets = [s for s in keep_sheets if s in destfile.sheetnames]
        for sheet_name in target_sheets:
            worksheet = destfile[sheet_name]
            worksheet.insert_cols(3, 6) # Insert cols C-H
            worksheet['C1']='IsFile'; worksheet['D1']='IsSQL'; worksheet['E1']='IsOrcl'
            worksheet['F1']='IsPGres'; worksheet['G1']='IsExch'; worksheet['H1']='IsTestDev'
            if sheet_name == 'vInfo': worksheet.insert_cols(9, 1); worksheet['I1'] = 'HasTools'
            elif sheet_name == 'vDisk': worksheet.insert_cols(9, 1); worksheet['I1'] = 'DiskCount'
        yield sse_message(" -> Columns added.")
        time.sleep(0.1)

        # Step 5: Match Workload Types & Set Defaults
        yield sse_message("Step 5: Identifying workloads...")
        time.sleep(0.1)
        match_fs(destfile); match_sql(destfile); match_orcl(destfile); match_pgres(destfile)
        match_exch(destfile); match_tstdev(destfile); match_gendb(destfile)
        yield sse_message(" -> Types matched.")
        yield sse_message("Step 5b: Setting default values...")
        time.sleep(0.1)
        set_no_values(destfile)
        yield sse_message(" -> Defaults set.")
        time.sleep(0.1)

        # Step 6: Compare vInfo/vPart for 'HasTools'
        yield sse_message("Step 6: Checking VMware Tools status...")
        time.sleep(0.1)
        vinfo_ws = destfile['vInfo'] if 'vInfo' in destfile.sheetnames else None
        vpart_ws = destfile['vPartition'] if 'vPartition' in destfile.sheetnames else None

        # The subsequent check 'if vinfo_ws and vpart_ws:' will then correctly
        # handle cases where one or both sheets might be missing (set to None).
        # Optional: Add a more specific message if sheets are missing
        if not (vinfo_ws and vpart_ws):
            missing = []
            if not vinfo_ws: missing.append("'vInfo'")
            if not vpart_ws: missing.append("'vPartition'")
            # Yield message indicating which specific sheet(s) caused the skip
            yield sse_message(f" -> Skipped Tools check ({', '.join(missing)} sheet(s) missing).")
            # The original 'else' part yielding this message can be removed or kept as a fallback.
        if vinfo_ws and vpart_ws:
             compare_vms(vinfo_ws, vpart_ws)
             yield sse_message(" -> Tools status checked.")
        #else: yield sse_message(" -> Skipped Tools check (sheet missing).")
        time.sleep(0.1)

        # Step 7: Fill DiskCount in vDisk
        yield sse_message("Step 7: Setting DiskCount...")
        time.sleep(0.1)
        if 'vDisk' in destfile:
            vdisk_ws = destfile['vDisk']
            if not vdisk_diskcount_val(vdisk_ws): yield sse_message(" -> Warning: Failed to set DiskCount.")
            else: yield sse_message(" -> DiskCount set.")
        else: yield sse_message(" -> Skipped DiskCount (sheet missing).")
        time.sleep(0.1)

        # Step 8: Insert GB Columns and Calculate
        yield sse_message("Step 8: Calculating GB values...")
        time.sleep(0.1)
        if 'vInfo' in destfile:
            vinfo_ws = destfile['vInfo']
            vinfo_insert_gb_col(vinfo_ws, "Provisioned MiB", "Provisioned GB")
            vinfo_insert_gb_col(vinfo_ws, "In Use MiB", "In Use GB")
        if 'vDisk' in destfile:
            vdisk_ws = destfile['vDisk']
            vdisk_insert_gb_col(vdisk_ws, "Capacity MiB", "Capacity GB")
        if 'vPartition' in destfile:
            vpart_ws = destfile['vPartition']
            vpart_insert_gb_col(vpart_ws, "Capacity MiB", "Capacity GB")
            vpart_insert_gb_col(vpart_ws, "Consumed MiB", "Consumed GB")
            vpart_insert_gb_col(vpart_ws, "Free MiB", "Free GB")
        yield sse_message(" -> GB values calculated.")
        time.sleep(0.1)

        # --- Intermediate Save before Pandas ---
        # Save the openpyxl changes to the final output file path before pandas reads it
        yield sse_message("Saving intermediate results before summary...")
        time.sleep(0.1)
        destfile.save(output_filepath)
        destfile.close()
        destfile = None # Clear variable
        yield sse_message(" -> Intermediate save complete.")
        time.sleep(0.1)

        # Step 9: Delete unnecessary columns (Pandas)
        yield sse_message("Step 9: Removing unused columns...")
        time.sleep(0.1)
        try:
            del_cols_vInfo(output_filepath) # Reads/writes the file at output_filepath
            yield sse_message(" -> Unused columns removed.")
        except Exception as pd_del_err:
            print(f"ERROR during Pandas column deletion: {pd_del_err}")
            yield from yield_result(False, f"Error removing columns: {pd_del_err}")
            return
        time.sleep(0.1)

        # Step 10: Aggregate vPartition (Pandas)
        yield sse_message("Step 10: Creating summary sheet...")
        time.sleep(0.1)
        try:
            trunc_cols_vPart(output_filepath) # Reads/adds sheet/writes file
            yield sse_message(" -> Summary sheet created.")
        except Exception as pd_trunc_err:
            print(f"ERROR during Pandas summary creation: {pd_trunc_err}")
            yield from yield_result(False, f"Error creating summary: {pd_trunc_err}")
            return
        time.sleep(0.1)

        # --- Reload with openpyxl for final cleanup ---
        yield sse_message("Reloading workbook for final cleanup...")
        time.sleep(0.1)
        destfile = xl.load_workbook(output_filepath)

        # Step 11: Final Formatting, Trim, Consolidate
        yield sse_message("Step 11: Final cleanup...")
        time.sleep(0.1)
        removeFormatting(destfile)
        if 'vSummary' in destfile:
            trimvSum1(destfile); consol_vSum(destfile)
            yield sse_message(" -> Summary sheet cleaned.")
        else:
            yield sse_message(" -> Skipped summary cleanup (sheet missing).")
            print("Warning: vSummary sheet not found for final cleanup.")
        time.sleep(0.1)

        # Step 12: Apply Filters
        yield sse_message("Step 12: Applying filters...")
        time.sleep(0.1)
        filter_rows(destfile)
        yield sse_message(" -> Filters applied.")
        time.sleep(0.1)

        # Final save (of the openpyxl object)
        yield sse_message("Saving final workbook...")
        time.sleep(0.1)
        destfile.save(output_filepath)
        yield sse_message(" -> Final save complete.")
        if destfile: destfile.close(); destfile = None
        time.sleep(0.5)

        # --- Success ---
        yield sse_message("--- Analysis Completed Successfully ---")
        print(f"--- Analysis Completed: {original_basename} -> {output_filename} ---")
        yield from yield_result(True, output_filename) # Yield success with BASE filename

    except Exception as e:
        if destfile:
            try: destfile.close()
            except Exception as close_err: print(f"Warning: Error closing workbook: {close_err}")

        print(f"\n--- ERROR processing {original_basename} ---"); traceback.print_exc()
        error_message = f"An error occurred: {type(e).__name__}"
        yield sse_message(f"ERROR: {error_message} - See server logs.")
        time.sleep(0.1)
        yield from yield_result(False, error_message)

# ===============================================================================================================================================================================================================================================
# ================================================================================================================ END OF SCRIPT ================================================================================================================
# ===============================================================================================================================================================================================================================================